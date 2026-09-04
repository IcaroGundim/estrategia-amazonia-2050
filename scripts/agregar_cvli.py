#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Agrega CVLI (homicídio doloso + latrocínio + lesão corporal seguida de morte)
por UF e ano a partir das bases Sinesp/VDE do MJ (bancovde-YYYY.xlsx)."""
import csv, os
from collections import defaultdict
from openpyxl import load_workbook

DADOS = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "dados")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
CVLI_EVENTOS = {"Homicídio doloso", "Roubo seguido de morte (latrocínio)", "Lesão corporal seguida de morte"}

def num(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return 0

agreg = defaultdict(lambda: defaultdict(int))  # uf -> ano -> total
info = {}
for ano in range(2020, 2027):
    path = os.path.join(DADOS, "sinesp", f"bancovde-{ano}.xlsx")
    if not os.path.exists(path):
        continue
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(min_row=2, values_only=True)
    datas = set()
    muns = set()
    n = 0
    for r in rows:
        uf, mun, evento, data = r[0], r[1], r[2], r[3]
        if uf is None:
            continue
        n += 1
        uf = str(uf).strip().upper()
        a = str(data)[:4]
        datas.add(a)
        muns.add(str(mun).strip().upper())
        if evento and evento.strip() in CVLI_EVENTOS:
            total = num(r[7]) + num(r[8]) + num(r[9])
            agreg[uf][a] += total
    info[ano] = (n, sorted(datas), len(muns))
    wb.close()

for ano, (n, datas, m) in sorted(info.items()):
    print(f"base {ano}: {n} linhas | anos cobertos: {datas} | municípios: {m}")

print()
print("CVLI (nº de vítimas/ocorrências) por UF/ano:")
anos_todos = sorted({a for uf in agreg for a in agreg[uf]})
print("UF | " + " | ".join(anos_todos))
for uf in sorted(agreg):
    print(uf + " | " + " | ".join(str(agreg[uf].get(a, "")) for a in anos_todos))

# salvar CSV
with open(os.path.join(DADOS, "sinesp", "cvli_uf_ano.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["uf", "ano", "cvli"])
    for uf in sorted(agreg):
        for a in anos_todos:
            if a in agreg[uf]:
                w.writerow([uf, a, agreg[uf][a]])
print("\nCSV salvo: dados/cvli_uf_ano.csv")
