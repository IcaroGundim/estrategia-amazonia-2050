#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Monta 'Indicadores_Resultado_Eixo4_Amazonia2050.xlsx' — Eixo 4: Infraestrutura
e integração regional sustentável (Estratégia Amazônia 2050), 9 estados da AL."""
import csv, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
DADOS = os.path.join(PASTA, "dados")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
NOMES_UF = {"AC": "Acre", "AP": "Amapá", "AM": "Amazonas", "MA": "Maranhão",
            "MT": "Mato Grosso", "PA": "Pará", "RO": "Rondônia", "RR": "Roraima", "TO": "Tocantins"}

# ---------- IBC (ANATEL) ----------
ibc = defaultdict(dict)  # uf -> ano -> (ibc_uf, ibc_ponderado)
with open(os.path.join(DADOS, "anatel", "ibc_uf_ano.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        ibc[r["uf"]][int(r["ano"])] = (r["ibc_uf"], r["ibc_ponderado_pop"])
al_ibc = {}
for ano in (2021, 2022, 2023, 2024, 2025):
    al_ibc[ano] = ibc[UFS[0]].get(ano, ("", ""))[1] if ano in ibc[UFS[0]] else ""
# AL ponderado (recalculado a partir das UFs não funciona — média ponderada exige população municipal;
# usa o valor calculado no script de coleta, gravado na linha 'AL' abaixo se existir)
al_ibc_file = os.path.join(DADOS, "anatel", "ibc_al_ano.csv")
if os.path.exists(al_ibc_file):
    with open(al_ibc_file, encoding="utf-8") as f:
        for r in csv.DictReader(f):
            al_ibc[int(r["ano"])] = r["ibc_al"]

# ---------- PER (ANEEL SIGA) ----------
per = {}
with open(os.path.join(DADOS, "aneel", "per_uf.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        per[r["uf"]] = r
pot_fonte = defaultdict(dict)  # uf -> origem -> MW
with open(os.path.join(DADOS, "aneel", "potencia_uf_fonte.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        pot_fonte[r["uf"]][r["origem"]] = float(r["potencia_fiscalizada_kw"]) / 1000

# ---------- ISGR (saneamento) ----------
isgr = {}
with open(os.path.join(DADOS, "saneamento", "isgr_uf.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        isgr[r["uf"]] = r

# ---------- Catálogo Eixo 4 ----------
CATALOGO = [
    ("I4.1.1", "4.1", "Conectividade digital (IBC-AMZ)",
     "Elevar o IBC médio da AL de 53,52 pontos (2025) para 80,00 pontos até 2050",
     "Índice de Conectividade da Amazônia Legal (IBC-AMZ) ponderado pela população",
     "0 a 100", "ANATEL (painel Meu Município)", 2050,
     "coletado (IBC estadual e ponderado pela população municipal, 2021-2025; AL 2025 = 53,52 — confere com o baseline da ficha)"),
    ("I4.2.1", "4.2", "Adequação e trafegabilidade de transportes",
     "Elevar a adequação e trafegabilidade da infraestrutura estadual de 46,5% para 65% até 2050",
     "Taxa de adequação e trafegabilidade da infraestrutura de transportes (%)",
     "%", "Min. Transportes; DNIT; CNT", 2050,
     "pendente (painel CNT em Power BI sem API; DNIT vgeo só publica geometria, sem estado de conservação)"),
    ("I4.3.1", "4.3", "Transição energética (ITEQ)",
     "Aumentar 12,37% o acesso estruturado a energia limpa (baseline 60,13% em 2025)",
     "Indicador de Transição Energética e Qualidade Distributiva Amazônica (%)",
     "0 a 100%", "ANEEL; EPE (PASI)", 2035,
     "parcial (componentes coletáveis: renovabilidade da geração AL via SIGA; POP_SIN/POP_isolado e DEC/FEC exigem PASI/ANEEL sem API aberta)"),
    ("I4.3.2", "4.3", "Participação de renováveis (PER)",
     "Atingir 80% de renováveis na oferta de energia até 2050 (baseline 65,24% em 2025)",
     "Participação de Energias Renováveis e Sustentáveis na oferta de energia (%)",
     "%", "ANEEL (SIGA); EPE", 2035,
     "coletado (proxy: % renovável da potência fiscalizada de usinas em operação — SIGA/CKAN, base 25/08/2026)"),
    ("I4.4.1", "4.4", "Saneamento básico e gestão de riscos (ISGR)",
     "Atingir 80% no ISGR até 2050 (baseline 41,52% em 2022/2024)",
     "Índice de Saneamento Básico e Gestão de Riscos",
     "0 a 100", "IBGE (Censo 2022; MUNIC 2024)", 2050,
     "coletado (proxy: Censo 2022 + fatores FClima/FGov da MUNIC 2024; FHidro não publicado => 1)"),
    ("I4.4.2", "4.4", "Capacidade adaptativa urbana",
     "80% dos municípios com capacidade adaptativa média acima de 0,5",
     "% de municípios com média de capacidade adaptativa acima de 0.5",
     "% de municípios", "AdaptaBrasil MCTI (Painel Cidades)", 2035,
     "pendente (Painel Cidades é SPA sem API; API do AdaptaBrasil 403)"),
]

# ---------- Estilo ----------
HDR_FILL = PatternFill("solid", fgColor="1B4332")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13, color="1B4332")
NOTE_FONT = Font(italic=True, size=9, color="555555")
BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")


def estilo_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def largura(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


wb = Workbook()

# ---------- Sobre ----------
ws = wb.active
ws.title = "Sobre"
ws["A1"] = "ESTRATÉGIA REGIONAL AMAZÔNIA 2050 — Eixo 4: Infraestrutura e integração regional sustentável"
ws["A1"].font = TITLE_FONT
ws["A3"] = "Indicadores de RESULTADO (somente resultado), organizados para os 9 estados da Amazônia Legal."
ws["A4"] = "Base: '3. Matriz Metas x Indicadores .xlsx' (aba atualizada, coluna INDICADOR DE RESULTADO) e fichas técnicas."
ws["A6"] = "Data de coleta: 30/08/2026."
ws["A7"] = "Status: coletado = dado oficial obtido; parcial = parte do indicador; pendente = fonte sem API aberta ou indisponível no dia."
ws["A9"] = "FONTES E LIMITAÇÕES"
ws["A9"].font = Font(bold=True)
ws["A10"] = "• IBC-AMZ (I4.1.1): dados brutos oficiais do painel Meu Município/ANATEL (ibc.zip, séries 2021-2025). IBC-AMZ ponderado pela população municipal calculado com a população do Censo 2022 (SIDRA 4709). Validação: AL 2025 = 53,52 pontos, idêntico ao baseline da ficha técnica."
ws["A11"] = "• Transportes (I4.2.1): painel da Pesquisa CNT de Rodovias está em Power BI (sem API/dados abertos); DNIT vgeo publica só geometria das rodovias, sem estado de conservação/trafegabilidade — o índice ponderado por qualidade (pesos 1/0,8/0,5/0,2/0,05 da ficha) não é replicável sem os dados da pesquisa."
ws["A12"] = "• ITEQ (I4.3.1): componente renovável coletável via SIGA; os demais (POP_SIN/POP_isolado por município, DEC/FEC por município, fatores de qualidade distributiva) dependem do PASI/EPE (endpoints de exportação protegidos por autenticação) e do portal de relatórios da ANEEL (sem API documentada)."
ws["A13"] = "• PER (I4.3.2): SIGA (dadosabertos.aneel.gov.br, CKAN), empreendimentos EM OPERAÇÃO na AL, potência fiscalizada (kW). Renováveis da ficha: hídrica, solar, eólica, biomassa (biogás sem registro na AL). Proxy sobre POTÊNCIA, não sobre energia ofertada — o PER oficial pondera a oferta de energia."
ws["A14"] = "• ISGR (I4.4.1): Censo 2022 (SIDRA 6803/6805) nos 808 municípios da AL + MUNIC 2024 (FTP IBGE, base 07/11/2025) para FClima (Magr18/Mhab088/Mtic266) e FGov (Mgov086), conforme fórmula da ficha. O fator FHidro (vulnerabilidade à escassez hídrica) não é publicado — assumido 1; por isso o resultado AL (47,17%) fica acima do baseline oficial (41,52%)."
ws["A15"] = "• Ressalvas MUNIC/ISGR: (a) Mhab088 ('-' = não aplicável) só tem resposta em 260 dos 808 municípios — onde não se aplica vale a penalidade 0,9, conforme a regra da ficha ('NÃO ou em branco'); (b) 'água adequada' da ficha inclui a categoria 'tem ligação à rede geral mas usa principalmente outra forma' (SIDRA 72145 — poço raso, carro-pipa, chuva, rios). Definição estrita (só rede utilizada + poço profundo) daria 79,1% na AL em vez de 86,2%."
ws["A16"] = "• Capacidade adaptativa (I4.4.2): Painel Cidades do AdaptaBrasil é aplicação SPA sem API acessível; a API geral do AdaptaBrasil responde 403. Coleta alternativa: navegação manual no painel ou solicitação ao MCTI."
ws["A17"] = "• Revisão independente (scripts/revisar_eixo4.py): população municipal conferida contra SIDRA n3 (9/9 UFs); agregação municipal água/esgoto = SIDRA n3 (0 divergências); esgoto adequado Brasil 77,4% coerente com a divulgação do Censo 2022; spot-checks municipais (Porto Velho, Boa Vista) exatos; SIGA: Belo Monte 11,2 GW + Tucuruí 8,5 GW no topo do PA e total nacional 219 GW; re-ponderação IBC reproduz o baseline 53,52 com implementação independente."
ws["A19"] = "Scripts: scripts/eixo4_ibc_anatel.py, eixo4_aneel_siga.py, eixo4_saneamento_sidra.py. Dados brutos: dados/anatel/, dados/aneel/, dados/saneamento/, dados/ibge_munic/."
ws["A19"].font = NOTE_FONT
largura(ws, [130])

# ---------- Catálogo ----------
ws = wb.create_sheet("Catalogo_Indicadores")
headers = ["Código", "Linha de Ação", "Indicador (nome curto)", "Meta", "Indicador de RESULTADO (texto da matriz)",
           "Unidade", "Fonte", "Prazo", "Status da coleta"]
ws.append(headers)
for row in CATALOGO:
    ws.append(list(row))
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = WRAP
largura(ws, [8, 9, 32, 40, 55, 12, 30, 9, 32])

# ---------- Matriz_UF ----------
ws = wb.create_sheet("Matriz_UF")
headers = ["Código", "Indicador de RESULTADO (resumo)", "Unidade", "Ano ref.", "AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO", "Status", "Fonte"]
ws.append(headers)
for cod, la, curto, meta, indicador, unid, fonte, prazo, status in CATALOGO:
    if cod == "I4.1.1":
        vals = [float(ibc[uf][2025][1]) if uf in ibc and 2025 in ibc[uf] else "—" for uf in UFS]
        anoref, status_, unid2 = "2025", "coletado (IBC ponderado pela população)", "IBC-AMZ 0-100"
    elif cod == "I4.3.2":
        vals = [round(float(per[uf]["per_renovaveis_pct"]), 1) if uf in per else "—" for uf in UFS]
        anoref, status_, unid2 = "ago/2026 (SIGA)", "coletado (proxy sobre potência fiscalizada)", "% renovável"
    elif cod == "I4.4.1":
        vals = [round(float(isgr[uf]["isgr_pct"]), 1) if uf in isgr else "—" for uf in UFS]
        anoref, status_, unid2 = "2022 (Censo) + MUNIC 2024", "coletado (proxy — FHidro=1)", "ISGR 0-100"
    else:
        vals = ["—"] * 9
        anoref, status_, unid2 = "—", status, unid
    ws.append([cod, curto, unid2, anoref] + vals + [status_, fonte])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = WRAP
largura(ws, [9, 44, 16, 20, 8, 8, 8, 8, 8, 8, 8, 8, 8, 26, 24])
ws.freeze_panes = "E2"

# ---------- 4.1.1 IBC-AMZ ----------
ws = wb.create_sheet("4.1.1_IBC_AMZ_Anatel")
headers = ["UF", "Estado", "IBC 2021", "IBC 2022", "IBC 2023", "IBC 2024", "IBC 2025",
           "IBC ponderado 2024", "IBC ponderado 2025"]
ws.append(headers)
for uf in UFS:
    row = [uf, NOMES_UF[uf]]
    for a in (2021, 2022, 2023, 2024, 2025):
        v = ibc.get(uf, {}).get(a)
        row.append(float(v[0]) if v and v[0] != "" else "—")
    for a in (2024, 2025):
        v = ibc.get(uf, {}).get(a)
        row.append(float(v[1]) if v and v[1] != "" else "—")
    ws.append(row)
row = ["AL", "Amazônia Legal"] + ["—"] * 5
for a in (2024, 2025):
    row.append(float(al_ibc.get(a)) if al_ibc.get(a) not in (None, "") else "—")
ws.append(row)
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 10, 10, 10, 10, 10, 17, 17])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: ANATEL, dados brutos do painel Meu Município — Índice Brasileiro de Conectividade (ibc.zip, dadosabertos ANATEL). 'IBC ponderado' = Σ(IBC_mun × pop_mun)/Σ pop_mun com população do Censo 2022 (SIDRA 4709). AL ponderada 2025 = 53,52 = baseline da ficha (meta: 80,00 até 2050). Valores com vírgula decimal no original (conferência: 2025 AC 52,43; PA 53,05).").font = NOTE_FONT

# ---------- 4.3.2 PER ----------
ws = wb.create_sheet("4.3.2_PER_renovaveis_aneel")
headers = ["UF", "Estado", "Potência fiscalizada total (MW)", "Potência renovável (MW)", "PER (% renovável)"]
ws.append(headers)
t = rr_t = 0.0
for uf in UFS:
    p = per.get(uf, {})
    tot_mw = float(p["potencia_total_kw"]) / 1000
    ren_mw = float(p["potencia_renovavel_kw"]) / 1000
    t += tot_mw
    rr_t += ren_mw
    ws.append([uf, NOMES_UF[uf], round(tot_mw, 1), round(ren_mw, 1), round(float(p["per_renovaveis_pct"]), 1)])
ws.append(["AL", "Amazônia Legal", round(t, 1), round(rr_t, 1), round(rr_t / t * 100, 1)])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 24, 22, 18])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: ANEEL — SIGA (dadosabertos.aneel.gov.br, CKAN), base de 25/08/2026. Usinas EM OPERAÇÃO; potência FISCALIZADA. Renováveis (ficha): hídrica, solar, eólica, biomassa. Proxy: participação na potência, não na energia ofertada. Baseline da ficha: 65,24% (2025); meta 80% (2050). Detalhe por fonte: aba de detalhe abaixo.").font = NOTE_FONT

ws = wb.create_sheet("4.3.2_PER_detalhe_fonte")
headers = ["UF"] + ["Hídrica", "Solar", "Eólica", "Biomassa", "Fóssil", "Nuclear"]
ws.append(headers)
for uf in UFS:
    f_ = pot_fonte.get(uf, {})
    ws.append([uf] + [round(f_.get(k, 0.0), 1) for k in ("Hídrica", "Solar", "Eólica", "Biomassa", "Fóssil", "Nuclear")])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 12, 12, 12, 12, 12, 12])
ws.cell(row=ws.max_row + 2, column=1, value="Potência fiscalizada (MW) por origem do combustível, usinas em operação (SIGA/ANEEL, ago/2026).").font = NOTE_FONT

# ---------- 4.4.1 ISGR ----------
ws = wb.create_sheet("4.4.1_ISGR_saneamento")
headers = ["UF", "Estado", "Domicílios ocupados (Censo 2022)", "% água adequada", "% esgoto adequado",
           "DOM efetivo (com FClima×FGov)", "ISGR (%)"]
ws.append(headers)
for uf in UFS:
    g = isgr.get(uf, {})
    ws.append([uf, NOMES_UF[uf], int(g["dom_total"]), float(g["pct_agua_adeq"]),
               float(g["pct_esgoto_adeq"]), round(float(g["dom_efetivo"]), 0), float(g["isgr_pct"])])
g = isgr.get("AL", {})
ws.append(["AL", "Amazônia Legal", int(g["dom_total"]), float(g["pct_agua_adeq"]),
           float(g["pct_esgoto_adeq"]), round(float(g["dom_efetivo"]), 0), float(g["isgr_pct"])])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 22, 15, 16, 24, 11])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: IBGE — Censo 2022 (SIDRA 6803 água; 6805 esgoto, 808 municípios da AL) + MUNIC 2024 (FTP IBGE). Fórmula da ficha: DOM_efetivo = min(água_adeq, esgoto_adeq) × FClima × FGov; ISGR = Σ DOM_efetivo/Σ DOM_total × 100. Água adequada = rede geral (usa/outra forma) + poço profundo; esgoto adequado = rede geral/pluvial/fossa ligada + fossa séptica não ligada. FClima=1 se município respondeu 'Sim' em Magr18/Mhab088/Mtic266 (0,9 caso contrário); FGov=1 se Mgov086='Sim' (0,95 c/c). FHidro não publicado => 1; baseline oficial 41,52% inclui esse fator.").font = NOTE_FONT

# ---------- Pendentes ----------
ws = wb.create_sheet("Pendentes_anotacoes")
ws["A1"] = "Indicadores do Eixo 4 sem dado automatizado completo na coleta (30/08/2026)"
ws["A1"].font = TITLE_FONT
anot = [
    ("I4.2.1 Taxa de adequação e trafegabilidade de transportes",
     "Painel da Pesquisa CNT de Rodovias (data.cnt.org.br) é embed Power BI — sem API ou download aberto; o 'Painel de Dados' não expõe os trechos por estado de conservação. DNIT vgeo (servicos.dnit.gov.br/vgeo) publica geometria das rodovias sem atributo de qualidade. A fórmula da ficha exige extensão ponderada por estado (ótimo 1,0 / bom 0,8 / regular 0,5 / ruim 0,2 / péssimo 0,05) para rodovias estaduais, ferrovias e hidrovias. Caminho: solicitar planilha da Pesquisa CNT ao CNT/SENCALL; ou compilar relatórios anuais da Pesquisa CNT (PDF) por UF — levantamento documental."),
    ("I4.3.1 Indicador de Transição Energética (ITEQ)",
     "Fórmula completa exige: (a) POP_SIN/POP_total por município — PASI/EPE (pasi.epe.gov.br) tem endpoints de exportação (ExportarDadosMercadoLocalidade etc.) mas retornam erro de autenticação sem token da aplicação; (b) DEC/FEC por município — portalrelatorios.aneel.gov.br/recalc/desempenhoMunicipio sem API documentada; (c) renovabilidade do SIN na AL — componente obtido via SIGA (ver aba PER). Coletado: parcela renovável da potência na AL (85,2%). Caminho: obter base de localidades atendidas por Sistemas Isolados do PASI (download manual no portal) + DEC/FEC via portal de relatórios ANEEL (exportação manual)."),
    ("I4.4.2 Capacidade adaptativa urbana (>0,5)",
     "Painel Cidades do AdaptaBrasil MCTI (painelcidades.adaptabrasil.mcti.gov.br) é SPA; env.js não expõe endpoints e a API do AdaptaBrasil responde 403 (verificado 18/08 e 30/08/2026). Caminho: navegação manual no painel por município (772 da AL) ou solicitação de base ao MCTI/AdaptaBrasil."),
]
r = 3
for titulo, texto in anot:
    ws.cell(row=r, column=1, value=titulo).font = Font(bold=True, size=10)
    ws.cell(row=r + 1, column=1, value=texto).alignment = WRAP
    r += 3
largura(ws, [130])

os.makedirs(os.path.join(PASTA, "entregaveis"), exist_ok=True)
out = os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo4_Amazonia2050.xlsx")
wb.save(out)
print("Workbook salvo:", out)
