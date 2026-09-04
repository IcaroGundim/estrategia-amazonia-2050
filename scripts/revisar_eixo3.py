#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Revisão independente do Eixo 3 (30/08/2026).

1. PEVS (SIDRA 289): re-consulta com padrões de URL diferentes (2020 e 2024,
   chamada por UF individual, produto×total) vs CSVs salvos.
2. PIA (1849/10457): re-consulta 2023 e 2024 vs CSVs; soma das divisões vs Total.
3. RAIS: re-extrai os 7z brutos e re-agrega com implementação independente
   (parse e contagem por outro caminho) vs CSVs; sensibilidade do filtro
   'Ind Atividade Ano = 1'; total Brasil vs divulgação pública.
4. Workbook vs CSVs (aba PEVS, RAIS, PIA e Matriz_UF).
"""
import csv
import glob
import gzip
import io
import json
import os
import shutil
import sys
import tempfile
import time
import urllib.request

import py7zr
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
EIXO3 = os.path.join(PASTA, "dados", "eixo3")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
COD_UF = {"RO": "11", "AC": "12", "AM": "13", "RR": "14", "PA": "15", "AP": "16", "TO": "17", "MA": "21", "MT": "51"}
problemas = []


def check(nome, cond, detalhe=""):
    print(("[OK]    " if cond else "[FALHA] ") + nome + ((" — " + detalhe) if detalhe else ""))
    if not cond:
        problemas.append(nome)


def sidra(url, tentativas=3):
    for i in range(tentativas):
        try:
            req = urllib.request.Request(url, headers={"Accept-Encoding": "identity", "User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=300) as r:
                data = r.read()
                if data[:2] == b"\x1f\x8b":
                    data = gzip.decompress(data)
                return json.loads(data)
        except Exception:
            if i == tentativas - 1:
                raise
            time.sleep(3)


def num(v):
    if v in (None, "", "-", "...", "X", ".."):
        return None
    try:
        return float(str(v).strip())
    except ValueError:
        return None


# ================= 1. PEVS =================
print("===== 1. PEVS (SIDRA 289) =====")
pevs_csv = {}
for r in csv.DictReader(open(os.path.join(EIXO3, "pevs_total_uf_ano.csv"), encoding="utf-8")):
    pevs_csv[(r["uf"], int(r["ano"]))] = num(r["valor_mil_rs"]) if r["valor_mil_rs"] else None

# 1a. re-consulta combinada (2024 e 2020) — caminho direto
for ano in (2024, 2020):
    rows = sidra(f"https://apisidra.ibge.gov.br/values/t/289/n3/11,12,13,14,15,16,17,21,51/v/145/p/{ano}/c193/0?formato=json")
    ok = 0
    for x in rows[1:]:
        uf = COD_UF_INV = {v: k for k, v in COD_UF.items()}[x["D1C"]]
        alvo = pevs_csv.get((uf, ano))
        ok += (num(x["V"]) == alvo)
    check(f"PEVS {ano}: re-consulta n3 (9 UFs) = CSV", ok == 9, f"{ok}/9 conferem")

# 1b. chamadas individuais por UF (detecta troca de linha/UF)
difs = []
for uf in UFS:
    rows = sidra(f"https://apisidra.ibge.gov.br/values/t/289/n3/{COD_UF[uf]}/v/145/p/2023/c193/0?formato=json")
    if num(rows[1]["V"]) != pevs_csv.get((uf, 2023)):
        difs.append(uf)
check("PEVS 2023: 9 chamadas individuais por UF = CSV", not difs, str(difs))

# 1c. soma dos SUBPRODUTOS (folhas "7.1 - ...") = Total (2024, todas as UFs)
#     grupos ("1 - Alimentícios") são somatório dos filhos e não entram
import re as _re
FOLHA = _re.compile(r"^\d+\.\d+ ")
rows = sidra("https://apisidra.ibge.gov.br/values/t/289/n3/11,12,13,14,15,16,17,21,51/v/145/p/2024/c193/all?formato=json")
soma_prod = {}
for x in rows[1:]:
    if FOLHA.match(x["D4N"]):
        soma_prod.setdefault(x["D1C"], 0.0)
        soma_prod[x["D1C"]] += num(x["V"]) or 0.0
difs = []
for x in rows[1:]:
    if x["D4C"] == "0":
        uf = {v: k for k, v in COD_UF.items()}[x["D1C"]]
        if abs(soma_prod[x["D1C"]] - num(x["V"])) > 1:
            difs.append(f"{uf}: {soma_prod[x['D1C']]:.0f} vs {num(x['V']):.0f}")
check("PEVS 2024: soma dos SUBPRODUTOS = categoria Total (9 UFs)", not difs, "; ".join(difs))

# 1d. Brasil (n1) vs soma AL — concentração extrativista da AL (açaí, castanha,
#     madeira nativa, babaçu são quase 100% AL; fora da AL: erva-mate, carnaúba, pinho)
rows = sidra("https://apisidra.ibge.gov.br/values/t/289/n1/all/v/145/p/2024/c193/0?formato=json")
br = num(rows[1]["V"])
al = sum(v for (uf, a), v in pevs_csv.items() if a == 2024)
check("PEVS 2024: soma AL = 55-85% do Brasil (extrativismo concentrado na AL)", 0.55 <= al / br <= 0.85, f"AL {al:,.0f} vs BR {br:,.0f} ({al/br*100:.1f}%)")

# ================= 2. PIA =================
print("\n===== 2. PIA-Empresa (1849 / 10457) =====")
pia_csv = {}
for r in csv.DictReader(open(os.path.join(EIXO3, "pia_industria_uf_ano.csv"), encoding="utf-8")):
    pia_csv[(r["uf"], int(r["ano"]), r["tabela"])] = {
        "vti": num(r["valor_transf_ind_mil_rs"]),
        "po": num(r["pessoal_ocupado_3112"]),
        "ul": num(r["unidades_locais"]),
        "empresas": num(r["empresas"]),
    }

for tabela, ano in ((1849, 2023), (10457, 2024)):
    rows = sidra(f"https://apisidra.ibge.gov.br/values/t/{tabela}/n3/11,12,13,14,15,16,17,21,51/v/811,631/p/{ano}/c12762/117897?formato=json")
    ok_vti = ok_po = 0
    for x in rows[1:]:
        uf = {v: k for k, v in COD_UF.items()}[x["D1C"]]
        alvo = pia_csv[(uf, ano, str(tabela))]
        if x["D2C"] == "811":
            ok_vti += (num(x["V"]) == alvo["vti"])
        else:
            ok_po += (num(x["V"]) == alvo["po"])
    check(f"PIA {tabela} ({ano}): VTI e pessoal re-consultados = CSV", ok_vti == 9 and ok_po == 9, f"VTI {ok_vti}/9, PO {ok_po}/9")

# soma das divisões vs Total (2023 na 1849 e 2024 na 10457) — níveis de divisão
# com supressão por sigilo ("X" da SIDRA): a soma pode ser MENOR que o Total,
# mas a lacuna só é válida se houver categorias suprimidas naquele UF
for tabela, ano in ((1849, 2023), (10457, 2024)):
    rows = sidra(f"https://apisidra.ibge.gov.br/values/t/{tabela}/n3/11,12,13,14,15,16,17,21,51/v/811/p/{ano}/c12762/all?formato=json")
    soma, tot, suprimidos = {}, {}, {}
    for x in rows[1:]:
        if x["D4C"] == "117897":
            tot[x["D1C"]] = num(x["V"]) or 0.0
        elif len(x["D4N"]) >= 3 and x["D4N"][0].isdigit() and x["D4N"][1].isdigit() and x["D4N"][2] == " ":
            if x["V"] in ("X", "x"):
                suprimidos.setdefault(x["D1C"], 0)
                suprimidos[x["D1C"]] += 1
            soma.setdefault(x["D1C"], 0.0)
            soma[x["D1C"]] += num(x["V"]) or 0.0
    difs, infos = [], []
    for cod, v in sorted(tot.items()):
        uf = {v_: k for k, v_ in COD_UF.items()}[cod]
        s = soma.get(cod, 0)
        lacuna = v - s
        if lacuna < -1:
            difs.append(f"{uf}: soma {s:,.0f} > total {v:,.0f}")
        elif lacuna > 1 and suprimidos.get(cod, 0) == 0:
            difs.append(f"{uf}: lacuna de {lacuna:,.0f} sem categorias suprimidas")
        elif lacuna > 1:
            infos.append(f"{uf} {lacuna/v*100:.1f}% sigilo ({suprimidos[cod]} div.)")
    check(f"PIA {tabela} ({ano}): soma das divisões ≤ Total e toda lacuna explicada por 'X' de sigilo",
          not difs, "; ".join(difs))
    print(f"        lacunas por sigilo: {', '.join(infos) if infos else 'nenhuma'}")

# 10457: empresas e unidades locais (só existem na 10457)
rows = sidra("https://apisidra.ibge.gov.br/values/t/10457/n3/15,13/v/2086,13816,631/p/2024/c12762/117897?formato=json")
ok = all(num(x["V"]) is not None for x in rows[1:])
check("PIA 10457 (2024): empresas/unidades/pessoal PA e AM não nulos", ok)

# ================= 3. RAIS =================
print("\n===== 3. RAIS Estabelecimentos (re-agregação independente dos 7z) =====")
UF_SIGLA = {"11": "RO", "12": "AC", "13": "AM", "14": "RR", "15": "PA", "16": "AP", "17": "TO", "21": "MA", "51": "MT"}


def re_agregar(ano):
    """Re-extrai e re-agrega com caminho de código diferente do coletor original."""
    arq = os.path.join(EIXO3, "rais", f"RAIS_ESTAB_PUB_{ano}.7z")
    dest = os.path.join(tempfile.gettempdir(), f"revisao3_rais_{ano}")
    comt = os.path.join(dest, "RAIS_ESTAB_PUB.COMT")
    if not os.path.exists(comt):
        if os.path.exists(dest):
            shutil.rmtree(dest, ignore_errors=True)
        os.makedirs(dest)
        with py7zr.SevenZipFile(arq) as z:
            z.extractall(path=dest)
    com_uf, com_div, fora_ativo = {}, {}, [0, 0]
    with open(comt, "r", encoding="latin-1", newline="") as f:
        rd = csv.reader(f)
        cols = [c.strip() for c in next(rd)]
        i_uf, i_cl, i_va, i_at = (cols.index("UF - Código"), cols.index("CNAE 2.0 Classe - Código"),
                                  cols.index("Qtd Vínculos Ativos"), cols.index("Ind Atividade Ano - Código"))
        for linha in rd:
            cod = linha[i_uf].strip()
            ativo = linha[i_at].strip() == "1"
            if not ativo:
                v = linha[i_va].strip()
                if v.isdigit() and int(v) > 0:
                    fora_ativo[1] += int(v)
                fora_ativo[0] += 1
            if cod not in UF_SIGLA:
                continue
            cl = linha[i_cl].strip().strip('"').strip()
            div = cl[:2] if cl[:2].isdigit() else "NA"
            v = linha[i_va].strip()
            v = int(v) if v.isdigit() else 0
            if ativo:
                com_uf[cod] = com_uf.get(cod, 0) + 1
                e, t = com_div.get((cod, div), (0, 0))
                com_div[(cod, div)] = (e + 1, t + v)
    shutil.rmtree(dest, ignore_errors=True)
    return com_uf, com_div, fora_ativo


for ano in (2023, 2024):
    csv_uf = {r["uf"]: int(r["estabelecimentos_ativos"])
              for r in csv.DictReader(open(os.path.join(EIXO3, f"rais_estab_uf_ano.csv"), encoding="utf-8"))
              if int(r["ano"]) == ano}
    csv_div = {}
    for r in csv.DictReader(open(os.path.join(EIXO3, "rais_estab_uf_divisao_ano.csv"), encoding="utf-8")):
        if int(r["ano"]) == ano:
            csv_div[(r["uf"], r["divisao_cnae"])] = (int(r["estabelecimentos_ativos"]), int(r["vinculos_ativos"]))
    uf2, div2, fora = re_agregar(ano)
    sigla2 = {COD_UF[s]: s for s in UFS}
    div2_sigla = {(sigla2[cod], div): v for (cod, div), v in div2.items()}
    iguais_uf = all(uf2.get(cod) == csv_uf.get(sigla) for cod, sigla in UF_SIGLA.items())
    iguais_div = set(div2_sigla) == set(csv_div) and all(div2_sigla[k] == csv_div[k] for k in div2_sigla)
    check(f"RAIS {ano}: re-agregação independente = CSV (estabelecimentos por UF)", iguais_uf)
    check(f"RAIS {ano}: re-agregação independente = CSV (UF × divisão, {len(div2_sigla)} pares)", iguais_div)
    tot_v = sum(t for (_, _), (_, t) in csv_div.items())
    print(f"        sensibilidade: vínculos em estabelecimentos INATIVOS no ano (fora do filtro) = {fora[1]:,} em {fora[0]:,} linhas")

# total Brasil (o arquivo é nacional) vs divulgação pública (~53,6 mi em 2023/2024)
arq = os.path.join(EIXO3, "rais", "RAIS_ESTAB_PUB_2024.7z")
dest = os.path.join(tempfile.gettempdir(), "revisao3_rais_br")
comt = os.path.join(dest, "RAIS_ESTAB_PUB.COMT")
if not os.path.exists(comt):
    if os.path.exists(dest):
        shutil.rmtree(dest, ignore_errors=True)
    os.makedirs(dest)
    with py7zr.SevenZipFile(arq) as z:
        z.extractall(path=dest)
br_uf = {}
with open(comt, "r", encoding="latin-1", newline="") as f:
    rd = csv.reader(f)
    cols = [c.strip() for c in next(rd)]
    i_uf, i_va, i_at = cols.index("UF - Código"), cols.index("Qtd Vínculos Ativos"), cols.index("Ind Atividade Ano - Código")
    for linha in rd:
        if linha[i_at].strip() != "1":
            continue
        cod = linha[i_uf].strip()
        v = linha[i_va].strip()
        if v.isdigit():
            br_uf[cod] = br_uf.get(cod, 0) + int(v)
br_total = sum(br_uf.values())
shutil.rmtree(dest, ignore_errors=True)
al_total = sum(br_uf.get(cod, 0) for cod in UF_SIGLA)
check("RAIS 2024: total BRASIL dos vínculos ativos em 53-58 mi (faixa da divulgação)", 53_000_000 <= br_total <= 58_000_000, f"{br_total:,}")
check("RAIS 2024: AL = 9-12% do Brasil (emprego formal menor que o share de população)", 0.09 <= al_total / br_total <= 0.12, f"{al_total/br_total*100:.1f}%")

# ================= 4. WORKBOOK vs CSVs =================
print("\n===== 4. Workbook vs CSVs =====")
wb = load_workbook(os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo3_Amazonia2050.xlsx"), data_only=True)


def parse_br(s):
    if s in (None, "—", ""):
        return None
    return float(str(s).replace(".", "").replace(",", "."))


ws = wb["3.1.1_Pevs_serie"]
dif = []
for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
    uf = row[0]
    for j, ano in enumerate(range(2015, 2025)):
        cel = parse_br(row[2 + j])
        if (cel or None) != pevs_csv.get((uf, ano)):
            dif.append(f"{uf}/{ano}: {cel} vs {pevs_csv.get((uf, ano))}")
check("Workbook PEVS série (9 UFs × 10 anos) = CSV", not dif, "; ".join(dif[:4]))

ws = wb["F3.5_Pia_transformacao"]
dif = []
for row in ws.iter_rows(min_row=2, max_row=10, values_only=True):
    uf = row[0]
    for j, ano in enumerate(range(2015, 2025)):
        cel = parse_br(row[2 + j])
        alvo = pia_csv.get((uf, ano, "1849" if ano <= 2023 else "10457"))["vti"]
        if (cel or None) != alvo:
            dif.append(f"{uf}/{ano}: {cel} vs {alvo}")
check("Workbook PIA série (9 UFs × 10 anos) = CSV", not dif, "; ".join(dif[:4]))

ws = wb["F3.2_Rais_empregos"]
linhas = {row[0]: row for row in ws.iter_rows(min_row=2, max_row=10, values_only=True)}
dif = []
for uf in UFS:
    vinc_csv = {}
    est_csv = {}
    for r in csv.DictReader(open(os.path.join(EIXO3, "rais_estab_uf_divisao_ano.csv"), encoding="utf-8")):
        if r["uf"] == uf:
            vinc_csv[int(r["ano"])] = vinc_csv.get(int(r["ano"]), 0) + int(r["vinculos_ativos"])
            est_csv[int(r["ano"])] = est_csv.get(int(r["ano"]), 0) + int(r["estabelecimentos_ativos"])
    row = linhas[uf]
    if parse_br(row[2]) != est_csv[2023] or parse_br(row[3]) != est_csv[2024] or \
       parse_br(row[4]) != vinc_csv[2023] or parse_br(row[5]) != vinc_csv[2024]:
        dif.append(uf)
check("Workbook RAIS empregos (9 UFs, 4 colunas) = CSV", not dif, str(dif))

ws = wb["Matriz_UF"]
linhas = {row[0]: row for row in ws.iter_rows(min_row=2, values_only=True)}
dif = []
for uf in UFS:
    esperado = pevs_csv.get((uf, 2024))
    if parse_br(linhas["I3.1.1"][4 + UFS.index(uf)]) != esperado:
        dif.append(uf)
check("Workbook Matriz_UF linha I3.1.1 = PEVS 2024", not dif, str(dif))

# divisões RAIS do workbook = CSV (amostra: PA 2024; workbook: G=vínculos 2024, I=estab 2024)
ws = wb["F3.2_Rais_divisoes"]
wb_div = {row[2]: (parse_br(row[6]), parse_br(row[8])) for row in ws.iter_rows(min_row=2, values_only=True) if row[0] == "PA"}
csv_pa = {}
for r in csv.DictReader(open(os.path.join(EIXO3, "rais_estab_uf_divisao_ano.csv"), encoding="utf-8")):
    if r["uf"] == "PA" and int(r["ano"]) == 2024:
        csv_pa[r["divisao_cnae"]] = (int(r["vinculos_ativos"]), int(r["estabelecimentos_ativos"]))
check(f"Workbook RAIS divisões PA 2024 = CSV ({len(csv_pa)} divisões)", wb_div == csv_pa)

print("\n===== RESUMO =====")
if problemas:
    print(f"{len(problemas)} FALHA(S):")
    for p in problemas:
        print(" -", p)
    sys.exit(1)
print("Todas as verificações passaram.")
