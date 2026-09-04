# -*- coding: utf-8 -*-
"""Auditoria de sanidade do workbook do Eixo 5."""
import openpyxl, os

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
ARQ = os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo5_Amazonia2050.xlsx")
wb = openpyxl.load_workbook(ARQ)
print("ABAS:", wb.sheetnames)

# Catalogo
ws = wb["Catalogo_Indicadores"]
print("\nCatalogo:", ws.max_row - 1, "indicadores")
for r in ws.iter_rows(min_row=2, values_only=True):
    print(" ", r[0], "|", str(r[2])[:45], "|", str(r[8])[:70])

# Matriz_UF
ws = wb["Matriz_UF"]
print("\nMatriz_UF:")
for r in ws.iter_rows(min_row=2, values_only=True):
    print(" ", r[0], "|", r[3], "|", r[4:13], "|", str(r[13])[:24])

# CAPAG: conferir contagens AL
ws = wb["5.5.1_CAPAG_STN"]
rows = list(ws.iter_rows(values_only=True))
for r in rows:
    if r and r[0] in ("AL",) or (r and r[0] in ("Ano", "A ou B (nº UF)", "% da AL", "A/B estritos (nº UF)")):
        print("CAPAG>", r[:11])

# P&D: soma AL por ano
ws = wb["5.4.1_PnD_MCTI_IBGE"]
rows = list(ws.iter_rows(values_only=True))
for r in rows:
    if r and r[0] == "AL":
        print("PD AL R$ mi>", r[:12])

# conferência cruzada com CSV bruto
import csv
tot = {}
with open(os.path.join(PASTA, "dados", "eixo5", "pd_uf_ano.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        if r["pd_mi"]:
            tot[r["ano"]] = tot.get(r["ano"], 0) + float(r["pd_mi"])
print("CSV AL 2000:", round(tot.get("2000", 0), 2), "| 2010:", round(tot.get("2010", 0), 2), "| 2023:", round(tot.get("2023", 0), 2))
