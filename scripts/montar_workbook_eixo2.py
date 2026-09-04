#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Monta 'Indicadores_Resultado_Eixo2_Amazonia2050.xlsx' — Eixo 2: Inclusão social,
segurança e qualidade de vida (Estratégia Amazônia 2050), 9 estados da AL."""
import csv, os, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
DADOS = os.path.join(PASTA, "dados")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
NOMES_UF = {"AC": "Acre", "AP": "Amapá", "AM": "Amazonas", "MA": "Maranhão",
            "MT": "Mato Grosso", "PA": "Pará", "RO": "Rondônia", "RR": "Roraima", "TO": "Tocantins"}

# ---------- CNES: estabelecimentos tipo TELESSAUDE por UF (Jul/2026) ----------
telessaude = {}
idx_tel = None
for linha in open(os.path.join(DADOS, "cnes", "cnes_uf_tipo.csv"), encoding="latin-1"):
    linha = linha.strip()
    if not linha:
        continue
    partes = [p.strip().strip('"') for p in linha.split(";")]
    p0 = partes[0].lower()
    if p0.startswith("cnes") or p0.startswith("por ") or p0.startswith("per") \
       or p0.startswith("fonte") or p0.startswith("nota") or p0.startswith("-"):
        continue
    if p0.startswith("unidade da federação"):
        idx_tel = partes.index("TELESSAUDE")
        continue
    if partes[0] == "Total":
        continue
    m = re.match(r"^(\d+)\s*([A-Za-zÀ-Üà-ü ]+)", partes[0])
    if not m:
        continue
    import unicodedata
    norm = lambda s: unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode()
    uf = m.group(2).strip()
    sigla = next((s for s, n in NOMES_UF.items() if norm(n).upper() == norm(uf).upper()), None)
    if sigla and idx_tel is not None:
        telessaude[sigla] = int(partes[idx_tel])

# ---------- CVLI por UF/ano (Sinesp/MJ) ----------
cvli = {}
with open(os.path.join(DADOS, "sinesp", "cvli_uf_ano.csv")) as f:
    for r in csv.DictReader(f):
        cvli.setdefault(r["uf"], {})[int(r["ano"])] = int(r["cvli"])

# ---------- População por UF/ano (IBGE, projeção 2024) ----------
pop = {}
with open(os.path.join(DADOS, "ibge_pop", "populacao_uf_ano.csv")) as f:
    for r in csv.DictReader(f):
        pop.setdefault(r["uf"], {})[int(r["ano"])] = int(r["populacao"])

# ---------- Equipes de saúde por UF (CNES Jul/2026) ----------
equipes = {}
idx_esf01 = idx_esf70 = idx_eap = None
for linha in open(os.path.join(DADOS, "cnes", "cnes_equipes_uf.csv"), encoding="latin-1"):
    linha = linha.strip()
    if not linha:
        continue
    partes = [p.strip().strip('"') for p in linha.split(";")]
    p0 = partes[0].lower()
    if p0.startswith("cnes") or p0.startswith("por ") or p0.startswith("per") \
       or p0.startswith("fonte") or p0.startswith("nota") or p0.startswith("-"):
        continue
    if p0.startswith("unidade da federação"):
        idx_esf01 = partes.index("01 ESF - EQUIPE DE SAUDE DA FAMILIA") if "01 ESF - EQUIPE DE SAUDE DA FAMILIA" in partes else None
        idx_esf70 = partes.index("70 ESF - EQUIPE DE SAUDE DA FAMILIA") if "70 ESF - EQUIPE DE SAUDE DA FAMILIA" in partes else None
        idx_eap = partes.index("76 EAP - EQUIPE DE ATENCAO PRIMARIA") if "76 EAP - EQUIPE DE ATENCAO PRIMARIA" in partes else None
        continue
    if partes[0] == "Total":
        continue
    import re as _re, unicodedata as _u
    m = _re.match(r"^(\d+)\s*([A-Za-zÀ-Üà-ü ]+)", partes[0])
    if not m:
        continue
    norm = lambda s: _u.normalize("NFD", s).encode("ascii", "ignore").decode()
    uf = m.group(2).strip()
    sigla = next((s for s, n in NOMES_UF.items() if norm(n).upper() == norm(uf).upper()), None)
    if sigla:
        getv = lambda i: int(partes[i]) if (i is not None and partes[i] not in ("-", "")) else 0
        equipes[sigla] = (getv(idx_esf01), getv(idx_esf70), getv(idx_eap))

# ---------- SIS 2025 (PNADc 2024): pobreza e frequência escolar por UF ----------
pobreza = {}
with open(os.path.join(DADOS, "ibge_sis", "sis_pobreza_uf.csv")) as f:
    for r in csv.DictReader(f):
        pobreza[r["uf"]] = r
freq = {}
with open(os.path.join(DADOS, "ibge_sis", "sis_freq_escolar_uf.csv")) as f:
    for r in csv.DictReader(f):
        freq[r["uf"]] = r

# ---------- População por faixa 0-17 (para média ponderada 4-17) ----------
pop_faixa = defaultdict(dict)  # uf -> faixa -> {ano: pop}
with open(os.path.join(DADOS, "ibge_pop", "populacao_faixas_0_17.csv")) as f:
    for r in csv.DictReader(f):
        pop_faixa[r["uf"]].setdefault(r["faixa"], {})[int(r["ano"])] = int(r["populacao"])

def freq_4_17(uf, ano=2024):
    """média ponderada da taxa de atendimento 4-17 (faixas 4-5, 6-14, 15-17)."""
    pesos = {fx: pop_faixa.get(uf, {}).get(fx, {}).get(ano, 0) for fx in ("4_5", "6_14", "15_17")}
    tot = sum(pesos.values())
    if not tot:
        return None
    taxas = {"4_5": freq.get(uf, {}).get("4_5"), "6_14": freq.get(uf, {}).get("6_14"), "15_17": freq.get(uf, {}).get("15_17")}
    num = sum(float(taxas[fx]) * pesos[fx] for fx in pesos if taxas[fx] not in (None, ""))
    return num / tot

# ---------- Mortalidade evitável <5 anos (SIM/TabNet, 2024) ----------
morte5 = {}
for linha in open(os.path.join(DADOS, "sim", "sim_evitaveis_menores5_2024.csv"), encoding="latin-1"):
    linha = linha.strip()
    if not linha or linha.startswith("Óbitos") or linha.startswith(" por ") or linha.startswith("Per"):
        continue
    partes = [p.strip().strip('"') for p in linha.split(";")]
    if not partes or not partes[0][:1].isdigit():
        continue
    m = re.match(r"^(\d+)\s*([A-Za-zÀ-Üà-ü ]+)", partes[0])
    if not m:
        continue
    import unicodedata as _u2
    norm2 = lambda s: _u2.normalize("NFD", s).encode("ascii", "ignore").decode()
    uf = m.group(2).strip()
    sigla = next((s for s, n in NOMES_UF.items() if norm2(n).upper() == norm2(uf).upper()), None)
    if sigla:
        morte5[sigla] = int(partes[-1])  # última coluna = Total

# ---------- Catálogo Eixo 2 ----------
CATALOGO = [
    ("I2.1.1", "2.1", "Pobreza (CadÚnico)", "Redução da taxa de pobreza para patamar ≤ 3%",
     "Menos de 3% da população em situação de pobreza (CadUnico)",
     "%", "MDS (Cadastro Único); proxy: SIS/PNADc 2024 (IBGE)", 2050, "coletado (proxy SIS/PNADc 2024: % abaixo das linhas de pobreza)"),
    ("I2.1.2", "2.1", "Cobertura de programas de inclusão produtiva", "Ampliação da população apoiada por programas integrados de inclusão produtiva e social",
     "Todos os territórios (municípios) com cobertura de pelo menos 4 programas federais (Brasil Sem Fome, Bolsa Verde, PNAE e PAA)",
     "nº de programas", "Estados", 2050, "coleta manual"),
    ("I2.2.1", "2.2", "Mortalidade por causas evitáveis", "Redução de mortes evitáveis em pelo menos 50% na região",
     "Taxa de mortalidade por causas evitáveis, segundo metodologia do SUS (Lista Brasileira de Causas de Mortes Evitáveis)",
     "taxa / 100 mil", "DATASUS (SIM/TabNet)", 2050, "parcial (TabNet só disponibiliza LBE <5 anos — dado de referência incluído; LBE completa exige microdados SIM)"),
    ("I2.2.2", "2.2", "Cobertura da Atenção Primária", "Ampliar a cobertura da APS para 100% da população",
     "% de cobertura populacional de APS",
     "%", "Ministério da Saúde (e-Gestor Atenção Básica)", 2035, "parcial (nº eSF/eAP do CNES como referência; % oficial exige e-Gestor)"),
    ("I2.2.3", "2.2", "Telessaúde em municípios remotos", "Atendimento em territórios remotos, com pelo menos 50% dos municípios da AL com telessaúde ativa",
     "Municípios com serviço de telessaúde ativo",
     "nº de municípios", "Ministério da Saúde; Secretarias Estaduais de Saúde (CNES)", 2050, "coletado (proxy: estabelecimentos TELESSAUDE ativos no CNES)"),
    ("I2.3.1", "2.3", "IDEB", "Alcançar as metas de proficiência (IDEB) do INEP nas três etapas",
     "Percentual de municípios/estados da AL que atingiram ou superaram a meta do IDEB (Anos Iniciais, Anos Finais e Ensino Médio)",
     "%", "INEP (IDEB)", 2050, "pendente (INEP Data sem API aberta; download.inep.gov.br bloqueado)"),
    ("I2.3.2", "2.3", "Atendimento escolar 4-17 anos", "Universalizar o acesso à educação básica de qualidade, inclusive em territórios remotos",
     "Taxa de atendimento escolar da população de 4 a 17 anos (%) e creche (0 a 3 anos)",
     "%", "INEP e IBGE (SIS/PNADc 2024)", 2050, "coletado (SIS/PNADc 2024: faixas etárias por UF)"),
    ("I2.4.1", "2.4", "CVLI", "Reduzir as mortes por CVLI para 10 mortes por 100 mil habitantes",
     "Taxa de Crimes Violentos Letais Intencionais (CVLI) por 100 mil habitantes",
     "taxa / 100 mil", "Atlas da Violência; MJ (Sinesp VDE) + IBGE", 2050, "coletado (Sinesp + população IBGE)"),
    ("I2.4.2", "2.4", "Segurança cidadã e planos de adaptação", "Estratégias integradas de segurança cidadã e ambiental em 100% dos municípios",
     "100% dos municípios com aderência formal a programas de segurança cidadã; 100% com Planos de Adaptação Climática produzidos e ativos",
     "% / Sim-Não", "PRONASCI 2 (MJ); AdaptaBrasil (MCTI)", 2050, "coleta manual"),
    ("I2.5.1", "2.5", "Renda da sociobioeconomia", "Criação e fomento de empregos verdes e renda com foco em jovens, mulheres e populações tradicionais",
     "Volume de renda no painel da sociobioeconomia do MMA",
     "R$", "MMA (Plataforma da Ecosociobiodiversidade)", 2050, "pendente (plataforma do MMA em construção)"),
]

# ---------- Estilo ----------
HDR_FILL = PatternFill("solid", fgColor="1B4332")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13, color="1B4332")
NOTE_FONT = Font(italic=True, size=9, color="555555")
BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")

def estilo_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER

def largura(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# ---------- Sobre ----------
ws = wb.active
ws.title = "Sobre"
ws["A1"] = "ESTRATÉGIA REGIONAL AMAZÔNIA 2050 — Eixo 2: Inclusão social, segurança e qualidade de vida"
ws["A1"].font = TITLE_FONT
ws["A3"] = "Indicadores de RESULTADO (somente resultado), organizados para os 9 estados da Amazônia Legal."
ws["A4"] = "Base: '3. Matriz Metas x Indicadores .xlsx' (aba atualizada, coluna INDICADOR DE RESULTADO) e fichas técnicas."
ws["A6"] = "Data de coleta: 18/08/2026."
ws["A7"] = "Status: coletado = dado oficial obtido; parcial = parte do indicador; pendente = fonte sem API aberta ou indisponível no dia; coleta manual = levantamento documental/estadual."
ws["A9"] = "FONTES E LIMITAÇÕES"
ws["A9"].font = Font(bold=True)
ws["A10"] = "• Telessaúde (I2.2.3): CNES via TabNet (competência Jul/2026) — nº de estabelecimentos do tipo TELESSAUDE com status ATIVO por UF. Proxy do indicador da matriz ('municípios com serviço de telessaúde ativo'); o número de municípios distintos exige extração por município no CNES."
ws["A11"] = "• CVLI (I2.4.1): bases oficiais Sinesp/VDE do Ministério da Justiça (bancovde-2020 a 2026). CVLI = homicídio doloso + latrocínio + lesão corporal seguida de morte (vítimas). 2026 = parcial (jan-jul). Taxa/100 mil calculada com a Projeção da População IBGE 2024 (ftp.ibge.gov.br)."
ws["A12"] = "• Mortalidade evitável (I2.2.1): o TabNet só mantém a tabela de evitáveis em <5 anos (LBE infantil, 2024) — incluída como referência; a LBE completa (0-74) exige microdados SIM/DATASUS."
ws["A13"] = "• Pobreza (I2.1.1): fonte da matriz é o CadÚnico (MDS); proxy coletada: SIS 2025/PNADc 2024 (IBGE, FTP oficial) com linhas de pobreza do Banco Mundial."
ws["A14"] = "• IDEB (I2.3.1): INEP Data é Power BI (sem API pública); download.inep.gov.br bloqueado — alternativa: INEP Data/portal ou solicitação direta."
ws["A15"] = "• Frequência escolar (I2.3.2): SIS 2025/PNADc 2024 (IBGE). APS (I2.2.2): % oficial exige e-Gestor (autenticado); incluído nº de equipes eSF/eAP do CNES como referência. Bioeconomia (I2.5.1): plataforma MMA em construção."
ws["A17"] = "Scripts: scripts/agregar_cvli.py. Dados brutos: dados/bancovde-*.xlsx, dados/cvli_uf_ano.csv."
ws["A17"].font = NOTE_FONT
largura(ws, [130])

# ---------- Catálogo ----------
ws = wb.create_sheet("Catalogo_Indicadores")
headers = ["Código", "Linha de Ação", "Indicador (nome curto)", "Meta", "Indicador de RESULTADO (texto da matriz)",
           "Unidade", "Fonte", "Prazo", "Status da coleta"]
ws.append(headers)
for row in CATALOGO:
    ws.append(list(row))
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = WRAP
largura(ws, [8, 9, 32, 40, 55, 12, 34, 9, 30])

# ---------- Matriz_UF ----------
ws = wb.create_sheet("Matriz_UF")
headers = ["Código", "Indicador de RESULTADO (resumo)", "Unidade", "Ano ref.", "AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO", "Status", "Fonte"]
ws.append(headers)
M = []
for cod, la, curto, meta, indicador, unid, fonte, prazo, status in CATALOGO:
    if cod == "I2.2.3":
        vals = [telessaude.get(uf, "—") for uf in UFS]
        anoref, status_, unid2 = "Jul/2026 (CNES)", "coletado", "estab. TELESSAUDE"
    elif cod == "I2.4.1":
        vals = [round(cvli.get(uf, {}).get(2025, 0) / pop.get(uf, {}).get(2025, 1) * 100000, 1) for uf in UFS]
        anoref, status_, unid2 = "2025", "coletado (Sinesp + pop. IBGE)", "taxa /100 mil"
    elif cod == "I2.2.2":
        vals = [equipes.get(uf, (0, 0, 0))[0] + equipes.get(uf, (0, 0, 0))[1] for uf in UFS]
        anoref, status_, unid2 = "Jul/2026 (CNES)", "parcial (nº eSF de referência; % oficial no e-Gestor)", "nº equipes ESF"
    elif cod == "I2.1.1":
        vals = [round(float(pobreza.get(uf, {}).get("pct_pobreza_usd365", 0)), 1) for uf in UFS]
        anoref, status_, unid2 = "2024 (SIS/PNADc)", "coletado (proxy — % abaixo de US$ 3,65/dia)", "% < US$ 3,65"
    elif cod == "I2.2.1":
        vals = [morte5.get(uf, "—") for uf in UFS]
        anoref, status_, unid2 = "2024 (SIM)", "parcial (referência: evitáveis <5 anos)", "óbitos <5 anos"
    elif cod == "I2.3.2":
        vals = [round(freq_4_17(uf), 1) if freq_4_17(uf) is not None else "—" for uf in UFS]
        anoref, status_, unid2 = "2024 (SIS/PNADc)", "coletado (média ponderada 4-17)", "% 4-17 (ponderado)"
    else:
        vals = ["—"] * 9
        anoref, status_, unid2 = "—", status, unid
    M.append([cod, curto, unid2, anoref] + vals + [status_, fonte])
for row in M:
    ws.append(row)
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = WRAP
largura(ws, [9, 44, 16, 14, 8, 8, 8, 8, 8, 8, 8, 8, 8, 26, 26])
ws.freeze_panes = "E2"

# ---------- 2.2.3 Telessaúde ----------
ws = wb.create_sheet("2.2.3_Telessaude_CNES")
headers = ["UF", "Estado", "Estabelecimentos TELESSAUDE ativos (Jul/2026)", "Participação na AL (%)"]
ws.append(headers)
tot = sum(telessaude.values())
for uf in UFS:
    v = telessaude.get(uf, 0)
    ws.append([uf, NOMES_UF[uf], v, round(v / tot * 100, 1)])
ws.append(["TOTAL", "Amazônia Legal", tot, 100.0])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 40, 22])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: CNES/DATASUS (TabNet, competência Jul/2026), estabelecimentos com status ATIVO classificados como TELESSAUDE. Proxy do indicador 'municípios com serviço de telessaúde ativo'.").font = NOTE_FONT

# ---------- 2.4.1 CVLI ----------
ws = wb.create_sheet("2.4.1_CVLI_Sinesp")
headers = ["UF", "Estado", "2020", "2021", "2022", "2023", "2024", "2025", "2026 (jan-jul)", "Pop. 2025 (IBGE)", "Taxa CVLI 2025 /100 mil"]
ws.append(headers)
for uf in UFS:
    d = cvli.get(uf, {})
    row = [uf, NOMES_UF[uf]]
    for a in range(2020, 2027):
        row.append(d.get(a, "—"))
    p2025 = pop.get(uf, {}).get(2025)
    row.append(f"{p2025:,}" if p2025 else "—")
    taxa = d.get(2025, 0) / p2025 * 100000 if p2025 else None
    row.append(round(taxa, 1) if taxa is not None else "—")
    ws.append(row)
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 9, 9, 9, 9, 9, 9, 13, 16, 18])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: Sinesp/VDE — MJ (bancovde-2020 a 2026). CVLI = homicídio doloso + latrocínio + lesão corporal seguida de morte. 2026 parcial (jan-jul). População: IBGE, Projeção 2024 (ftp.ibge.gov.br). Meta da matriz: 10 mortes/100 mil (2050).").font = NOTE_FONT

# ---------- 2.2.2 Equipes APS (referência) ----------
ws = wb.create_sheet("2.2.2_Equipes_APS_ref")
headers = ["UF", "Estado", "eSF (tipo 01)", "eSF (tipo 70)", "Total eSF", "eAP (tipo 76)", "Pop. 2026 (IBGE)"]
ws.append(headers)
for uf in UFS:
    e01, e70, eap = equipes.get(uf, (0, 0, 0))
    p2026 = pop.get(uf, {}).get(2026)
    ws.append([uf, NOMES_UF[uf], e01, e70, e01 + e70, eap, f"{p2026:,}" if p2026 else "—"])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 13, 13, 11, 13, 15])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: CNES/DATASUS (Jul/2026), equipes com status ATIVO. Dado de REFERÊNCIA: o indicador da matriz é % de cobertura populacional de APS (e-Gestor, exige autenticação). Fórmula da ficha: (nº eSF × 3.500 + eAP20h × 1.750 + eAP30h × 2.625 + vinculados eCR/eSFR/eAPP)/população × 100.").font = NOTE_FONT

# ---------- 2.1.1 Pobreza (SIS) ----------
ws = wb.create_sheet("2.1.1_Pobreza_SIS")
headers = ["UF", "Estado", "População (mil)", "% extrema pobreza (<US$ 2,15)", "% pobreza (<US$ 3,65)", "% <US$ 6,85", "% até 50% da mediana", "Linha 50% mediana (R$/mês)"]
ws.append(headers)
for uf in UFS:
    p = pobreza.get(uf)
    if p:
        ws.append([uf, NOMES_UF[uf], f"{float(p['pop_mil']):,.0f}", round(float(p["pct_extrema_usd215"]), 1),
                   round(float(p["pct_pobreza_usd365"]), 1), round(float(p["pct_usd685"]), 1),
                   round(float(p["pct_ate50mediana"]), 1), round(float(p["linha50_mediana_r$"]), 2)])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 14, 26, 22, 14, 20, 22])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: IBGE, Síntese de Indicadores Sociais 2025 (PNAD Contínua 2024), Tabela 2.18. Linhas de pobreza do Banco Mundial (PPC 2017). O indicador da matriz usa o CadÚnico (MDS); a coluna % < US$ 3,65/dia é a proxy mais próxima da linha do Bolsa Família. Meta da matriz: ≤3% da população em situação de pobreza (2050).").font = NOTE_FONT

# ---------- 2.3.2 Frequência escolar (SIS) ----------
ws = wb.create_sheet("2.3.2_Freq_escolar_SIS")
headers = ["UF", "Estado", "0-3 anos (creche)", "4-5 anos", "6-10 anos", "11-14 anos", "6-14 anos", "15-17 anos", "4-17 (média ponderada)"]
ws.append(headers)
for uf in UFS:
    fr = freq.get(uf)
    if fr:
        ws.append([uf, NOMES_UF[uf], round(float(fr["0_3"]), 1), round(float(fr["4_5"]), 1),
                   round(float(fr["6_10"]), 1), round(float(fr["11_14"]), 1), round(float(fr["6_14"]), 1),
                   round(float(fr["15_17"]), 1), round(freq_4_17(uf), 1) if freq_4_17(uf) is not None else "—"])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 16, 10, 10, 11, 10, 11, 18])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: IBGE, SIS 2025 (PNADc 2024), Tabela 4.1 (taxa de frequência escolar bruta). '4-17 (média ponderada)' = ponderação pelas populações de 4-5, 6-14 e 15-17 (Projeção IBGE 2024). Indicador da matriz: taxa de atendimento 4-17 e creche 0-3.").font = NOTE_FONT

# ---------- 2.2.1 Mortalidade evitável (referência <5) ----------
ws = wb.create_sheet("2.2.1_Mortalidade_evitavel_ref")
headers = ["UF", "Estado", "Óbitos evitáveis <5 anos (2024)", "Pop. 0-4 (2024)", "Taxa /1.000 nascidos (ref.)"]
ws.append(headers)
for uf in UFS:
    n = morte5.get(uf)
    p04 = sum(pop_faixa.get(uf, {}).get(fx, {}).get(2024, 0) for fx in ("0_3", "4_5"))
    ws.append([uf, NOMES_UF[uf], n if n is not None else "—",
               f"{p04:,}" if p04 else "—",
               round(n / p04 * 1000, 2) if (n is not None and p04) else "—"])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 26, 15, 24])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: SIM/DATASUS (TabNet, tabela de causas evitáveis em menores de 5 anos — LBE infantil, 2024). Dado de REFERÊNCIA: o indicador da matriz é a taxa de mortalidade por causas evitáveis segundo a LBE completa (0-74 anos), indisponível via TabNet — requer microdados SIM.").font = NOTE_FONT

# ---------- Anotações ----------
ws = wb.create_sheet("Pendentes_anotacoes")
ws["A1"] = "Indicadores do Eixo 2 sem dado automatizado na coleta (18/08/2026)"
ws["A1"].font = TITLE_FONT
anot = [
    ("I2.1.2 Territórios com ≥4 programas", "Dados dispersos sem API unificada: PNAE (FNDE — todos os municípios aderem), PAA (MDS/SAGI), Bolsa Verde (MMA/ICMBio), Brasil Sem Fome (adesão municipal). Coleta documental por estado/município (772 municípios da AL) — sem automação viável no momento."),
    ("I2.2.1 Mortalidade evitável (LBE completa 0-74)", "TabNet só tem a tabela de evitáveis <5 anos (incluída como referência). LBE completa: microdados SIM no FTP do DATASUS (dissemin/publicos/SIM) ou solicitação ao MS."),
    ("I2.2.2 Cobertura APS (% oficial)", "e-Gestor (relatorioaps.saude.gov.br) é aplicação autenticada; sem API pública. Pedir extração por UF ao MS. Nº de equipes eSF/eAP do CNES incluído como referência."),
    ("I2.3.1 IDEB", "Testadas 12+ abordagens (download.inep.gov.br em 8 padrões, INEP Data, Qedu, Power BI, busca gov.br) — todas bloqueadas/sem API. O portal ideb.inep.gov.br redireciona para Power BI (exportação manual possível). Obtenção: baixar a planilha de divulgação do IDEB 2025 na página de Resultados do INEP (gov.br/inep) ou solicitar via SIC ao INEP."),
    ("I2.4.2 Segurança cidadã + Planos de Adaptação", "PRONASCI 2: página institucional do MJ sem lista pública de municípios aderentes (verificar portarias de habilitação por UF). Planos de Adaptação municipais: sem base consolidada pública — verificar AdaptaBrasil (API protegida) e prefeituras."),
    ("I2.5.1 Renda da sociobioeconomia", "Plataforma da Ecosociobiodiversidade do MMA em construção (confirmado na página de bioeconomia do MMA) — sem dado público. Alternativa futura: IBGE PEVS (valor da extração vegetal por UF) quando a SIDRA estiver disponível."),
]
r = 3
for titulo, texto in anot:
    ws.cell(row=r, column=1, value=titulo).font = Font(bold=True, size=10)
    ws.cell(row=r + 1, column=1, value=texto).alignment = WRAP
    r += 3
largura(ws, [130])

os.makedirs(os.path.join(PASTA, "entregaveis"), exist_ok=True)
out = os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo2_Amazonia2050.xlsx")
wb.save(out)
print("Workbook salvo:", out)
print("\nTelessaúde por UF:", telessaude)
print("Total AL:", tot)
print("\nCVLI 2025 por UF (AL):", {uf: cvli.get(uf, {}).get(2025) for uf in UFS})
