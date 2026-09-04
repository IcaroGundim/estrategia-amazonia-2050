#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Monta 'Indicadores_Resultado_Eixo3_Amazonia2050.xlsx' — Eixo 3: Desenvolvimento
econômico sustentável (Estratégia Amazônia 2050), 9 estados da Amazônia Legal."""
import csv
import os
import re
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
DADOS = os.path.join(PASTA, "dados")
EIXO3 = os.path.join(DADOS, "eixo3")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
NOMES_UF = {"AC": "Acre", "AP": "Amapá", "AM": "Amazonas", "MA": "Maranhão",
            "MT": "Mato Grosso", "PA": "Pará", "RO": "Rondônia", "RR": "Roraima", "TO": "Tocantins"}

# ---------- PEVS: total por UF/ano (I3.1.1) ----------
pevs = {}  # uf -> ano -> valor_mil_rs
pevs_qt = {}
with open(os.path.join(EIXO3, "pevs_total_uf_ano.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        pevs.setdefault(r["uf"], {})[int(r["ano"])] = float(r["valor_mil_rs"]) if r["valor_mil_rs"] else None
        pevs_qt.setdefault(r["uf"], {})[int(r["ano"])] = float(r["quantidade_t"]) if r["quantidade_t"] else None

# ---------- PEVS: madeireiros x não-madeireiros (2024) ----------
pevs_mad = defaultdict(dict)  # uf -> grupo -> valor
with open(os.path.join(EIXO3, "pevs_madeireiro_uf_ano.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        if int(r["ano"]) == 2024:
            pevs_mad[r["uf"]][r["grupo"]] = float(r["valor_mil_rs"])

# ---------- PEVS: produtos 2024 (somente subprodutos — grupos são somatório dos filhos) ----------
FOLHA = re.compile(r"^\d+\.\d+ ")
pevs_prod = []
with open(os.path.join(EIXO3, "pevs_por_produto_uf.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        if int(r["ano"]) == 2024 and r["valor_mil_rs"] and FOLHA.match(r["produto"]):
            pevs_prod.append((r["uf"], r["cod_produto"], r["produto"], float(r["valor_mil_rs"]), float(r["quantidade_t"]) if r["quantidade_t"] else None))
pevs_prod.sort(key=lambda t: (t[0], -t[3]))

# ---------- RAIS: UF/ano (F3.2) ----------
rais_uf = {}  # uf -> ano -> estabelecimentos
with open(os.path.join(EIXO3, "rais_estab_uf_ano.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        rais_uf.setdefault(r["uf"], {})[int(r["ano"])] = int(r["estabelecimentos_ativos"])

# ---------- RAIS: UF x divisão (F3.2) ----------
rais_div = defaultdict(lambda: defaultdict(dict))  # uf -> divisao -> ano -> (estab, vinc)
with open(os.path.join(EIXO3, "rais_estab_uf_divisao_ano.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        uf, div, ano = r["uf"], r["divisao_cnae"], int(r["ano"])
        rais_div[uf][div][ano] = (int(r["estabelecimentos_ativos"]), int(r["vinculos_ativos"]))

# ---------- CNAE: nomes das divisões ----------
cnae_nomes, cnae_secoes = {}, {}
with open(os.path.join(EIXO3, "cnae_divisoes.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        cnae_nomes[r["divisao"]] = r["descricao"]
        cnae_secoes[r["divisao"]] = f'{r["secao"]} - {r["secao_descricao"]}'

# ---------- PIA: série por UF/ano (F3.5) ----------
pia = {}  # uf -> ano -> dict variáveis
with open(os.path.join(EIXO3, "pia_industria_uf_ano.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        uf, ano = r["uf"], int(r["ano"])
        pia.setdefault(uf, {})[ano] = {
            "vti": float(r["valor_transf_ind_mil_rs"]) if r["valor_transf_ind_mil_rs"] else None,
            "po": float(r["pessoal_ocupado_3112"]) if r["pessoal_ocupado_3112"] else None,
            "ul": float(r["unidades_locais"]) if r["unidades_locais"] else None,
            "empresas": float(r["empresas"]) if r["empresas"] else None,
        }

# ---------- Catálogo Eixo 3 ----------
CATALOGO = [
    ("I3.1.1", "3.1", "Valor da produção da sociobioeconomia",
     "Aumentar em 100 milhões de reais o valor da produção das atividades da sociobioeconomia",
     "Valor da produção das atividades da sociobioeconomia (MMA)",
     "R$", "MMA (Plataforma da Ecosociobiodiversidade, em construção); proxy: IBGE PEVS (SIDRA 289)", 2050,
     "coletado (proxy PEVS/IBGE: valor da produção da extração vegetal, 2015-2024)"),
    ("I3.2.1", "3.2", "Cadeias produtivas estruturadas",
     "Estruturação de pelo menos 02 cadeias produtivas sustentáveis em todos os estados da AL",
     "Número de cadeias estruturadas que atendam a critérios de sociobioeconomia",
     "nº de cadeias", "Estados", 2050, "pendente (documental — critérios de cadeia não definidos publicamente)"),
    ("I3.3.1", "3.3", "Política de trabalho verde",
     "Estabelecimento de políticas estaduais de promoção do trabalho verde",
     "Número e percentual de UFs com política de trabalho verde regulamentada, orçamento definido e ações em execução",
     "nº / %", "Estados", 2050, "pendente (documental)"),
    ("I3.3.2", "3.3", "Pessoas capacitadas em sociobioeconomia",
     "Capacitação de pelo menos 10 mil pessoas em atividades ligadas à sociobioeconomia",
     "Colaboradores capacitados; Cidadãos capacitados",
     "pessoas", "Estados", 2050, "pendente (documental)"),
    ("I3.3.3", "3.3", "Programas permanentes de formação técnica",
     "Institucionalização de programas permanentes de formação técnica regional",
     "Número de UFs com programa permanente formalizado e oferta regular",
     "nº de UFs", "Estados", 2050, "pendente (documental)"),
    ("I3.4.1", "3.4", "Beneficiários de PSA",
     "Aumento do número de beneficiários de PSA na Amazônia Legal em 100%",
     "Número de beneficiários ativos de programas de PSA",
     "nº de beneficiários", "Estados (devem possuir esta informação)", 2050, "pendente (documental)"),
    ("I3.4.2", "3.4", "Operações de PSA e mercados de serviços ecossistêmicos",
     "Consolidação de mercados regionais de serviços ecossistêmicos financeiramente sustentáveis",
     "Número e valor de operações de PSA, créditos ambientais, fundos verdes e outros instrumentos de financiamento sustentável",
     "nº / R$", "Estados (devem possuir esta informação)", 2050, "pendente (documental)"),
    ("I3.5.1", "3.5", "Política de agregação de valor mineral",
     "Implementação de políticas de agregação de valor mineral em 100% dos estados mineradores",
     "Percentual de estados com política formalizada, regulamentada, financiada e em execução",
     "%", "Estados", 2050, "pendente (documental)"),
    ("I3.5.2", "3.5", "Projetos minerais com critérios socioambientais",
     "100% dos projetos minerais estratégicos estaduais com critérios socioambientais atendidos",
     "Percentual de projetos estratégicos avaliados e conformes com uma matriz padronizada de critérios socioambientais",
     "%", "Estados", 2050, "pendente (documental — matriz de critérios a definir)"),
    ("F3.2", "3.2 (ficha)", "Empregos e empresas nas cadeias produtivas",
     "Estruturação de cadeias produtivas sustentáveis (ficha técnica: aumento do valor da produção, do nº de empregos e do nº de empresas)",
     "Aumento do valor da produção, do número de empregos e do número de empresas nas cadeias produtivas definidas pelos estados",
     "nº de vínculos", "MMA; proxy: MTE (RAIS Estabelecimentos 2023-2024)", 2050,
     "coletado (proxy RAIS: vínculos ativos e estabelecimentos ativos por UF x divisão CNAE, 2023-2024)"),
    ("F3.5", "3.5 (ficha)", "Valor da transformação industrial",
     "Aumento de 10% ao ano do valor agregado de produtos de transformação industrial",
     "Valor da renda líquida total de empresas de transformação industrial (ficha: Valor da transformação industrial)",
     "R$ (mil)", "IBGE (PIA-Empresa, SIDRA 1849/10457)", 2050,
     "coletado (PIA-Empresa: valor da transformação industrial por UF, 2015-2024)"),
]

# ---------- Estilo ----------
HDR_FILL = PatternFill("solid", fgColor="1B4332")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13, color="1B4332")
NOTE_FONT = Font(italic=True, size=9, color="555555")
BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")


def estilo_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER


def largura(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def bordas(ws, ncols, ini=2):
    for r in range(ini, ws.max_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = BORDER


def fmt(v, dec=0):
    if v is None:
        return "—"
    return f"{v:,.{dec}f}".replace(",", "X").replace(".", ",").replace("X", ".")


wb = Workbook()

# ---------- Sobre ----------
ws = wb.active
ws.title = "Sobre"
ws["A1"] = "ESTRATÉGIA REGIONAL AMAZÔNIA 2050 — Eixo 3: Desenvolvimento econômico sustentável"
ws["A1"].font = TITLE_FONT
ws["A3"] = "Indicadores de RESULTADO, organizados para os 9 estados da Amazônia Legal."
ws["A4"] = "Base: '3. Matriz Metas x Indicadores .xlsx' (aba atualizada, coluna INDICADOR DE RESULTADO) e fichas técnicas."
ws["A6"] = "Data de coleta: 30/08/2026."
ws["A7"] = "Status: coletado = dado oficial obtido via API/FTP; parcial = parte do indicador; pendente = fonte sem automação viável (coleta documental/estadual)."
ws["A9"] = "FONTES E LIMITAÇÕES"
ws["A9"].font = Font(bold=True)
notas = [
    "• Sociobioeconomia (I3.1.1): a fonte da matriz (Plataforma da Ecosociobiodiversidade do MMA) segue em construção. Proxy coletada: IBGE PEVS — Produção da Extração Vegetal (SIDRA tabela 289), valor da produção (R$ mil correntes) e quantidade (t), série 2015-2024 por UF e por produto. A proxy NÃO cobre toda a sociobioeconomia (exclui agricultura familiar não extrativista, pesca e manejo de fauna).",
    "• Empregos e empresas (F3.2): proxy coletada do MTE — RAIS Estabelecimentos 2023 e 2024 (FTP ftp.mtps.gov.br), agregação por UF x divisão CNAE (2 dígitos) com 'Qtd Vínculos Ativos' (estoque em 31/12) e nº de estabelecimentos com 'Ind Atividade Ano = 1'. A definição oficial das cadeias por estado é documental.",
    "• Transformação industrial (F3.5): IBGE PIA-Empresa — série antiga tabela 1849 (2007-2023, empresas com 5+ pessoas) e série nova tabela 10457 (2024). Variáveis: valor da transformação industrial, pessoal ocupado, unidades locais, empresas. Valores em R$ mil correntes. ATENÇÃO: a variação 2023→2024 envolve mudança de série (antiga→nova); 2024 = linha de base oficial.",
    "• Indicadores I3.3.1-I3.5.2 (matriz): informação jurisdicional dos estados (políticas, capacitados, PSA, mineração) — sem base pública consolidada; caminho de obtenção na aba 'Pendentes_anotacoes'.",
    "• Fichas técnicas também listam indicadores 3.3 (capacitação) e 3.4 (crédito de carbono jurisdicional) — todos documentais.",
]
r = 10
for n in notas:
    ws.cell(row=r, column=1, value=n).alignment = WRAP
    ws.row_dimensions[r].height = 45
    r += 1
ws.cell(row=r + 1, column=1, value="Scripts: scripts/eixo3_pevs.py, eixo3_pia.py, eixo3_rais.py. Dados brutos: dados/eixo3/.").font = NOTE_FONT
largura(ws, [150])

# ---------- Catálogo ----------
ws = wb.create_sheet("Catalogo_Indicadores")
headers = ["Código", "Linha de Ação", "Indicador (nome curto)", "Meta", "Indicador de RESULTADO (texto da matriz/ficha)",
           "Unidade", "Fonte", "Prazo", "Status da coleta"]
ws.append(headers)
for row in CATALOGO:
    ws.append(list(row))
estilo_header(ws, len(headers))
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
    for cell in row:
        cell.border = BORDER
        cell.alignment = WRAP
largura(ws, [9, 12, 30, 38, 52, 14, 34, 8, 38])

# ---------- Matriz_UF ----------
ws = wb.create_sheet("Matriz_UF")
headers = ["Código", "Indicador (resumo)", "Unidade", "Ano ref.", "AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO", "Status", "Fonte"]
ws.append(headers)


def pevs_2024(uf):
    return pevs.get(uf, {}).get(2024)


def rais_vinc(uf, ano):
    return sum(v[1] for d in rais_div.get(uf, {}).values() for k, v in d.items() if k == ano)


def pia_vti(uf, ano):
    d = pia.get(uf, {}).get(ano)
    return d["vti"] if d else None


for cod, la, curto, meta, indicador, unid, fonte, prazo, status in CATALOGO:
    if cod == "I3.1.1":
        vals = [pevs_2024(uf) for uf in UFS]
        vals = [fmt(v) if v is not None else "—" for v in vals]
        anoref, status_, unid2 = "2024 (PEVS)", "coletado (proxy PEVS — valor extração vegetal)", "R$ mil (PEVS)"
    elif cod == "F3.2":
        vals = [fmt(rais_vinc(uf, 2024)) for uf in UFS]
        anoref, status_, unid2 = "2024 (RAIS)", "coletado (proxy RAIS — vínculos ativos 31/12)", "nº vínculos"
    elif cod == "F3.5":
        vals = [fmt(pia_vti(uf, 2024)) for uf in UFS]
        anoref, status_, unid2 = "2024 (PIA)", "coletado (PIA-Empresa — valor da transformação industrial)", "R$ mil"
    else:
        vals = ["—"] * 9
        anoref, status_, unid2 = "—", status, unid
    ws.append([cod, curto, unid2, anoref] + vals + [status_, fonte])
estilo_header(ws, len(headers))
bordas(ws, len(headers))
largura(ws, [9, 34, 16, 14, 12, 12, 12, 12, 12, 12, 12, 12, 12, 34, 30])
ws.freeze_panes = "E2"

# ---------- 3.1.1 PEVS série ----------
ws = wb.create_sheet("3.1.1_Pevs_serie")
headers = ["UF", "Estado"] + [str(a) for a in range(2015, 2025)] + ["Madeireiros 2024 (R$ mil)", "Não-madeireiros 2024 (R$ mil)"]
ws.append(headers)
tot_anos = {a: 0.0 for a in range(2015, 2025)}
tot_mad = tot_nmad = 0.0
for uf in UFS:
    row = [uf, NOMES_UF[uf]]
    for a in range(2015, 2025):
        v = pevs.get(uf, {}).get(a)
        if v is not None:
            tot_anos[a] += v
        row.append(fmt(v) if v is not None else "—")
    mad = pevs_mad.get(uf, {}).get("madeireiro")
    nmad = pevs_mad.get(uf, {}).get("nao_madeireiro")
    if mad is not None:
        tot_mad += mad
    if nmad is not None:
        tot_nmad += nmad
    row += [fmt(mad) if mad is not None else "—", fmt(nmad) if nmad is not None else "—"]
    ws.append(row)
ws.append(["TOTAL", "Amazônia Legal"] + [fmt(tot_anos[a]) for a in range(2015, 2025)]
          + [fmt(tot_mad) if tot_mad else "—", fmt(tot_nmad) if tot_nmad else "—"])
estilo_header(ws, len(headers))
bordas(ws, len(headers))
largura(ws, [6, 16] + [11] * 10 + [20, 22])
ws.append([])
ws.cell(row=ws.max_row, column=1,
        value="Fonte: IBGE — PEVS, Produção da Extração Vegetal (SIDRA tabela 289; API apisidra.ibge.gov.br), valor da produção em R$ mil correntes. Proxy do indicador I3.1.1 (valor da produção das atividades da sociobioeconomia); o painel oficial do MMA (Ecosociobiodiversidade) está em construção. Madeireiros = grupos 7 (carvão, lenha, madeira em tora) e 9 (pinheiro — irrelevante na AL); demais grupos = não-madeireiros. Quantidades (t) por produto na aba 3.1.1_Pevs_produtos_2024 — o IBGE não publica o total de quantidade (soma heterogênea).").font = NOTE_FONT

# ---------- 3.1.1 PEVS produtos 2024 ----------
ws = wb.create_sheet("3.1.1_Pevs_produtos_2024")
headers = ["UF", "Estado", "Cód. produto", "Produto", "Valor 2024 (R$ mil)", "Participação na UF (%)", "Quantidade 2024 (t)"]
ws.append(headers)
por_uf = defaultdict(float)
for uf, _, _, v, _ in pevs_prod:
    por_uf[uf] += v
for uf, cod_p, prod, v, qt in pevs_prod:
    part = v / por_uf[uf] * 100 if por_uf[uf] else None
    ws.append([uf, NOMES_UF[uf], cod_p, prod, fmt(v), fmt(part, 1) if part else "—", fmt(qt) if qt else "—"])
estilo_header(ws, len(headers))
bordas(ws, len(headers))
largura(ws, [6, 14, 12, 42, 18, 20, 18])
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:G{ws.max_row}"
ws.append([])
ws.cell(row=ws.max_row, column=1,
        value="Fonte: IBGE — PEVS, Produção da Extração Vegetal (SIDRA tabela 289; API apisidra.ibge.gov.br), ano de 2024. Somente SUBPRODUTOS (categorias folha da classificação); os grupos (ex.: '1 - Alimentícios') são somatório dos filhos e foram omitidos para evitar dupla contagem — a soma por UF bate com o Total da SIDRA. 'Participação na UF' = share do valor do subproduto na soma dos subprodutos da UF. Valor em R$ mil correntes.").font = NOTE_FONT

# ---------- F3.2 RAIS empregos ----------
ws = wb.create_sheet("F3.2_Rais_empregos")
headers = ["UF", "Estado", "Estab. ativos 2023", "Estab. ativos 2024", "Vínculos ativos 31/12/2023", "Vínculos ativos 31/12/2024", "Variação vínculos (%)"]
ws.append(headers)
tv = {"23": 0, "24": 0}
te = {"23": 0, "24": 0}
for uf in UFS:
    e23 = rais_uf.get(uf, {}).get(2023)
    e24 = rais_uf.get(uf, {}).get(2024)
    v23 = rais_vinc(uf, 2023)
    v24 = rais_vinc(uf, 2024)
    var = (v24 / v23 - 1) * 100 if v23 else None
    te["23"] += e23 or 0
    te["24"] += e24 or 0
    tv["23"] += v23
    tv["24"] += v24
    ws.append([uf, NOMES_UF[uf], fmt(e23), fmt(e24), fmt(v23), fmt(v24), fmt(var, 1) if var is not None else "—"])
ws.append(["TOTAL", "Amazônia Legal", fmt(te["23"]), fmt(te["24"]), fmt(tv["23"]), fmt(tv["24"]),
           fmt((tv["24"] / tv["23"] - 1) * 100, 1)])
estilo_header(ws, len(headers))
bordas(ws, len(headers))
largura(ws, [6, 16, 15, 15, 22, 22, 18])
ws.append([])
ws.cell(row=ws.max_row, column=1,
        value="Fonte: MTE — RAIS Estabelecimentos 2023 e 2024 (FTP ftp.mtps.gov.br/pdet/microdados/RAIS). Proxy da ficha do indicador 3.2: número de empregos formais (vínculos ativos em 31/12) e empresas (estabelecimentos com Ind Atividade Ano = 1) por UF. As cadeias produtivas definidas por estado são documentais (aba Pendentes_anotacoes).").font = NOTE_FONT

# ---------- F3.2 RAIS divisões ----------
ws = wb.create_sheet("F3.2_Rais_divisoes")
headers = ["UF", "Estado", "Divisão CNAE", "Descrição da divisão", "Seção CNAE", "Vínculos 2023", "Vínculos 2024", "Variação (%)", "Estab. 2024"]
ws.append(headers)
rows_div = []
for uf in UFS:
    for div, anos in rais_div.get(uf, {}).items():
        v23 = anos.get(2023, (0, 0))[1]
        v24 = anos.get(2024, (0, 0))[1]
        e24 = anos.get(2024, (0, 0))[0]
        var = (v24 / v23 - 1) * 100 if v23 else None
        rows_div.append([uf, NOMES_UF[uf], div, cnae_nomes.get(div, ""), cnae_secoes.get(div, ""),
                         fmt(v23), fmt(v24), fmt(var, 1) if var is not None else "—", fmt(e24)])
rows_div.sort(key=lambda r: (r[0], -(int(r[6].replace(".", "")) if r[6] != "—" else 0)))
for row in rows_div:
    ws.append(row)
estilo_header(ws, len(headers))
bordas(ws, len(headers))
largura(ws, [6, 14, 12, 46, 34, 14, 14, 11, 13])
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:I{ws.max_row}"

# ---------- F3.5 PIA transformação ----------
ws = wb.create_sheet("F3.5_Pia_transformacao")
headers = ["UF", "Estado"] + [str(a) for a in range(2015, 2025)] + ["Var. 2023→24 (%)", "Var. média anual 2015→24 (%)", "Pessoal ocupado 2024", "Empresas 2024", "Unid. locais 2024"]
ws.append(headers)
tot_pia = {a: 0.0 for a in range(2015, 2025)}
for uf in UFS:
    row = [uf, NOMES_UF[uf]]
    for a in range(2015, 2025):
        v = pia_vti(uf, a)
        if v is not None:
            tot_pia[a] += v
        row.append(fmt(v) if v is not None else "—")
    v15 = pia_vti(uf, 2015)
    v23 = pia_vti(uf, 2023)
    v24 = pia_vti(uf, 2024)
    var24 = (v24 / v23 - 1) * 100 if (v23 and v24) else None
    cagr = ((v24 / v15) ** (1 / 9) - 1) * 100 if (v15 and v24) else None
    d24 = pia.get(uf, {}).get(2024, {})
    row += [fmt(var24, 1) if var24 is not None else "—", fmt(cagr, 1) if cagr is not None else "—",
            fmt(d24.get("po")) if d24.get("po") is not None else "—",
            fmt(d24.get("empresas")) if d24.get("empresas") is not None else "—",
            fmt(d24.get("ul")) if d24.get("ul") is not None else "—"]
    ws.append(row)
row = ["TOTAL", "Amazônia Legal"] + [fmt(tot_pia[a]) for a in range(2015, 2025)] + ["—", "—", "—", "—", "—"]
ws.append(row)
estilo_header(ws, len(headers))
bordas(ws, len(headers))
largura(ws, [6, 16] + [13] * 10 + [15, 22, 18, 14, 14])
ws.append([])
ws.cell(row=ws.max_row, column=1,
        value="Fonte: IBGE — PIA-Empresa: tabela 1849 (série antiga, 2015-2023) e tabela 10457 (série nova, 2024), empresas com 5+ pessoas, Total das divisões CNAE. Valores em R$ mil correntes. 'Var. média anual' = CAGR 2015→2024. ATENÇÃO: quebra de série entre 2023 (tabela 1849) e 2024 (tabela 10457); a ficha do indicador 3.5 cita a tabela 10457 como fonte oficial. Indicador da ficha: aumento de 10% ao ano do valor agregado. A série usa a categoria Total da SIDRA — em nível de DIVISÃO (dados/eixo3/pia_divisoes_uf_ano.csv) parte dos valores é suprimida por sigilo ('X'), e a soma das divisões fica abaixo do Total.").font = NOTE_FONT

# ---------- Pendentes ----------
ws = wb.create_sheet("Pendentes_anotacoes")
ws["A1"] = "Indicadores do Eixo 3 sem dado automatizado (coleta 30/08/2026)"
ws["A1"].font = TITLE_FONT
anot = [
    ("I3.1.1 Valor da produção da sociobioeconomia (fonte oficial MMA)",
     "A Plataforma da Ecosociobiodiversidade (MMA) segue em construção. Proxy PEVS/IBGE incluída neste workbook. Quando a plataforma abrir, extrair a renda da sociobioeconomia por UF e substituir a proxy. Alternativa complementar: SIDRA PAM/PEAM (produção agrícola extrativa não florestal) e RUMEA/IBGE (pesca)."),
    ("I3.2.1 Cadeias produtivas estruturadas (matriz)",
     "Requer a lista de cadeias definidas por cada estado e critérios de 'cadeia estruturada' — levantamento documental junto às secretarias estaduais (SEMEC/SEDUC, SDI etc.). A proxy RAIS (aba F3.2_Rais_divisoes) permite acompanhar empregos e empresas por divisão CNAE em qualquer cadeia escolhida."),
    ("I3.3.1 Política estadual de trabalho verde",
     "Informação jurisdicional: verificar leis/decretos estaduais sobre trabalho verde, emprego verde e transição justa nos portais legislativos e Diários Oficiais dos 9 estados; registrar orçamento e ações em execução."),
    ("I3.3.2 Pessoas capacitadas (10 mil) e I3.3.3 programas de formação",
     "Fontes possíveis: secretarias estaduais de qualificação (SETAS/SINE), SENAI/SENAC/Sensai amazônicos, e programas como o PQGA (Plano de Qualificação dos Guardiões da Amazônia? verificar nomenclatura oficial). Levantamento documental por estado; sistemas de qualificação (SINE) sem API pública."),
    ("I3.4.1 Beneficiários de PSA e I3.4.2 operações de PSA/créditos",
     "Os 9 estados devem fornecer (ficha: 'Estados devem possuir esta informação'). Bases auxiliares possíveis: programas estaduais de PSA (REDD+ estaduais, leis de PSI), Bolsa Verde (ICMBio/MMA), Floresta+ (IBAMA), SISRAP/APIS. Sem API pública consolidada — coleta documental/contato com os estados."),
    ("I3.5.1 Política estadual de agregação de valor mineral (matriz)",
     "Documental: verificar políticas estaduais de mineração (PEM/Planos Estaduais de Mineração, agendas minerais) nos 9 estados; 'PNAD' citado como fonte na matriz é provável erro tipográfico."),
    ("I3.5.2 Projetos minerais estratégicos conformes",
     "Requer matriz padronizada de critérios socioambientais (a definir junto aos estados) e lista de projetos estratégicos. Bases auxiliares: SIGMINE/ANM (projetos ativos por UF), cadastro estadual de mineradoras. Sem automação até a definição da matriz."),
    ("CAGED (fluxo mensal de empregos) — não usado",
     "A API antiga apidatalake.mte.gov.br está fora do ar (DNS inexistente em 30/08/2026). Os microdados mensais do Novo CAGED estão no mesmo FTP do MTE (pdet/microdados/NOVO CAGED) e podem ser agregados depois se for necessário um fluxo (saldo) além do estoque RAIS."),
]
r = 3
for titulo, texto in anot:
    ws.cell(row=r, column=1, value=titulo).font = Font(bold=True, size=10)
    ws.cell(row=r + 1, column=1, value=texto).alignment = WRAP
    ws.row_dimensions[r + 1].height = 60
    r += 3
largura(ws, [140])

os.makedirs(os.path.join(PASTA, "entregaveis"), exist_ok=True)
out = os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo3_Amazonia2050.xlsx")
wb.save(out)
print("Workbook salvo:", out)

# ---------- verificação rápida ----------
print("\nPEVS 2024 TOTAL AL:", fmt(sum(pevs.get(uf, {}).get(2024) or 0 for uf in UFS)), "mil R$")
print("RAIS vínculos 2024:", fmt(sum(rais_vinc(uf, 2024) for uf in UFS)))
print("PIA VTI 2024 AL:", fmt(sum(pia_vti(uf, 2024) or 0 for uf in UFS)), "mil R$")
