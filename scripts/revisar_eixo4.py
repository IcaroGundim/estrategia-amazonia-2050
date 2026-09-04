#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Revisão independente do Eixo 4 — validações externas (v2, pós-correções).

Correções v2 (achados da 1ª revisão):
- Revisão v1 lia a coluna 8 (ponderado 2024) achando que era 2025 → bug da revisão
- Tolerância PER (rounding 1,55→1,6) → tolerância 0,06
- SIDRA n3 pode devolver '-' transitório → retry
- Domínio MUNIC real: Sim/Não/'-' (não aplicável)/'Recusa'/'Não informou' — a regra da ficha
  (só 'Sim' = 1,0; qualquer outra = penalidade) é a implementada → check ajustado
- Novos spot-checks: município individual (Porto Velho) vs SIDRA direto;
  fatores MUNIC de 2 municípios conferidos na planilha; sensibilidade da definição
  de água adequada (sem a categoria 72145); contagem de fallback do dom_total.
"""
import csv, json, os, time, urllib.request
from collections import Counter, defaultdict

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
DADOS = os.path.join(PASTA, "dados")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
problemas = []


def check(nome, cond, detalhe=""):
    print(("[OK]  " if cond else "[FALHA] ") + nome + (" — " + detalhe if detalhe else ""))
    if not cond:
        problemas.append(nome)


def sidra(url, tentativas=3):
    for i in range(tentativas):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=300) as r:
                return json.loads(r.read().decode("utf-8"))
        except Exception:
            if i == tentativas - 1:
                raise
            time.sleep(3)


# ============ 1. POPULAÇÃO ============
print("===== 1. POPULAÇÃO MUNICIPAL (Censo 2022) =====")
soma_mun = Counter()
for r in csv.DictReader(open(os.path.join(DADOS, "ibge_pop", "pop_mun_censo2022.csv"), encoding="utf-8")):
    soma_mun[r["cod_ibge"][:2]] += int(r["populacao"])
d3 = sidra("https://apisidra.ibge.gov.br/values/t/4709/n3/all/v/93/p/2022?formato=json")
sidra_uf = {x["D1C"]: int(x["V"]) for x in d3[1:]}
for cod2, uf in {"11": "RO", "12": "AC", "13": "AM", "14": "RR", "15": "PA", "16": "AP", "17": "TO", "21": "MA", "51": "MT"}.items():
    check(f"pop {uf}: soma municipal = SIDRA estadual", soma_mun[cod2] == sidra_uf[cod2],
          f"{soma_mun[cod2]:,} vs {sidra_uf[cod2]:,}")

# ============ 2. ÁGUA/ESGOTO ============
print("\n===== 2. ÁGUA/ESGOTO (Censo 2022) — recálculo independente =====")
cats_a = "72129,72144,72145,72154"
cats_e = "46292,46290,72112"
SET_UF = {"11", "12", "13", "14", "15", "16", "17", "21", "51"}


def para_map(rows, idx):
    m = {}
    for x in rows[1:]:
        if x["V"] not in ("-", ""):
            m.setdefault(x[idx], {})[x["D4C"]] = int(x["V"])
    return m


n1_a = para_map(sidra(f"https://apisidra.ibge.gov.br/values/t/6803/n1/all/v/381/p/2022/c1821/{cats_a}?formato=json"), "D1C")["1"]
n1_e = para_map(sidra(f"https://apisidra.ibge.gov.br/values/t/6805/n1/all/v/381/p/2022/c11558/{cats_e}?formato=json"), "D1C")["1"]
pct_agua_br = (n1_a["72144"] + n1_a["72145"] + n1_a["72154"]) / n1_a["72129"] * 100
pct_esg_br = (n1_e["46290"] + n1_e["72112"]) / n1_e["46292"] * 100
pct_agua_br_strict = (n1_a["72144"] + n1_a["72154"]) / n1_a["72129"] * 100
print(f"Brasil (SIDRA n1): água 'adequada' (ficha) {pct_agua_br:.1f}% | variante estrita {pct_agua_br_strict:.1f}% | esgoto adequado {pct_esg_br:.1f}%")
# IBGE publicou para o Censo 2022: esgoto adequado (rede+fossa séptica) ≈ 77%; ligação à rede de água ≈ 84,6%
check("esgoto adequado Brasil ≈ 77% (coerente com divulgação do Censo 2022)", 75 <= pct_esg_br <= 80, f"{pct_esg_br:.1f}%")
check("água 'adequada' Brasil (ficha, ampla) ≥ ligação à rede publicada (~84,6%)", 84 <= pct_agua_br <= 96, f"{pct_agua_br:.1f}%")

n3_a = para_map(sidra(f"https://apisidra.ibge.gov.br/values/t/6803/n3/all/v/381/p/2022/c1821/{cats_a}?formato=json"), "D1C")
n3_e = para_map(sidra(f"https://apisidra.ibge.gov.br/values/t/6805/n3/all/v/381/p/2022/c11558/{cats_e}?formato=json"), "D1C")
cod_uf2 = {"11": "RO", "12": "AC", "13": "AM", "14": "RR", "15": "PA", "16": "AP", "17": "TO", "21": "MA", "51": "MT"}
a_mun = Counter(); e_mun = Counter(); t_mun = Counter()
for r in csv.DictReader(open(os.path.join(DADOS, "saneamento", "saneamento_mun.csv"), encoding="utf-8")):
    t_mun[r["uf"]] += float(r["dom_total"])
    a_mun[r["uf"]] += float(r["agua_adeq"])
    e_mun[r["uf"]] += float(r["esgoto_adeq"])
dif = 0
for cod2, uf in cod_uf2.items():
    s_a, s_e = n3_a.get(cod2, {}), n3_e.get(cod2, {})
    sa = s_a.get("72144", 0) + s_a.get("72145", 0) + s_a.get("72154", 0)
    se = s_e.get("46290", 0) + s_e.get("72112", 0)
    st = s_a.get("72129") or s_e.get("46292") or 0
    if int(a_mun[uf]) != sa or int(e_mun[uf]) != se or int(t_mun[uf]) != st:
        dif += 1
        print(f"  DIF {uf}: agua {a_mun[uf]:.0f} vs {sa} | esgoto {e_mun[uf]:.0f} vs {se} | total {t_mun[uf]:.0f} vs {st}")
check("agregação municipal = SIDRA n3 por UF (água, esgoto, total)", dif == 0, f"{dif} UFs divergentes")

# sensibilidade: definição estrita de água (sem 72145) — UF-level
str_uf = {}
for cod2, uf in cod_uf2.items():
    s_a = n3_a.get(cod2, {})
    str_uf[uf] = (s_a.get("72144", 0) + s_a.get("72154", 0)) / s_a.get("72129") * 100
num = sum(n3_a[c]["72144"] + n3_a[c]["72154"] for c in cod_uf2)
den = sum(n3_a[c]["72129"] for c in cod_uf2)
print(f"  AL água adequada estrita (sem 72145): {num/den*100:.1f}% (usada na coleta, ampla: 86,2%)")

# fallback do dom_total: quantos municípios sem total de água usaram o de esgoto
n6_tot_a = para_map(sidra(f"https://apisidra.ibge.gov.br/values/t/6803/n6/all/v/381/p/2022/c1821/72129?formato=json"), "D1C")
cod_al = set()
for r in csv.DictReader(open(os.path.join(DADOS, "iivcm", "iivcm.csv"), encoding="utf-8")):
    cod_al.add(r["codigo_ibge"])
sem_tot_agua = [c for c in cod_al if c not in n6_tot_a or "72129" not in n6_tot_a.get(c, {})]
print(f"  municípios da AL sem total de água (72129) na SIDRA: {len(sem_tot_agua)}")
check("dom_total municipal consistente com a fonte", len(sem_tot_agua) <= 20, f"{len(sem_tot_agua)} municípios dependem do fallback")

# spot-check municipal: Porto Velho direto na SIDRA vs saneamento_mun.csv
pv = [r for r in csv.DictReader(open(os.path.join(DADOS, "saneamento", "saneamento_mun.csv"), encoding="utf-8")) if r["cod_ibge"] == "1100205"][0]
d6a = para_map(sidra("https://apisidra.ibge.gov.br/values/t/6803/n6/1100205/v/381/p/2022/c1821/" + cats_a + "?formato=json"), "D1C")["1100205"]
d6e = para_map(sidra("https://apisidra.ibge.gov.br/values/t/6805/n6/1100205/v/381/p/2022/c11558/" + cats_e + "?formato=json"), "D1C")["1100205"]
sa = d6a["72144"] + d6a["72145"] + d6a["72154"]
se = d6e["46290"] + d6e["72112"]
check("spot-check Porto Velho: água/esgoto/total = SIDRA municipal direto",
      int(pv["agua_adeq"]) == sa and int(pv["esgoto_adeq"]) == se and int(float(pv["dom_total"])) == d6a["72129"],
      f"água {pv['agua_adeq']} vs {sa}; esgoto {pv['esgoto_adeq']} vs {se}")

# ============ 3. MUNIC ============
print("\n===== 3. MUNIC 2024 — domínio e spot-check dos fatores =====")
import openpyxl
MUNIC = os.path.join(DADOS, "ibge_munic", "Base_MUNIC_2024_20251107.xlsx")
COLS = {"Informática e comunicação": "Mtic266", "Governanca": "Mgov086",
        "Habitacao": "Mhab088", "Agropecuária": "Magr18"}
wb = openpyxl.load_workbook(MUNIC, read_only=True)
vals = {v: Counter() for v in COLS.values()}
spot = {"1100205": {}, "1400100": {}}  # Porto Velho, Boa Vista
for aba, alvo in COLS.items():
    ws = wb[aba]
    idx = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        row = [str(c).strip() if c is not None else "" for c in row]
        idx = row.index(alvo)
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = str(row[0]).strip()
        if cod in cod_al:
            raw = row[idx] if idx < len(row) else None
            vals[alvo][str(raw)] += 1
            if cod in spot:
                spot[cod][alvo] = str(raw)
wb.close()
for v, c in vals.items():
    top = ", ".join(f"{k}: {n}" for k, n in c.most_common(5))
    print(f"  {v}: {top}")
    check(f"{v}: valores são rótulos MUNIC (Sim/Não/-/Recusa/Não informou)",
          set(c) <= {"Sim", "Não", "-", "Recusa", "Não informou", "None"}, top)
dash_mhab = vals["Mhab088"].get("-", 0)
print(f"  (Mhab088 '-' = não aplicável em {dash_mhab} dos 808 municípios — tratado como branco pela regra da ficha: penalidade 0,9)")

# fatores dos 2 municípios spot
mun_csv = {r["cod_ibge"]: r for r in csv.DictReader(open(os.path.join(DADOS, "saneamento", "saneamento_mun.csv"), encoding="utf-8"))}
for cod, raws in spot.items():
    fclima = 1.0 if any(raws.get(v) == "Sim" for v in ("Magr18", "Mhab088", "Mtic266")) else 0.9
    fgov = 1.0 if raws.get("Mgov086") == "Sim" else 0.95
    ok_ = abs(float(mun_csv[cod]["fclima"]) - fclima) < 0.01 and abs(float(mun_csv[cod]["fgov"]) - fgov) < 0.01
    check(f"spot-check fatores {cod} ({'Porto Velho' if cod=='1100205' else 'Boa Vista'})",
          ok_, f"raw={raws} → fclima={fclima} fgov={fgov} | csv={mun_csv[cod]['fclima']},{mun_csv[cod]['fgov']}")

# ============ 4. SIGA ============
print("\n===== 4. ANEEL SIGA =====")
pot_uf = defaultdict(float)
top_pa = []
nacional = 0.0
with open(os.path.join(DADOS, "aneel", "siga.csv"), encoding="utf-8-sig") as f:
    for r in csv.DictReader(f, delimiter=";"):
        try:
            kw = float(r["MdaPotenciaFiscalizadaKw"].replace(".", "").replace(",", "."))
        except ValueError:
            continue
        if r["DscFaseUsina"] == "Operação":
            nacional += kw
            if r["SigUFPrincipal"] in UFS:
                pot_uf[r["SigUFPrincipal"]] += kw
                if r["SigUFPrincipal"] == "PA":
                    top_pa.append((kw, r["NomEmpreendimento"]))
top_pa.sort(reverse=True)
print("  Top 5 PA:", [(n[:30], f"{k/1e6:.2f} GW") for k, n in top_pa[:5]])
check("Belo Monte (~11,2 GW) e Tucuruí (~8,5 GW) no topo do PA",
      any(10_000_000 < k < 13_000_000 for k, _ in top_pa) and any(7_500_000 < k < 9_000_000 for k, _ in top_pa))
check("total nacional em operação entre 180 e 260 GW", 180e6 <= nacional <= 260e6, f"{nacional/1e6:.0f} GW")
per = {r["uf"]: r for r in csv.DictReader(open(os.path.join(DADOS, "aneel", "per_uf.csv"), encoding="utf-8"))}
dif_sig = [uf for uf in UFS if abs(pot_uf[uf] - float(per[uf]["potencia_total_kw"])) >= 1]
check("re-soma SIGA por UF = per_uf.csv", not dif_sig, f"{dif_sig}")

# ============ 5. IBC ============
print("\n===== 5. IBC ponderado (re-cálculo independente) =====")
pop = {r["cod_ibge"]: int(r["populacao"]) for r in csv.DictReader(open(os.path.join(DADOS, "ibge_pop", "pop_mun_censo2022.csv"), encoding="utf-8"))}
mun_ibc = defaultdict(dict)
uf_cod = {}
for r in csv.DictReader(open(os.path.join(DADOS, "anatel", "IBC_municipios_indicadores_originais.csv"), encoding="utf-8-sig"), delimiter=";"):
    uf_cod[r["Código Município"]] = r["UF"]
    try:
        v = float(r["IBC"].replace(",", "."))
    except ValueError:
        continue
    mun_ibc[r["Código Município"]][int(r["Ano"])] = v
sem_pop = [c for c in mun_ibc if c not in pop and uf_cod[c] in UFS]
n_dist = len({c for c in mun_ibc if uf_cod[c] in UFS})
check("808 municípios da AL no arquivo ANATEL, todos com população", not sem_pop and n_dist == 808,
      f"{n_dist} municípios, {len(sem_pop)} sem pop")
num = sum(mun_ibc[c][2025] * pop[c] for c in mun_ibc if uf_cod[c] in UFS and 2025 in mun_ibc[c] and c in pop)
den = sum(pop[c] for c in mun_ibc if uf_cod[c] in UFS and 2025 in mun_ibc[c] and c in pop)
check("re-cálculo AL 2025 = 53,52 (baseline da ficha)", abs(num / den - 53.52) < 0.01, f"{num/den:.4f}")
ibc_proc = {r["uf"]: (float(r["ibc_uf"]), float(r["ibc_ponderado_pop"]))
            for r in csv.DictReader(open(os.path.join(DADOS, "anatel", "ibc_uf_ano.csv"), encoding="utf-8")) if r["ano"] == "2025"}
dif_uf = []
for uf in UFS:
    num = sum(mun_ibc[c][2025] * pop[c] for c in mun_ibc if uf_cod[c] == uf and c in pop and 2025 in mun_ibc[c])
    den = sum(pop[c] for c in mun_ibc if uf_cod[c] == uf and c in pop and 2025 in mun_ibc[c])
    if abs(num / den - ibc_proc[uf][1]) > 0.01:
        dif_uf.append(uf)
check("re-ponderação por UF = valores do eixo4_ibc_anatel.py", not dif_uf, f"{dif_uf}")

# ============ 6. WORKBOOK vs CSVs ============
print("\n===== 6. WORKBOOK vs CSVs =====")
import openpyxl as op
wbk = op.load_workbook(os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo4_Amazonia2050.xlsx"))
ws = wbk["4.1.1_IBC_AMZ_Anatel"]
linha = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}
dif = [uf for uf in UFS if abs(float(ws.cell(linha[uf], 9).value) - ibc_proc[uf][1]) > 0.01]  # col 9 = ponderado 2025
check("aba IBC: ponderado 2025 por UF = CSV (coluna 9)", not dif, f"{dif}")
ws = wbk["4.3.2_PER_renovaveis_aneel"]
linha = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}
dif = [uf for uf in UFS if abs(float(ws.cell(linha[uf], 5).value) - float(per[uf]["per_renovaveis_pct"])) > 0.06]
check("aba PER: % por UF = CSV", not dif, f"{dif}")
ws = wbk["4.4.1_ISGR_saneamento"]
linha = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}
isgr_f = {r["uf"]: r for r in csv.DictReader(open(os.path.join(DADOS, "saneamento", "isgr_uf.csv"), encoding="utf-8"))}
dif = [uf for uf in UFS if abs(float(ws.cell(linha[uf], 7).value) - float(isgr_f[uf]["isgr_pct"])) > 0.01]
check("aba ISGR: % por UF = CSV", not dif, f"{dif}")

print("\n===== RESULTADO =====")
print("REVISÃO SEM PROBLEMAS" if not problemas else f"PROBLEMAS ({len(problemas)}): {problemas}")
