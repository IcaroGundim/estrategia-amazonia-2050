#!/usr/bin/env python3
"""Exporta o catálogo oficial de indicadores (Estratégia Amazônia 2050) dos
workbooks de entregáveis para um único JSON consumido pelo dashboard.
Consolida: catálogo (meta/fonte/prazo/status) + valores por UF + séries anuais
onde existem (PRODES, focos, CVLI), a partir das abas de detalhe de cada eixo.
"""
import json
import re
from pathlib import Path
import openpyxl

ROOT = Path(__file__).resolve().parent.parent
ENTREGAVEIS = ROOT / "entregaveis"
OUT_DIR = ROOT / "dados" / "catalogo"
OUT_DIR.mkdir(parents=True, exist_ok=True)

NAME_TO_UF = {
    "Acre": "AC", "Amapá": "AP", "Amazonas": "AM", "Maranhão": "MA",
    "Mato Grosso": "MT", "Pará": "PA", "Rondônia": "RO", "Roraima": "RR",
    "Tocantins": "TO",
}

MILHAR_RE = re.compile(r"^\d{1,3}(\.\d{3})+$")


def num(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text in ("", "—", "-", "X"):
        return None
    negative = text.startswith("(") and text.endswith(")")
    if negative:
        text = text[1:-1].strip()
    if "," in text:
        # vírgula decimal; pontos são separadores de milhar
        text = text.replace(".", "").replace(",", ".")
    elif "." in text and MILHAR_RE.match(text) and text.split(".")[0] != "0":
        # milhar brasileiro (ex.: 80.754); "0.xxx" é sempre decimal (ex.: 0.023)
        text = text.replace(".", "")
    try:
        parsed = float(text)
    except ValueError:
        return None
    return -parsed if negative else parsed


def rows_by_uf(ws):
    """Itera linhas de uma aba de detalhe e agrupa por UF usando a coluna
    'Estado' (coluna B), que é confiável mesmo quando a coluna A carrega uma
    nota de rodapé em vez da sigla (linha do Tocantins em várias abas)."""
    header = None
    out = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row
            continue
        estado = row[1]
        uf = NAME_TO_UF.get(str(estado).strip()) if estado else None
        if not uf:
            continue
        out[uf] = dict(zip(header, row))
    return out


def load_catalog(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Catalogo_Indicadores"]
    header = None
    catalog = {}
    for row in ws.iter_rows(values_only=True):
        if header is None:
            header = row
            continue
        if not row[0]:
            continue
        rec = dict(zip(header, row))
        descricao = next((rec[k] for k in rec if str(k).startswith("Indicador de RESULTADO")), None)
        catalog[rec["Código"]] = {
            "codigo": rec["Código"],
            "linhaAcao": rec["Linha de Ação"],
            "nome": rec["Indicador (nome curto)"],
            "meta": rec["Meta"],
            "descricao": descricao,
            "unidade": rec["Unidade"],
            "fonte": rec["Fonte"],
            "prazo": rec["Prazo"],
            "status": rec["Status da coleta"],
            "valores": None,
            "serieAnual": None,
            "extra": None,
            "anoRef": None,
        }
    return wb, catalog


def attach(catalog, codigo, ano_ref, valores=None, serie=None, extra=None):
    if codigo not in catalog:
        return
    catalog[codigo]["anoRef"] = ano_ref
    if valores:
        catalog[codigo]["valores"] = valores
    if serie:
        catalog[codigo]["serieAnual"] = serie
    if extra:
        catalog[codigo]["extra"] = extra


def build_eixo1():
    wb, catalog = load_catalog(ENTREGAVEIS / "Indicadores_Resultado_Eixo1_Amazonia2050.xlsx")

    ucs = rows_by_uf(wb["1.1.2_UCs_estaduais"])
    attach(catalog, "I1.1.2", "03/2026",
           valores={uf: num(r["% com ambos (meta 40%)"]) for uf, r in ucs.items()},
           extra={uf: {"total": num(r["UCs estaduais (nº)"]), "comPlano": num(r["Com plano de manejo (Sim)"]),
                        "comConselho": num(r["Com conselho gestor (Sim)"]), "comAmbos": num(r["Com ambos (PM + CG)"])}
                  for uf, r in ucs.items()})

    iivcm = rows_by_uf(wb["1.3.1_IIVCM"])
    attach(catalog, "I1.3.1", "2025",
           valores={uf: num(r["Média IIVCM dos prioritários"]) for uf, r in iivcm.items()},
           extra={uf: {"municipiosPrioritarios": num(r["Municípios prioritários (nº)"]),
                        "metaConvergencia": num(r["Meta de convergência (ref.)"])}
                  for uf, r in iivcm.items()})

    prodes = rows_by_uf(wb["1.3.2_Desmatamento_PRODES"])
    anos_prodes = ["2020", "2021", "2022", "2023", "2024"]
    attach(catalog, "I1.3.2", "2024",
           valores={uf: num(r["2024"]) for uf, r in prodes.items()},
           serie={uf: {a: num(r[a]) for a in anos_prodes} for uf, r in prodes.items()},
           extra={uf: {"media2020_2024": num(r["Média 2020-2024"]), "totalAcumulado2008_2024": num(r["Total acumulado 2008-2024"])}
                  for uf, r in prodes.items()})

    focos = rows_by_uf(wb["1.3.4_Focos_calor"])
    anos_focos = [str(a) for a in range(2015, 2025)]
    attach(catalog, "I1.3.4", "média 2015-2024",
           valores={uf: num(r["Média 2015-2024 (baseline)"]) for uf, r in focos.items()},
           serie={uf: {a: num(r[a]) for a in anos_focos} for uf, r in focos.items()})

    return [rec for rec in catalog.values()]


def build_eixo2():
    wb, catalog = load_catalog(ENTREGAVEIS / "Indicadores_Resultado_Eixo2_Amazonia2050.xlsx")

    pobreza = rows_by_uf(wb["2.1.1_Pobreza_SIS"])
    attach(catalog, "I2.1.1", "2024 (SIS/PNADc)",
           valores={uf: num(r["% pobreza (<US$ 3,65)"]) for uf, r in pobreza.items()},
           extra={uf: {"populacaoMil": num(r["População (mil)"]), "pctExtrema215": num(r["% extrema pobreza (<US$ 2,15)"]),
                        "pct685": num(r["% <US$ 6,85"]), "pct50mediana": num(r["% até 50% da mediana"]),
                        "linha50medianaRs": num(r["Linha 50% mediana (R$/mês)"])}
                  for uf, r in pobreza.items()})

    mortalidade = rows_by_uf(wb["2.2.1_Mortalidade_evitavel_ref"])
    attach(catalog, "I2.2.1", "2024 (SIM)",
           valores={uf: num(r["Taxa /1.000 nascidos (ref.)"]) for uf, r in mortalidade.items()},
           extra={uf: {"obitosEvitaveisMenor5": num(r["Óbitos evitáveis <5 anos (2024)"]), "pop0a4": num(r["Pop. 0-4 (2024)"])}
                  for uf, r in mortalidade.items()})

    equipes = rows_by_uf(wb["2.2.2_Equipes_APS_ref"])
    attach(catalog, "I2.2.2", "jul/2026 (CNES)",
           valores={uf: num(r["Total eSF"]) for uf, r in equipes.items()},
           extra={uf: {"eap": num(r["eAP (tipo 76)"]), "populacao2026": num(r["Pop. 2026 (IBGE)"])}
                  for uf, r in equipes.items()})

    telessaude = rows_by_uf(wb["2.2.3_Telessaude_CNES"])
    attach(catalog, "I2.2.3", "jul/2026 (CNES)",
           valores={uf: num(r["Estabelecimentos TELESSAUDE ativos (Jul/2026)"]) for uf, r in telessaude.items()},
           extra={uf: {"participacaoAlPct": num(r["Participação na AL (%)"])} for uf, r in telessaude.items()})

    freq = rows_by_uf(wb["2.3.2_Freq_escolar_SIS"])
    attach(catalog, "I2.3.2", "2024 (SIS/PNADc)",
           valores={uf: num(r["4-17 (média ponderada)"]) for uf, r in freq.items()},
           extra={uf: {"creche0a3": num(r["0-3 anos (creche)"]), "faixa4a5": num(r["4-5 anos"]),
                        "faixa6a10": num(r["6-10 anos"]), "faixa11a14": num(r["11-14 anos"]),
                        "faixa15a17": num(r["15-17 anos"])}
                  for uf, r in freq.items()})

    cvli = rows_by_uf(wb["2.4.1_CVLI_Sinesp"])
    anos_cvli = [str(a) for a in range(2020, 2026)]
    attach(catalog, "I2.4.1", "2025",
           valores={uf: num(r["Taxa CVLI 2025 /100 mil"]) for uf, r in cvli.items()},
           serie={uf: {a: num(r[a]) for a in anos_cvli} for uf, r in cvli.items()},
           extra={uf: {"populacao2025": num(r["Pop. 2025 (IBGE)"])} for uf, r in cvli.items()})

    return [rec for rec in catalog.values()]


def build_eixo3():
    wb, catalog = load_catalog(ENTREGAVEIS / "Indicadores_Resultado_Eixo3_Amazonia2050.xlsx")

    pevs = rows_by_uf(wb["3.1.1_Pevs_serie"])
    anos_pevs = [str(a) for a in range(2015, 2025)]
    attach(catalog, "I3.1.1", "2024",
           valores={uf: num(r["2024"]) for uf, r in pevs.items()},
           serie={uf: {a: num(r[a]) for a in anos_pevs} for uf, r in pevs.items()},
           extra={uf: {"madeireirosMil2024": num(r["Madeireiros 2024 (R$ mil)"]),
                        "naoMadeireirosMil2024": num(r["Não-madeireiros 2024 (R$ mil)"])}
                  for uf, r in pevs.items()})

    rais = rows_by_uf(wb["F3.2_Rais_empregos"])
    attach(catalog, "F3.2", "2024",
           valores={uf: num(r["Vínculos ativos 31/12/2024"]) for uf, r in rais.items()},
           serie={uf: {"2023": num(r["Vínculos ativos 31/12/2023"]), "2024": num(r["Vínculos ativos 31/12/2024"])}
                  for uf, r in rais.items()},
           extra={uf: {"estabelecimentos2023": num(r["Estab. ativos 2023"]), "estabelecimentos2024": num(r["Estab. ativos 2024"]),
                        "variacaoVinculosPct": num(r["Variação vínculos (%)"])}
                  for uf, r in rais.items()})

    pia = rows_by_uf(wb["F3.5_Pia_transformacao"])
    anos_pia = [str(a) for a in range(2015, 2025)]
    attach(catalog, "F3.5", "2024",
           valores={uf: num(r["2024"]) for uf, r in pia.items()},
           serie={uf: {a: num(r[a]) for a in anos_pia} for uf, r in pia.items()},
           extra={uf: {"pessoalOcupado2024": num(r["Pessoal ocupado 2024"]), "empresas2024": num(r["Empresas 2024"]),
                        "unidadesLocais2024": num(r["Unid. locais 2024"]), "variacao2023_24Pct": num(r["Var. 2023→24 (%)"])}
                  for uf, r in pia.items()})

    return [rec for rec in catalog.values()]


def build_eixo4():
    wb, catalog = load_catalog(ENTREGAVEIS / "Indicadores_Resultado_Eixo4_Amazonia2050.xlsx")

    ibc = rows_by_uf(wb["4.1.1_IBC_AMZ_Anatel"])
    anos_ibc = [str(a) for a in range(2021, 2026)]
    attach(catalog, "I4.1.1", "2025",
           valores={uf: num(r["IBC ponderado 2025"]) for uf, r in ibc.items()},
           serie={uf: {a: num(r[f"IBC {a}"]) for a in anos_ibc} for uf, r in ibc.items()},
           extra={uf: {"ponderado2024": num(r["IBC ponderado 2024"])} for uf, r in ibc.items()})

    per = rows_by_uf(wb["4.3.2_PER_renovaveis_aneel"])
    attach(catalog, "I4.3.2", "base 25/08/2026",
           valores={uf: num(r["PER (% renovável)"]) for uf, r in per.items()},
           extra={uf: {"potenciaFiscalizadaMW": num(r["Potência fiscalizada total (MW)"]),
                        "potenciaRenovavelMW": num(r["Potência renovável (MW)"])}
                  for uf, r in per.items()})

    isgr = rows_by_uf(wb["4.4.1_ISGR_saneamento"])
    attach(catalog, "I4.4.1", "Censo 2022 + MUNIC 2024",
           valores={uf: num(r["ISGR (%)"]) for uf, r in isgr.items()},
           extra={uf: {"pctAguaAdequada": num(r["% água adequada"]), "pctEsgotoAdequado": num(r["% esgoto adequado"]),
                        "domiciliosOcupados": num(r["Domicílios ocupados (Censo 2022)"])}
                  for uf, r in isgr.items()})

    return [rec for rec in catalog.values()]


def build_eixo5():
    wb, catalog = load_catalog(ENTREGAVEIS / "Indicadores_Resultado_Eixo5_Amazonia2050.xlsx")

    pnd_rows = list(wb["5.4.1_PnD_MCTI_IBGE"].iter_rows(values_only=True))
    header = pnd_rows[0]
    anos_pd = [str(a) for a in range(2000, 2025)]
    pnd = {}
    pct_pib_2023 = {}
    block = "serie"
    for row in pnd_rows[1:]:
        if row[0] == "UF":
            block = "pct_pib"
            continue
        if block == "serie":
            uf = NAME_TO_UF.get(str(row[1]).strip()) if row[1] else None
            if uf:
                pnd[uf] = dict(zip(header, row))
        else:
            # bloco % do PIB: UF(sigla) | 2019 | 2020 | 2021 | 2022 | 2023 (sem coluna Estado)
            uf = row[0] if row[0] in NAME_TO_UF.values() else None
            if uf:
                pct_pib_2023[uf] = num(row[5])
    attach(catalog, "I5.4.1", "2024",
           valores={uf: num(r["2024"]) for uf, r in pnd.items()},
           serie={uf: {a: num(r[a]) for a in anos_pd} for uf, r in pnd.items()},
           extra={uf: {"pctPib2023": pct_pib_2023.get(uf)} for uf in pnd})

    capag = rows_by_uf(wb["5.5.1_CAPAG_STN"])
    anos_capag = [str(a) for a in range(2018, 2026)]
    attach(catalog, "I5.5.1", "2025",
           valores={uf: str(r["2025"]).strip() if r["2025"] else None for uf, r in capag.items()},
           serie={uf: {a: (str(r[a]).strip() if r[a] else None) for a in anos_capag} for uf, r in capag.items()},
           extra={uf: {"indicadoresAB2025": num(r["Nº indicadores com nota A/B (2025)"])} for uf, r in capag.items()})

    return [rec for rec in catalog.values()]


def main():
    payload = {
        "eixos": [
            {"numero": 1, "nome": "Território, Ambiente e Clima", "indicadores": build_eixo1()},
            {"numero": 2, "nome": "Pessoas e Bem-estar", "indicadores": build_eixo2()},
            {"numero": 3, "nome": "Desenvolvimento econômico sustentável", "indicadores": build_eixo3()},
            {"numero": 4, "nome": "Infraestrutura e integração regional sustentável", "indicadores": build_eixo4()},
            {"numero": 5, "nome": "Governança e parcerias", "indicadores": build_eixo5()},
        ]
    }
    out_path = OUT_DIR / "indicadores.json"
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    coletados = sum(1 for e in payload["eixos"] for i in e["indicadores"] if i["valores"])
    total = sum(len(e["indicadores"]) for e in payload["eixos"])
    print(f"Gravado {out_path} — {coletados}/{total} indicadores com valores por UF.")
    for eixo in payload["eixos"]:
        com_valores = sum(1 for i in eixo["indicadores"] if i["valores"])
        print(f"  Eixo {eixo['numero']}: {com_valores}/{len(eixo['indicadores'])} com valores.")


if __name__ == "__main__":
    main()
