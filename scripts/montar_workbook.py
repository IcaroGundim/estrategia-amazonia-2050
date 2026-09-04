#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Monta o workbook 'Indicadores_Resultado_Eixo1_Amazonia2050.xlsx' com os
indicadores de RESULTADO do Eixo 1 (Estratégia Amazônia 2050) organizados
para os 9 estados da Amazônia Legal."""
import csv, io, os, re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
DADOS = os.path.join(PASTA, "dados")
os.makedirs(DADOS, exist_ok=True)

UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
NOMES_UF = {"AC": "Acre", "AP": "Amapá", "AM": "Amazonas", "MA": "Maranhão",
            "MT": "Mato Grosso", "PA": "Pará", "RO": "Rondônia", "RR": "Roraima", "TO": "Tocantins"}

# ----------------------------------------------------------------------------
# 1. PRODES: taxas de desmatamento por UF/ano
# ----------------------------------------------------------------------------
prodes = defaultdict(dict)
with open(os.path.join(DADOS, "prodes", "prodes_rates_uf.csv")) as f:
    for row in csv.DictReader(f):
        prodes[row["uf"]][int(row["ano"])] = float(row["taxa_km2"])

# ----------------------------------------------------------------------------
# 2. Focos de calor por UF/ano
# ----------------------------------------------------------------------------
focos = defaultdict(dict)
with open(os.path.join(DADOS, "focos", "focos_calor_uf_ano.csv")) as f:
    for row in csv.DictReader(f):
        focos[row["uf"]][int(row["ano"])] = int(row["focos_sat_ref"])

# ----------------------------------------------------------------------------
# 3. CNUC: UCs estaduais na AL
# ----------------------------------------------------------------------------
raw = open(os.path.join(DADOS, "cnuc", "cnuc.csv"), "rb").read()
try:
    text = raw.decode("latin-1")
except Exception:
    text = raw.decode("utf-8", errors="replace")
ucs = list(csv.DictReader(io.StringIO(text), delimiter=";"))

def ufs_da_uc(r):
    """UFs cobertas pela UC (pode ser multivalorada, ex.: 'AM, MT')."""
    return [u.strip().upper() for u in str(r.get("UF", "")).split(",") if u.strip()]

ucs_est_al = [r for r in ucs if r.get("Esfera Administrativa", "").strip().lower() == "estadual"
              and any(u in UFS for u in ufs_da_uc(r))]
uc_uf = defaultdict(list)
for r in ucs_est_al:
    for u in ufs_da_uc(r):
        if u in UFS:
            uc_uf[u].append(r)

def sim(v):
    return str(v).strip().lower() in ("sim", "1", "formalizado", "atualizado")

# ----------------------------------------------------------------------------
# 4. IIVCM (AdaptaBrasil) por município
# ----------------------------------------------------------------------------
iivcm = []
with open(os.path.join(DADOS, "iivcm", "iivcm.csv"), encoding="utf-8-sig") as f:
    for row in csv.DictReader(f):
        try:
            v = float(str(row.get("adaptabrasil_iivcm", "")).replace(",", ".").strip() or "nan")
        except ValueError:
            v = float("nan")
        row["_val"] = v
        iivcm.append(row)

# ----------------------------------------------------------------------------
# Catálogo de indicadores de resultado do Eixo 1
# ----------------------------------------------------------------------------
CATALOGO = [
    ("I1.1.1", "1.1", "ZEE vigente e atualizado", "100% do território dos estados da AL com ZEE vigente, atualizado (revisão nos últimos 10 anos) e incorporado ao planejamento/licenciamento",
     "Percentual do território do Estado com ZEE aprovado por lei/decreto vigente e atualizado nos últimos 10 anos / uso formal em licenciamento (sim/não) / nº de políticas públicas estaduais que citam o ZEE",
     "% / Sim-Não", "Legislação/portal das secretarias estaduais de meio ambiente", 2035, "coleta manual"),
    ("I1.1.2", "1.1", "UCs estaduais com plano de manejo e conselho gestor", "40% de UCs estaduais da AL com plano de manejo elaborado/atualizado e conselho gestor em funcionamento",
     "% de UCs estaduais com plano de manejo e conselho gestor ativos",
     "%", "CNUC/MMA (dados.mma.gov.br)", 2035, "coletado"),
    ("I1.1.3", "1.1", "Cooperação técnica PNGATI/PNGTAQ", "Instrumento de cooperação técnica com órgão federal (PNGATI/PNGTAQ) nos estados, mediante demanda das comunidades",
     "Nº de estados da AL com instrumento de cooperação técnica vigente",
     "nº (0-9)", "Estados / Diário Oficial", 2030, "coleta manual"),
    ("I1.2.1", "1.2", "Planos Estaduais de Adaptação aprovados e em execução", "100% dos estados com PEA aprovados e em execução",
     "100% de estados com planos aprovados e com execução orçamentária iniciada (LOA estadual)",
     "% / Sim-Não", "Diário Oficial / portal da secretaria estadual de meio ambiente", 2030, "coleta manual"),
    ("I1.2.2", "1.2", "PEAs alinhados ao Plano Clima/ENA", "PEAs alinhados ao Plano Clima Nacional e com previsão de revisão periódica",
     "Nº de planos setoriais/temáticos da ENA (de 16) referenciados no PEA / cláusula formal de revisão periódica (sim ou não)",
     "nº / Sim-Não", "Texto do PEA (Diário Oficial/portal), cotejado com Plano Clima/ENA (MMA/Casa Civil)", 2028, "coleta manual"),
    ("I1.3.1", "1.3", "Vulnerabilidade climática municipal (IIVCM)", "Redução de 12,2% da vulnerabilidade climática de municípios prioritários da AL (convergência: média 52,1 pts do grupo menos vulnerável, baseline 2025)",
     "Avaliação periódica do Índice de Intensidade da Vulnerabilidade Climática Municipal (IIVCM)",
     "pontos (0-100)", "AdaptaBrasil/MCTI (lista IIVCM da ficha técnica)", 2050, "coletado"),
    ("I1.3.2", "1.3", "Desmatamento ilegal", "Desmatamento ilegal zero em todos os estados",
     "% de desmatamento ilegal = (ha desmatamento − ha autorizações de supressão)",
     "% / ha", "INPE (PRODES) + IBAMA (Sinaflor)", 2030, "parcial (PRODES ok; Sinaflor exige login)"),
    ("I1.3.3", "1.3", "Monitoramento por sensoriamento remoto", "Ampliar % do território estadual monitorado por sensoriamento remoto com alerta acionável e reduzir tempo de resposta",
     "% do território estadual monitorado com geração de alerta acionável / tempo médio alerta→fiscalização",
     "% / dias", "Estados", 2027, "coleta manual"),
    ("I1.3.4", "1.3", "Focos de calor", "Redução dos focos de incêndio em 30% em relação à baseline 2015-2025",
     "Nº de focos de calor",
     "nº", "INPE Queimadas (data.inpe.br/queimadas)", 2035, "coletado"),
    ("I1.3.5", "1.3", "Delegacias especializadas em crimes ambientais", "Estruturação de delegacias especializadas em crimes ambientais em todos os 9 estados",
     "Nº de delegacias especializadas e fluviais em operação; frota operacional por estado",
     "nº", "FBSP (Cartografias da Violência na Amazônia)", 2027, "coleta manual"),
    ("I1.3.6", "1.3", "Sistema regional de crimes ambientais", "Sistema regional integrado de registro/monitoramento de ocorrências com protocolos padronizados",
     "Nº de estados reportando ao sistema com protocolo padronizado",
     "nº (0-9)", "CAL (parceria com Igarapé)", 2030, "coleta manual"),
    ("I1.3.7", "1.3", "Plano de Manejo Integrado do Fogo", "100% dos estados com Plano de Manejo Integrado do Fogo elaborado (2027) e vigente (revisão a cada 4 anos até 2050)",
     "Nº de estados com Plano registrado no Sisfogo",
     "nº (0-9)", "Sisfogo (IBAMA)", 2030, "coleta manual"),
    ("I1.3.8", "1.3", "PPCDQ vigente", "100% dos estados com PPCDQ (ou equivalente) vigente, revisão nos últimos 4 anos, continuamente até 2050",
     "Nº de estados com PPCDQ vigente (revisão concluída nos últimos 4 anos)",
     "nº (0-9)", "Diário Oficial / portal da secretaria estadual de meio ambiente", 2027, "coleta manual"),
    ("I1.4.1", "1.4", "PRA regulamentado e em execução", "100% dos estados com PRA regulamentado e em execução, com TCAs firmados e monitorados",
     "Estado tem PRA regulamentado? (Sim/Não) / Hectares sob Termo de Compromisso de restauração no âmbito do PRA",
     "Sim-Não / ha", "Observatório do Código Florestal / CAR-SICAR / órgão ambiental estadual", 2030, "coleta manual"),
    ("I1.4.2", "1.4", "Restauração com SAFs", "40% da área restaurada utilizando Sistemas Agroflorestais",
     "Percentual da área restaurada implantada com SAFs",
     "%", "Registros de TCA/PRADA do PRA de cada estado", 2035, "coleta manual"),
    ("I1.5.1", "1.5", "Regularização de áreas públicas estaduais", "100% das áreas públicas prioritárias de domínio estadual destinadas (conservação + PIQCTAF), em especial áreas sob conflito",
     "Percentual da área pública prioritária de domínio estadual regularizada para conservação; Percentual regularizada para PIQCTAF",
     "%", "Órgãos fundiários estaduais", 2035, "coleta manual"),
    ("I1.5.2", "1.5", "Câmaras Técnicas de Destinação", "Câmaras Técnicas de Destinação estaduais constituídas e em operação com participação social nos 9 estados",
     "Nº de Câmaras Técnicas de Destinação estaduais constituídas e operantes",
     "nº (0-9)", "Órgãos fundiários estaduais", 2030, "coleta manual"),
    ("I1.5.3", "1.5", "Mediação de conflitos fundiários", "Institucionalizar espaços de governança de mediação e conciliação de conflitos de PCTs nos 9 estados",
     "Nº de resoluções administrativas de conflitos; % redução da judicialização; aumento da taxa de titulação de PCTs",
     "nº / %", "Órgãos fundiários estaduais", 2030, "coleta manual"),
    ("I1.5.4", "1.5", "Adesão ao SICARF Federativo", "Órgãos estaduais de terras utilizando sistemas (ex.: Terras do Brasil/SICARF-Federativo) na regularização",
     "% de adesão ao SICARF Federativo; aumento da taxa de titulação anual",
     "%", "Órgãos fundiários estaduais", 2030, "coleta manual"),
    ("I1.5.5", "1.5", "Destinação de florestas públicas estaduais", "Destinação de 100% das florestas públicas de domínio estadual não destinadas",
     "% de incremento de florestas públicas de domínio estadual destinadas",
     "%", "Órgãos fundiários estaduais", 2030, "coleta manual"),
    ("I1.5.6", "1.5", "Integração das bases fundiárias/ambientais", "Integração e interoperabilidade das bases fundiárias, ambientais e registrais (Decreto de Governança da Terra / ADPF 743)",
     "% de integração das bases com o uso do SICARF-Federativo",
     "%", "Órgãos fundiários estaduais", 2035, "coleta manual"),
    ("I1.5.7", "1.5", "Sobreposições de CAR", "Resolução de 100% das sobreposições de CAR para desbloqueio de validações",
     "Redução do número de sobreposições de CAR nas bases estaduais",
     "nº", "Órgãos fundiários estaduais", 2035, "coleta manual"),
    ("I1.5.8", "1.5", "CAR de povos e comunidades tradicionais", "Elaboração, análise e validação de 100% dos registros de CAR de PIQCTAFs",
     "Nº de registros de CAR/PCT elaborados e analisados por estado por ano",
     "nº", "SICAR / órgãos ambientais estaduais / INCRA", 2030, "coleta manual"),
]

# ----------------------------------------------------------------------------
# Estilo
# ----------------------------------------------------------------------------
HDR_FILL = PatternFill("solid", fgColor="1B4332")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13, color="1B4332")
NOTE_FONT = Font(italic=True, size=9, color="555555")
BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")

def estilo_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER

def largura(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# =========================== Aba Sobre =====================================
ws = wb.active
ws.title = "Sobre"
ws["A1"] = "ESTRATÉGIA REGIONAL AMAZÔNIA 2050 — Eixo 1: Gestão territorial, sustentabilidade ambiental e governança climática"
ws["A1"].font = TITLE_FONT
ws["A3"] = "Indicadores de RESULTADO (somente resultado — sem indicadores de impacto), organizados para os 9 estados da Amazônia Legal."
ws["A4"] = "Estados: AC (Acre), AP (Amapá), AM (Amazonas), MA (Maranhão), MT (Mato Grosso), PA (Pará), RO (Rondônia), RR (Roraima), TO (Tocantins)."
ws["A5"] = "Base: '3. Matriz Metas x Indicadores .xlsx' (aba 'Metas x Indicadores (atualizada)', coluna INDICADOR DE RESULTADO) e 'Fichas Técnicas Indicadores - Amazonia2050.docx'."
ws["A7"] = "Data de coleta dos dados: 18/08/2026."
ws["A8"] = "Status: coletado = dados obtidos de fonte oficial (aba com detalhes); parcial = apenas parte do indicador obtida; coleta manual = indicador depende de levantamento documental/estadual (ZEE, PEA, PPCDQ, Sisfogo, delegacias, fundiário etc.), sem API pública — fontes indicadas no catálogo."
ws["A10"] = "FONTES DE DADOS AUTOMATIZADAS"
ws["A10"].font = Font(bold=True)
ws["A11"] = "• Desmatamento (taxas anuais por UF): PRODES/INPE via TerraBrasilis (dados até 2024)."
ws["A12"] = "• Focos de calor (satélite de referência, 2015-2024): INPE Queimadas — data.inpe.br/queimadas/dados-abertos (2025 ainda não publicado)."
ws["A13"] = "• UCs estaduais (plano de manejo e conselho gestor): CNUC/MMA — dados.mma.gov.br (base de 03/2026)."
ws["A14"] = "• IIVCM por município: lista oficial citada na ficha técnica (AdaptaBrasil/MCTI, baseline 2025)."
ws["A15"] = "• Sinaflor (autorizações de supressão) exige login (SSO IBAMA) — % de desmatamento ilegal ficou parcial; alternativa: solicitar ao IBAMA ou usar painel público quando disponível."
ws["A16"] = "• Desmatamento: série de taxas anuais do TerraBrasilis (rates2025.json, INPE). Ano da taxa = ano final do período de medição (ago/jul). 2020-2023 conferem com a série oficial divulgada; 2024 (6.518 km²) reflete a revisão do ciclo 2025; 2025 = 5.731 km² (mais recente)."
ws["A17"] = "• CNUC: UCs estaduais que abrangem a Amazônia Legal; UCs interestaduais (ex.: 'AM, MT') contam para cada estado abrangido (pode somar >100% na agregação por estado)."
ws["A19"] = "Scripts de coleta: pasta 'scripts/'. Dados brutos: pasta 'dados/'."
ws["A19"].font = NOTE_FONT
largura(ws, [130])

# ====================== Aba Catalogo_Indicadores ============================
ws = wb.create_sheet("Catalogo_Indicadores")
headers = ["Código", "Linha de Ação", "Indicador (nome curto)", "Meta", "Indicador de RESULTADO (texto da matriz)",
           "Unidade", "Fonte", "Prazo", "Status da coleta"]
ws.append(headers)
for row in CATALOGO:
    ws.append(list(row))
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = WRAP
largura(ws, [8, 9, 34, 42, 52, 12, 34, 9, 16])

# ====================== Aba Matriz_UF =======================================
ws = wb.create_sheet("Matriz_UF")
headers = ["Código", "Indicador de RESULTADO (resumo)", "Unidade", "Ano ref.", "AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO", "Status", "Fonte"]
ws.append(headers)

def val_uf(uf):
    return None

rows_uf = []
# I1.1.2 UCs
uc_dados = {}
for uf in UFS:
    lst = uc_uf.get(uf, [])
    n = len(lst)
    pm = sum(1 for r in lst if sim(r.get("Plano de Manejo")))
    cg = sum(1 for r in lst if sim(r.get("Conselho Gestor")))
    ambos = sum(1 for r in lst if sim(r.get("Plano de Manejo")) and sim(r.get("Conselho Gestor")))
    pct = (ambos / n * 100) if n else None
    uc_dados[uf] = (n, pm, cg, ambos, pct)

# I1.3.1 IIVCM por UF (média dos municípios prioritários)
iivcm_uf = {}
for uf in UFS:
    vals = [r["_val"] for r in iivcm if r.get("uf", "").strip().upper() == uf and r.get("prioritario", "").strip().lower() == "sim"]
    vals = [v for v in vals if v == v]
    iivcm_uf[uf] = (sum(vals) / len(vals)) if vals else None

# I1.3.2 desmatamento PRODES (último ano 2024 + média 2020-2024)
prodes_uf = {}
for uf in UFS:
    d = prodes.get(uf, {})
    anos = sorted(d)
    recentes = [a for a in anos if a >= 2020]
    med = sum(d[a] for a in recentes) / len(recentes) if recentes else None
    prodes_uf[uf] = (d.get(2024), med)

# I1.3.4 focos (média 2015-2024 = baseline)
focos_uf = {}
for uf in UFS:
    d = focos.get(uf, {})
    anos = sorted(d)
    med = sum(d[a] for a in anos) / len(anos) if anos else None
    focos_uf[uf] = (d.get(max(anos)) if anos else None, med, len(anos))

M = []
for cod, la, curto, meta, indicador, unid, fonte, prazo, status in CATALOGO:
    if cod == "I1.1.2":
        vals = [round(uc_dados[uf][4], 1) if uc_dados[uf][4] is not None else "—" for uf in UFS]
        anoref, status_ = "03/2026", "coletado"
    elif cod == "I1.3.1":
        vals = [round(iivcm_uf[uf], 1) if iivcm_uf[uf] is not None else "—" for uf in UFS]
        anoref, status_ = "2025", "coletado"
    elif cod == "I1.3.2":
        vals = [round(prodes_uf[uf][0], 1) if prodes_uf[uf][0] is not None else "—" for uf in UFS]
        anoref, status_ = "2024 (PRODES)", "parcial (falta Sinaflor)"
        unid = "km²/ano (PRODES)"
    elif cod == "I1.3.4":
        vals = [f"{int(focos_uf[uf][1])}" if focos_uf[uf][1] is not None else "—" for uf in UFS]
        anoref, status_ = "média 2015-2024", "coletado"
    else:
        vals = ["—"] * 9
        anoref, status_ = "—", status
    M.append([cod, curto, unid, anoref] + vals + [status_, fonte])

for row in M:
    ws.append(row)
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = WRAP
largura(ws, [9, 44, 12, 12, 8, 8, 8, 8, 8, 8, 8, 8, 8, 20, 26])
ws.freeze_panes = "E2"

# ====================== Aba 1.1.2 UCs =======================================
ws = wb.create_sheet("1.1.2_UCs_estaduais")
headers = ["UF", "Estado", "UCs estaduais (nº)", "Com plano de manejo (Sim)", "Com conselho gestor (Sim)",
           "Com ambos (PM + CG)", "% com ambos (meta 40%)"]
ws.append(headers)
for uf in UFS:
    n, pm, cg, ambos, pct = uc_dados[uf]
    ws.append([uf, NOMES_UF[uf], n, pm, cg, ambos, round(pct, 1) if pct is not None else "—"])
# totais: UCs UNICAS na AL (cada UC conta 1x, inclusive interestaduais)
tot_n = len(ucs_est_al)
tot_ambos = sum(1 for r in ucs_est_al if sim(r.get("Plano de Manejo")) and sim(r.get("Conselho Gestor")))
ws.append(["TOTAL AL", "Amazônia Legal (UCs únicas)", tot_n,
           sum(1 for r in ucs_est_al if sim(r.get("Plano de Manejo"))),
           sum(1 for r in ucs_est_al if sim(r.get("Conselho Gestor"))),
           tot_ambos, round(tot_ambos / tot_n * 100, 1)])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
        ws.cell(row=r, column=c).alignment = WRAP
largura(ws, [6, 16, 14, 22, 22, 18, 20])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Nota: UCs de esfera administrativa ESTADUAL que abrangem a Amazônia Legal, base CNUC 03/2026 (dados.mma.gov.br). UCs interestaduais contam para cada estado abrangido (colunas por UF podem somar > total de UCs únicas). 'Plano de Manejo' e 'Conselho Gestor' = colunas da base (Sim/Não).").font = NOTE_FONT

# ====================== Aba 1.3.1 IIVCM =====================================
ws = wb.create_sheet("1.3.1_IIVCM")
headers = ["UF", "Estado", "Municípios prioritários (nº)", "Média IIVCM dos prioritários", "Meta de convergência (ref.)"]
ws.append(headers)
for uf in UFS:
    n = sum(1 for r in iivcm if r.get("uf", "").strip().upper() == uf and r.get("prioritario", "").strip().lower() == "sim")
    ws.append([uf, NOMES_UF[uf], n, round(iivcm_uf[uf], 2) if iivcm_uf[uf] is not None else "—", 52.1])
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 24, 26, 22])
# lista de municípios
ws2 = wb.create_sheet("IIVCM_municipios")
headers = ["Código IBGE", "Município", "UF", "Prioritário", "IIVCM"]
ws2.append(headers)
for r in sorted(iivcm, key=lambda x: (x.get("uf", ""), x.get("nome_municipio", ""))):
    ws2.append([r.get("codigo_ibge"), r.get("nome_municipio"), r.get("uf"), r.get("prioritario"),
                str(r.get("adaptabrasil_iivcm"))])
estilo_header(ws2, len(headers))
for r in range(2, ws2.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws2.cell(row=r, column=c).border = BORDER
largura(ws2, [12, 34, 6, 10, 10])
ws2.freeze_panes = "A2"

# ====================== Aba 1.3.2 Desmatamento ==============================
ws = wb.create_sheet("1.3.2_Desmatamento_PRODES")
headers = ["UF", "Estado", "2020", "2021", "2022", "2023", "2024", "Média 2020-2024", "Total acumulado 2008-2024"]
ws.append(headers)
for uf in UFS:
    d = prodes.get(uf, {})
    acum = sum(v for k, v in d.items() if 2008 <= k <= 2024)
    row = [uf, NOMES_UF[uf]]
    for a in (2020, 2021, 2022, 2023, 2024):
        row.append(round(d.get(a, 0), 1))
    recentes = [d[a] for a in (2020, 2021, 2022, 2023, 2024) if a in d]
    row.append(round(sum(recentes) / len(recentes), 1) if recentes else "—")
    row.append(round(acum, 1))
    ws.append(row)
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16, 10, 10, 10, 10, 10, 14, 22])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: PRODES/INPE via TerraBrasilis (rates2025.json), taxa anual em km²/ano, ano = fim do período de medição (ago/jul). 2024 = 6.518 km² na AL (série revisada do ciclo 2025). O indicador oficial da matriz é % de desmatamento ILEGAL = (desmatamento − autorizações Sinaflor); a componente Sinaflor exige login (SSO IBAMA) e não pôde ser automatizada — ver 'Sobre'.").font = NOTE_FONT

# ====================== Aba 1.3.4 Focos =====================================
ws = wb.create_sheet("1.3.4_Focos_calor")
headers = ["UF", "Estado"] + [str(a) for a in range(2015, 2025)] + ["Média 2015-2024 (baseline)", "Último ano com dado"]
ws.append(headers)
for uf in UFS:
    d = focos.get(uf, {})
    row = [uf, NOMES_UF[uf]]
    for a in range(2015, 2025):
        row.append(d.get(a, "—"))
    anos = sorted(d)
    med = sum(d[a] for a in anos) / len(anos) if anos else "—"
    row.append(int(med) if isinstance(med, float) else med)
    row.append(max(anos) if anos else "—")
    ws.append(row)
estilo_header(ws, len(headers))
for r in range(2, ws.max_row + 1):
    for c in range(1, len(headers) + 1):
        ws.cell(row=r, column=c).border = BORDER
largura(ws, [6, 16] + [9] * 10 + [22, 16])
ws.append([])
ws.cell(row=ws.max_row, column=1, value="Fonte: INPE Queimadas, satélite de referência (série homogênea), arquivos anuais por estado (focos_br_<uf>_ref_<ano>.zip). 2025 ainda não publicado. Meta: redução de 30% frente à baseline 2015-2025.").font = NOTE_FONT

# ====================== Aba 1.5.8 CAR (contexto) ============================
ws = wb.create_sheet("1.5.8_CAR_anotacoes")
ws["A1"] = "Indicador: Nº de registros de CAR/PCT elaborados e analisados por estado por ano (SICAR / órgãos ambientais / INCRA)"
ws["A1"].font = TITLE_FONT
ws["A3"] = "O SICAR público (consultapublica.car.gov.br) exige sessão interativa para consultas; não há API aberta de contagens de CAR por PCT/situação no momento da coleta (18/08/2026)."
ws["A4"] = "Para completar este indicador: (1) solicitar extração ao INCRA/SICAR ou às secretarias estaduais de meio ambiente; (2) ou usar o módulo de download do SICAR por município (manual, via navegador)."
ws["A6"] = "Itens relacionados (I1.5.4, I1.5.6, I1.5.7): adesão ao SICARF-Federativo, integração de bases e sobreposições de CAR também dependem de dados dos órgãos fundiários estaduais."
largura(ws, [120])

# Salvar
os.makedirs(os.path.join(PASTA, "entregaveis"), exist_ok=True)
out = os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo1_Amazonia2050.xlsx")
wb.save(out)
print("Workbook salvo:", out)

# Resumo rápido para conferência
print("\n== RESUMO ==")
print("UCs estaduais AL:", tot_n, "| com PM+CG:", tot_ambos, f"({tot_ambos/tot_n*100:.1f}%)")
for uf in UFS:
    iiv = iivcm_uf[uf]
    iiv_s = f"{iiv:.2f}" if iiv is not None else "—"
    print(f"{uf}: UCs={uc_dados[uf][0]} PM+CG={uc_dados[uf][3]} | IIVCM_med={iiv_s} | PRODES_2024={prodes_uf[uf][0]} km² | focos_med={int(focos_uf[uf][1]) if focos_uf[uf][1] else 0}")
