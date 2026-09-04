#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Extrai população projetada por UF/ano (IBGE 2024) e calcula taxas CVLI e
cobertura APS aproximada (CNES equipes) para os 9 estados da AL."""
import csv, os
from collections import defaultdict
from openpyxl import load_workbook

DADOS = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "dados")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]

# ---------- população por UF/ano ----------
wb = load_workbook(os.path.join(DADOS, "ibge_pop", "projecoes_2024_idade_simples.xlsx"), read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
it = ws.iter_rows(values_only=True)
for _ in range(5):
    next(it)  # pula cabeçalhos
header = next(it)
anos = [str(c) for c in header[5:]]
pop = defaultdict(lambda: defaultdict(float))
for row in it:
    if row[0] is None or row[3] is None:
        continue
    if str(row[1]).strip() != "Ambos":  # linha 'Ambos' já é o total (M+F)
        continue
    sigla = str(row[3]).strip().upper()
    if sigla not in UFS:
        continue
    for i, a in enumerate(anos):
        v = row[5 + i]
        if isinstance(v, (int, float)):
            pop[sigla][a] += float(v)
wb.close()
with open(os.path.join(DADOS, "ibge_pop", "populacao_uf_ano.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["uf", "ano", "populacao"])
    for uf in UFS:
        for a in sorted(pop[uf]):
            w.writerow([uf, a, int(pop[uf][a])])
print("população salva. Ex.:")
for uf in UFS:
    print(f"  {uf} 2025 = {int(pop[uf]['2025']):,}")

# ---------- CVLI por 100 mil (2025) ----------
cvli = defaultdict(dict)
with open(os.path.join(DADOS, "sinesp", "cvli_uf_ano.csv")) as f:
    for r in csv.DictReader(f):
        cvli[r["uf"]][int(r["ano"])] = int(r["cvli"])
print("\nTaxa CVLI 2025 (por 100 mil hab.):")
for uf in UFS:
    taxa = cvli[uf].get(2025, 0) / pop[uf]["2025"] * 100000
    print(f"  {uf}: {cvli[uf].get(2025, 0)} -> {taxa:.1f}")

# ---------- equipes de saúde (CNES Jul/2026) ----------
# colunas de interesse no CSV de equipes
with open(os.path.join(DADOS, "cnes", "cnes_equipes_uf.csv"), encoding="latin-1") as f:
    linhas = [l.strip() for l in f if l.strip()]
header = [c.strip().strip('"') for c in linhas[3].split(";")]
print("\ncolunas equipes (parcial):")
for i, c in enumerate(header):
    if "ESF" in c.upper() or "EAP" in c.upper() or "ESB" in c.upper() or "ACS" in c.upper():
        print(f"  [{i}] {c}")
