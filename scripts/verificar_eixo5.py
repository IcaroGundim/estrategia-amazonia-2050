# -*- coding: utf-8 -*-
"""Verificação independente: brutos → CSVs agregados → workbook (Eixo 5)."""
import csv, json, os, unicodedata, urllib.request

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
DADOS = os.path.join(PASTA, "dados", "eixo5")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
NOMES = {"AC": "Acre", "AP": "Amapá", "AM": "Amazonas", "MA": "Maranhão",
         "MT": "Mato Grosso", "PA": "Pará", "RO": "Rondônia", "RR": "Roraima",
         "TO": "Tocantins"}

def norm(s):
    return unicodedata.normalize("NFD", s).encode("ascii", "ignore").decode().strip().lower()

def num(s):
    s = (s or "").strip().replace(" ", "")
    if s in ("", "-", "—", "…"):
        return None
    try:
        return float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return None

erros = []
def check(cond, msg):
    print(("OK  " if cond else "FALHA") + " | " + msg)
    if not cond:
        erros.append(msg)

# ============ 1. CAPAG ============
# re-parse independente dos brutos (heurística: última coluna de classificação da linha)
import re
bruto_capag = {}  # (uf, ano) -> letra
ARQ = {2018: "capag_2018.csv", 2019: "capag_2019.csv", 2020: "capag_2020.csv",
       2021: "capag_2021.csv", 2022: "capag_2022.csv", 2023: "capag_2023.csv",
       2024: "capag_2024_revisao.csv", 2025: "capag_2025.csv"}
PAT_CLASS = re.compile(r"^[ABCD][+*]?$|^Suspensa$", re.I)
for ano, nome in ARQ.items():
    for raw in open(os.path.join(DADOS, "capag", nome), encoding="utf-8-sig"):
        p = [x.strip() for x in raw.rstrip("\n").split(";")]
        if p and p[0] in UFS:
            cands = [x for x in p[1:] if PAT_CLASS.match(x)]
            bruto_capag[(p[0], ano)] = cands[-1] if cands else ""

agg = {}
for r in csv.DictReader(open(os.path.join(DADOS, "capag_uf_ano.csv"), encoding="utf-8")):
    agg[(r["uf"], int(r["ano"]))] = r["capag"]
check(len(agg) == 72, f"capag_uf_ano.csv tem 72 linhas (72x? tem {len(agg)})")
diffs = [(k, bruto_capag.get(k), agg.get(k)) for k in sorted(agg) if bruto_capag.get(k) != agg.get(k)]
check(not diffs, f"CAPAG agregado == bruto para todas as 72 células {diffs[:3] if diffs else ''}")

# ============ 2. P&D MCTI ============
bruto_pd = {}   # (uf, ano) -> R$ mi
arq = os.path.join(DADOS, "pd", "tab_01_02_02_05_e_2024.csv")
cab = None
for raw in open(arq, encoding="latin-1"):
    p = [x.strip() for x in raw.rstrip("\n").split(";")]
    if any(h.endswith("P&D") for h in p):
        cab = p
        break
col = {int(h.split()[0]): i for i, h in enumerate(cab) if h.endswith("P&D")}
n2uf = {norm(n): u for u, n in NOMES.items()}
for raw in open(arq, encoding="latin-1"):
    p = [x.strip() for x in raw.rstrip("\n").split(";")]
    uf = n2uf.get(norm(p[0]))
    if uf:
        for a, i in col.items():
            v = num(p[i]) if i < len(p) else None
            if v is not None:
                bruto_pd[(uf, a)] = v

agg_pd = {}
for r in csv.DictReader(open(os.path.join(DADOS, "pd_uf_ano.csv"), encoding="utf-8")):
    if r["pd_mi"]:
        agg_pd[(r["uf"], int(r["ano"]))] = float(r["pd_mi"])
diffs = [(k, bruto_pd.get(k), agg_pd.get(k)) for k in set(bruto_pd) | set(agg_pd)
         if bruto_pd.get(k) != agg_pd.get(k)]
check(not diffs, f"P&D (R$ mi) agregado == bruto para todas as células {diffs[:3] if diffs else ''}")
check(len(bruto_pd) == len(agg_pd), f"P&D nº de células: bruto {len(bruto_pd)} == agregado {len(agg_pd)}")

# %/receita
bruto_rec = {}
arq = os.path.join(DADOS, "pd", "tab_01_02_02_07_e_2024.csv")
col_ano = {}
for raw in open(arq, encoding="latin-1"):
    p = [x.strip() for x in raw.rstrip("\n").split(";")]
    if not col_ano:
        col_ano = {int(h): i for i, h in enumerate(p) if h.isdigit()}
        continue
    uf = n2uf.get(norm(p[0]))
    if uf:
        for a, i in col_ano.items():
            if i < len(p):
                v = num(p[i])
                if v is not None:
                    bruto_rec[(uf, a)] = v
agg_rec = {}
for r in csv.DictReader(open(os.path.join(DADOS, "pd_uf_ano.csv"), encoding="utf-8")):
    if r["pct_receita"]:
        agg_rec[(r["uf"], int(r["ano"]))] = float(r["pct_receita"])
diffs = [(k, bruto_rec.get(k), agg_rec.get(k)) for k in set(bruto_rec) | set(agg_rec)
         if bruto_rec.get(k) != agg_rec.get(k)]
check(not diffs, f"%P&D/receita agregado == bruto {diffs[:3] if diffs else ''}")

# ============ 3. PIB: re-fetch SIDRA e comparar ============
url = "https://apisidra.ibge.gov.br/values/t/5938/n3/all/v/37/p/2019-2023"
req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
with urllib.request.urlopen(req, timeout=60) as r:
    d = json.loads(r.read().decode())
bruto_pib = {}
for row in d[1:]:
    uf = next((u for u, n in NOMES.items() if row["D1N"].upper() == n.upper()), None)
    if uf and row["V"].replace(".", "").isdigit():
        bruto_pib[(uf, int(row["D3C"]))] = float(row["V"]) / 1000.0
agg_pib = {}
for r in csv.DictReader(open(os.path.join(DADOS, "pd_uf_ano.csv"), encoding="utf-8")):
    if r["pib_mi"]:
        agg_pib[(r["uf"], int(r["ano"]))] = float(r["pib_mi"])
diffs = [(k, bruto_pib.get(k), agg_pib.get(k)) for k in set(bruto_pib) | set(agg_pib)
         if abs((bruto_pib.get(k) or 0) - (agg_pib.get(k) or 0)) > 0.01]
check(not diffs, f"PIB SIDRA (re-fetch) == agregado {diffs[:3] if diffs else ''}")

# % PIB recalculado independente
for uf in UFS:
    for ano in (2022, 2023):
        pd_ = agg_pd.get((uf, ano))
        pib = agg_pib.get((uf, ano))
        if pd_ and pib:
            esperado = round(pd_ / pib * 100, 4)

# ============ 4. Workbook ============
import openpyxl
wb = openpyxl.load_workbook(os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo5_Amazonia2050.xlsx"))
ws = wb["5.5.1_CAPAG_STN"]
linhas = {r[0]: r for r in ws.iter_rows(min_row=2, values_only=True) if r[0] in UFS}
ok = all(linhas[uf][2 + i] == bruto_capag[(uf, ano)] for uf in UFS for i, ano in enumerate(range(2018, 2026)))
check(ok, "workbook CAPAG == bruto (todas as 72 células)")

ws2 = wb["5.4.1_PnD_MCTI_IBGE"]
linhas2 = {}
for r in ws2.iter_rows(values_only=True):  # primeira ocorrência = bloco P&D (R$ mi)
    if r[0] in UFS and r[0] not in linhas2:
        linhas2[r[0]] = r
anos = list(range(2000, 2025))
ok = True
for uf in UFS:
    for i, a in enumerate(anos):
        wbv = linhas2[uf][2 + i]
        csvv = agg_pd.get((uf, a))
        wbv = float(wbv) if wbv not in (None, "") else None
        if (wbv is None) != (csvv is None) or (wbv is not None and abs(wbv - csvv) > 1e-9):
            ok = False
check(ok, "workbook P&D == CSV (todas as células)")

# Matriz_UF: CAPAG 2025 e P&D/PIB 2023
ws3 = wb["Matriz_UF"]
m = {r[0]: r for r in ws3.iter_rows(min_row=2, values_only=True)}
cap25_ok = all(m["I5.5.1"][4 + i] == agg[("x" if False else uf, 2025)] for i, uf in enumerate(UFS))
check(cap25_ok, "Matriz_UF CAPAG 2025 == CSV")
pd23_ok = True
for i, uf in enumerate(UFS):
    v = m["I5.4.1"][4 + i]
    esperado = agg_pd.get((uf, 2023)) and round(agg_pd[(uf, 2023)] / agg_pib[(uf, 2023)] * 100, 3)
    if (v == "—") != (esperado is None):
        pd23_ok = False
    elif esperado is not None and abs(float(v) - esperado) > 0.0006:
        pd23_ok = False
check(pd23_ok, "Matriz_UF P&D/PIB 2023 == recálculo independente")

print()
print("TOTAL DE FALHAS:", len(erros))
