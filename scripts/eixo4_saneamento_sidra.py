#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Eixo 4.4a — Índice de Saneamento Básico e Gestão de Riscos (ISGR).

Censo 2022 (SIDRA 6803 água / 6805 esgoto, municípios) + fatores da MUNIC 2024
(FClima = Magr18/Mhab088/Mtic266; FGov = Mgov086), conforme a ficha técnica.

  agua_adeq_m   = 72144 (rede geral e usa) + 72145 (rede geral, usa outra) + 72154 (poço profundo)
  esgoto_adeq_m = 46290 (rede geral/pluvial/fossa ligada à rede) + 72112 (fossa séptica não ligada)
  DOM_efetivo_m = min(agua_adeq, esgoto_adeq) × FClima × FGov     [FHidro não publicado => 1]
  ISGR_UF       = Σ DOM_efetivo / Σ DOM_total × 100

Saídas:
  dados/saneamento/saneamento_mun.csv
  dados/saneamento/isgr_uf.csv
"""
import csv, json, os, unicodedata, urllib.request

import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
DADOS = os.path.join(PASTA, "dados")
SAN = os.path.join(DADOS, "saneamento")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
URL_6803 = "https://apisidra.ibge.gov.br/values/t/6803/n6/all/v/381/p/2022/c1821/72129,72144,72145,72154?formato=json"
URL_6805 = "https://apisidra.ibge.gov.br/values/t/6805/n6/all/v/381/p/2022/c11558/46292,46290,72112?formato=json"
MUNIC = os.path.join(DADOS, "ibge_munic", "Base_MUNIC_2024_20251107.xlsx")


def baixar_json(url):
    print("SIDRA:", url[:80], "...")
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=600) as r:
        return json.loads(r.read().decode("utf-8"))


def sem_acento(s):
    return unicodedata.normalize("NFD", str(s)).encode("ascii", "ignore").decode()


# ---------- 1. Municípios da Amazônia Legal (códigos do IIVCM) ----------
cod_al = {}
with open(os.path.join(DADOS, "iivcm", "iivcm.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        cod_al[r["codigo_ibge"]] = r["uf"]
print("municípios da AL:", len(cod_al))

# ---------- 2. SIDRA 6803 (água) e 6805 (esgoto) ----------
agua = {}   # cod -> dict cat -> domicílios
for x in baixar_json(URL_6803)[1:]:
    cod = x["D1C"]
    if cod in cod_al and str(x["V"]).strip() not in ("-", ""):
        agua.setdefault(cod, {})[x["D4C"]] = int(x["V"])
esgoto = {}
for x in baixar_json(URL_6805)[1:]:
    cod = x["D1C"]
    if cod in cod_al and str(x["V"]).strip() not in ("-", ""):
        esgoto.setdefault(cod, {})[x["D4C"]] = int(x["V"])
print("municípios com dado de água:", len(agua), "| esgoto:", len(esgoto))

# ---------- 3. MUNIC 2024: fatores ----------
COLS = {"Informática e comunicação": "Mtic266", "Governanca": "Mgov086",
        "Habitacao": "Mhab088", "Agropecuária": "Magr18"}
munic = {}  # cod -> {var: True/False}
wb = openpyxl.load_workbook(MUNIC, read_only=True)
for aba, alvo in COLS.items():
    ws = wb[aba]
    idx = None
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        row = [str(c).strip() if c is not None else "" for c in row]
        idx = row.index(alvo)
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = str(row[0]).strip()
        if cod not in cod_al:
            continue
        val = str(row[idx]).strip().lower() if idx < len(row) and row[idx] is not None else ""
        munic.setdefault(cod, {})[alvo] = val == "sim"
wb.close()
print("municípios com MUNIC:", len(munic))

# ---------- 4. Cálculo municipal + agregação estadual ----------
fator_faltante = set()
mun_out = []
agreg = {uf: {"dom_total": 0, "agua": 0, "esgoto": 0, "efetivo": 0.0} for uf in UFS}
cods = sorted(cod_al)
for cod in cods:
    uf = cod_al[cod]
    a = agua.get(cod, {})
    e = esgoto.get(cod, {})
    dom_total = a.get("72129", 0) or e.get("46292", 0)
    agua_adeq = a.get("72144", 0) + a.get("72145", 0) + a.get("72154", 0)
    esgoto_adeq = e.get("46290", 0) + e.get("72112", 0)
    m = munic.get(cod, {})
    if set(m) != set(COLS.values()):
        fator_faltante.add(cod)
    fclima = 1.0 if any(m.get(v, False) for v in ("Magr18", "Mhab088", "Mtic266")) else 0.9
    fgov = 1.0 if m.get("Mgov086", False) else 0.95
    efetivo = min(agua_adeq, esgoto_adeq) * fclima * fgov
    mun_out.append([cod, uf, dom_total, agua_adeq, esgoto_adeq, round(fclima, 2), round(fgov, 2), round(efetivo, 1)])
    g = agreg[uf]
    g["dom_total"] += dom_total
    g["agua"] += agua_adeq
    g["esgoto"] += esgoto_adeq
    g["efetivo"] += efetivo

os.makedirs(SAN, exist_ok=True)
with open(os.path.join(SAN, "saneamento_mun.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["cod_ibge", "uf", "dom_total", "agua_adeq", "esgoto_adeq", "fclima", "fgov", "dom_efetivo"])
    w.writerows(mun_out)

with open(os.path.join(SAN, "isgr_uf.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.writer(f)
    w.writerow(["uf", "dom_total", "agua_adeq", "esgoto_adeq", "dom_efetivo", "isgr_pct",
                "pct_agua_adeq", "pct_esgoto_adeq"])
    tot = {"dom_total": 0, "agua": 0, "esgoto": 0, "efetivo": 0.0}
    for uf in UFS:
        g = agreg[uf]
        isgr = g["efetivo"] / g["dom_total"] * 100 if g["dom_total"] else 0
        w.writerow([uf, g["dom_total"], g["agua"], g["esgoto"], round(g["efetivo"], 1),
                    round(isgr, 2), round(g["agua"] / g["dom_total"] * 100, 1),
                    round(g["esgoto"] / g["dom_total"] * 100, 1)])
        for k in tot:
            tot[k] += g[k]
    isgr_al = tot["efetivo"] / tot["dom_total"] * 100
    w.writerow(["AL", tot["dom_total"], tot["agua"], tot["esgoto"], round(tot["efetivo"], 1),
                round(isgr_al, 2), round(tot["agua"] / tot["dom_total"] * 100, 1),
                round(tot["esgoto"] / tot["dom_total"] * 100, 1)])

print("fator faltante (sem MUNIC):", len(fator_faltante))
print("\nUF  ISGR%  %água  %esgoto")
for uf in UFS:
    g = agreg[uf]
    print(uf, f"{g['efetivo']/g['dom_total']*100:6.2f}", f"{g['agua']/g['dom_total']*100:6.1f}",
          f"{g['esgoto']/g['dom_total']*100:6.1f}")
print("AL", f"{tot['efetivo']/tot['dom_total']*100:6.2f}", f"{tot['agua']/tot['dom_total']*100:6.1f}",
      f"{tot['esgoto']/tot['dom_total']*100:6.1f}")
print("baseline da ficha: 41,52% (2022/2024) — calculado com FHidro=1 (fator não publicado)")
