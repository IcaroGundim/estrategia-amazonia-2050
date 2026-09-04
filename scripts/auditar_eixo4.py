#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Auditoria de sanidade do Eixo 4 — verificações cruzadas contra as fontes."""
import csv, os
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
DADOS = os.path.join(PASTA, "dados")
ENT = os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo4_Amazonia2050.xlsx")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
ok = True


def check(nome, cond, detalhe=""):
    global ok
    print(("[OK]  " if cond else "[FALHA] ") + nome + (" — " + detalhe if detalhe else ""))
    ok = ok and cond


# ---------- IBC ----------
def num(s):
    return float(str(s).replace(",", ".")) if str(s).strip() not in ("", "-") else None

ibc_uf_bruto = {}
for r in csv.DictReader(open(os.path.join(DADOS, "anatel", "IBC_UF_indicadores_originais.csv"), encoding="utf-8-sig"), delimiter=";"):
    if r["Ano"] == "2025" and r["UF"] in UFS:
        ibc_uf_bruto[r["UF"]] = num(r["IBC"])
proc = {r["uf"]: (float(r["ibc_uf"]), float(r["ibc_ponderado_pop"]) if r["ibc_ponderado_pop"] else None)
        for r in csv.DictReader(open(os.path.join(DADOS, "anatel", "ibc_uf_ano.csv"), encoding="utf-8")) if r["ano"] == "2025"}
check("IBC 2025 por UF bate com o arquivo bruto da ANATEL",
      all(abs(proc[uf][0] - ibc_uf_bruto[uf]) < 0.01 for uf in UFS),
      "ex.: AC " + str(ibc_uf_bruto["AC"]))
al2025 = float(csv.DictReader(open(os.path.join(DADOS, "anatel", "ibc_al_ano.csv"), encoding="utf-8")).__next__ if False else
                [r for r in csv.DictReader(open(os.path.join(DADOS, "anatel", "ibc_al_ano.csv"), encoding="utf-8")) if r["ano"] == "2025"][0]["ibc_al"])
check("AL ponderado 2025 = 53,52 (baseline da ficha)", abs(al2025 - 53.52) < 0.01, f"obtido {al2025}")
check("IBC ponderado dentro do intervalo municipal (10-80)",
      all(10 <= proc[uf][1] <= 80 for uf in UFS), "todos entre 10 e 80")

# ---------- SIGA / PER ----------
per = {r["uf"]: r for r in csv.DictReader(open(os.path.join(DADOS, "aneel", "per_uf.csv"), encoding="utf-8"))}
pot = {}
for r in csv.DictReader(open(os.path.join(DADOS, "aneel", "potencia_uf_fonte.csv"), encoding="utf-8")):
    pot.setdefault(r["uf"], {})[r["origem"]] = float(r["potencia_fiscalizada_kw"])
consist = all(abs(sum(pot[uf].values()) - float(per[uf]["potencia_total_kw"])) < 1 for uf in UFS)
check("potencia_uf_fonte soma igual a per_uf (consistência interna)", consist)
t_al = sum(float(per[uf]["potencia_total_kw"]) for uf in UFS)
r_al = sum(float(per[uf]["potencia_renovavel_kw"]) for uf in UFS)
per_al = r_al / t_al * 100
check("PER da AL entre 60% e 90% (plausível p/ proxy sobre potência)", 60 <= per_al <= 90, f"{per_al:.1f}%")
check("PA é o maior parque (Tucuruí + Belo Monte)", max(per, key=lambda u: float(per[u]["potencia_total_kw"])) == "PA",
      f"{float(per['PA']['potencia_total_kw'])/1e6:.1f} GW")
check("sem potência fóssil no AC zerada (AC deve ter baixo PER)", float(per["AC"]["per_renovaveis_pct"]) < 10,
      f"AC {per['AC']['per_renovaveis_pct']}% renovável")

# ---------- Saneamento ----------
mun = list(csv.DictReader(open(os.path.join(DADOS, "saneamento", "saneamento_mun.csv"), encoding="utf-8")))
check("808 municípios da AL com dado municipal", len(mun) == 808, f"{len(mun)}")
bad = [m for m in mun if float(m["dom_efetivo"]) > min(float(m["agua_adeq"]), float(m["esgoto_adeq"])) + 0.5]
check("DOM_efetivo ≤ min(água, esgoto) em todos os municípios", not bad)
neg = [m for m in mun if float(m["dom_total"]) <= 0]
check("nenhum município com domicílios ≤ 0", not neg)
isgr = {r["uf"]: r for r in csv.DictReader(open(os.path.join(DADOS, "saneamento", "isgr_uf.csv"), encoding="utf-8"))}
isgr_al = float(isgr["AL"]["isgr_pct"])
check("ISGR AL entre 40% e 55% (com FHidro=1, acima do baseline 41,52%)", 40 <= isgr_al <= 55, f"{isgr_al:.2f}%")
check("ISGR RR é o maior (Boa Vista: rede urbana)", max(UFS, key=lambda u: float(isgr[u]["isgr_pct"])) == "RR",
      f"RR {isgr['RR']['isgr_pct']}%")
soma_mun = sum(float(m["dom_total"]) for m in mun)
check("Σ dom_total municipal = total AL na aba UF", abs(soma_mun - float(isgr["AL"]["dom_total"])) < 10,
      f"{soma_mun:,.0f} domicílios")

# ---------- Workbook ----------
wb = openpyxl.load_workbook(ENT)
esperadas = {"Sobre", "Catalogo_Indicadores", "Matriz_UF", "4.1.1_IBC_AMZ_Anatel",
             "4.3.2_PER_renovaveis_aneel", "4.3.2_PER_detalhe_fonte", "4.4.1_ISGR_saneamento",
             "Pendentes_anotacoes"}
check("workbook com as 8 abas esperadas", set(wb.sheetnames) == esperadas, str(wb.sheetnames))
ws = wb["Catalogo_Indicadores"]
check("catálogo com 6 indicadores do Eixo 4", ws.max_row == 7, f"{ws.max_row - 1} linhas")
ws = wb["Matriz_UF"]
coletados = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, 5).value != "—")
check("Matriz_UF com 3 indicadores preenchidos (4.1.1, 4.3.2, 4.4.1)", coletados == 3, f"{coletados} preenchidos")

print("\nRESULTADO:", "AUDITORIA OK" if ok else "HÁ FALHAS — revisar")
