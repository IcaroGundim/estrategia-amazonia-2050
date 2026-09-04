#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Auditoria dos dados coletados para o Eixo 1 (Amazônia 2050)."""
import csv, io, os, zipfile, glob
from collections import defaultdict, Counter

PASTA = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DADOS = os.path.join(PASTA, "dados")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]

print("=" * 70)
print("AUDITORIA 1: FOCOS DE CALOR (INPE Queimadas)")
print("=" * 70)
# 1a. conferir o conteudo real de um zip (header + amostra)
z = os.path.join(DADOS, "focos", "focos_br_pa_ref_2024.zip")
with zipfile.ZipFile(z) as zf:
    names = zf.namelist()
    print("arquivos no zip:", names)
    with zf.open([n for n in names if n.endswith(".csv")][0]) as f:
        head = f.read(1200).decode("latin-1")
        print("header+amostra:")
        print(head[:700])
        # conta linhas
        f.seek(0)
        n = sum(1 for _ in f) - 1
        print("linhas de dados no zip PA 2024:", n)

# 1b. conferir se todos os UFs tem 10 anos (2015-2024)
print("\n-- cobertura por UF --")
with open(os.path.join(DADOS, "focos", "focos_calor_uf_ano.csv")) as f:
    rows = list(csv.DictReader(f))
cov = defaultdict(list)
for r in rows:
    cov[r["uf"]].append(int(r["ano"]))
for uf in UFS:
    anos = sorted(cov[uf])
    ok = anos == list(range(2015, 2025))
    print(f"{uf}: {len(anos)} anos ({anos[0]}-{anos[-1]}) completo={ok}")
tot = sum(int(r["focos_sat_ref"]) for r in rows)
print("TOTAL focos 2015-2024 (9 UFs):", tot)

# 1c. sanidade: PA deve ter mais focos que RR; valores positivos
for r in rows:
    if int(r["focos_sat_ref"]) <= 0:
        print("ALERTA valor <=0:", r)

print()
print("=" * 70)
print("AUDITORIA 2: PRODES (taxas por UF/ano)")
print("=" * 70)
prodes = defaultdict(dict)
with open(os.path.join(DADOS, "prodes", "prodes_rates_uf.csv")) as f:
    for r in csv.DictReader(f):
        prodes[r["uf"]][int(r["ano"])] = float(r["taxa_km2"])
# total AL por ano (soma dos 9 estados)
print("Total Amazônia Legal (soma 9 UFs) vs série oficial INPE:")
oficial = {2020: 10851, 2021: 13038, 2022: 11594, 2023: 9064, 2024: 6288}
for ano in range(2019, 2025):
    soma = sum(prodes[uf].get(ano, 0) for uf in UFS)
    ref = oficial.get(ano)
    diff = f"| oficial={ref} dif={soma-ref:+.0f} ({100*(soma-ref)/ref:+.1f}%)" if ref else "| sem ref oficial"
    print(f"  {ano}: {soma:,.0f} km² {diff}")
# ultimos anos por UF
print("2024 por UF:", {uf: prodes[uf].get(2024) for uf in UFS})

print()
print("=" * 70)
print("AUDITORIA 3: CNUC (UCs estaduais)")
print("=" * 70)
raw = open(os.path.join(DADOS, "cnuc", "cnuc.csv"), "rb").read()
text = raw.decode("latin-1")
ucs = list(csv.DictReader(io.StringIO(text), delimiter=";"))
print("total linhas CNUC:", len(ucs))
print("esferas:", Counter(r["Esfera Administrativa"].strip() for r in ucs).most_common())
print("UFs presentes:", sorted(set(r["UF"].strip() for r in ucs)))
def ufs_da_uc(r):
    return [u.strip().upper() for u in str(r.get("UF", "")).split(",") if u.strip()]

ucs_est_al = [r for r in ucs if r.get("Esfera Administrativa", "").strip().lower() == "estadual"
              and any(u in UFS for u in ufs_da_uc(r))]
print("UCs estaduais que abrangem a AL (únicas):", len(ucs_est_al))
print("por UF (com interestaduais contando em cada UF):", dict(Counter(u for r in ucs_est_al for u in ufs_da_uc(r) if u in UFS)))
print("interestaduais na AL:", [r["Nome da UC"] + " (" + r["UF"] + ")" for r in ucs_est_al if "," in r.get("UF", "")])
print("valores Plano de Manejo:", Counter(r["Plano de Manejo"].strip() for r in ucs_est_al).most_common())
print("valores Conselho Gestor:", Counter(r["Conselho Gestor"].strip() for r in ucs_est_al).most_common())
# exemplo de UC com plano de manejo sim
ex = [r for r in ucs_est_al if r["Plano de Manejo"].strip().lower() == "sim"][:3]
for r in ex:
    print("exemplo PM=Sim:", r["Nome da UC"], "|", r["UF"], "|", r["Categoria de Manejo"])

print()
print("=" * 70)
print("AUDITORIA 4: IIVCM (AdaptaBrasil)")
print("=" * 70)
with open(os.path.join(DADOS, "iivcm", "iivcm.csv"), encoding="utf-8-sig") as f:
    iivcm = list(csv.DictReader(f))
print("total municipios no CSV:", len(iivcm))
print("UFs:", sorted(set(r["uf"].strip().upper() for r in iivcm)))
prio = [r for r in iivcm if r["prioritario"].strip().lower() == "sim"]
print("prioritarios:", len(prio), "| nao prioritarios:", len(iivcm) - len(prio))
print("por UF (prio):", dict(Counter(r["uf"].strip().upper() for r in prio)))
# validar numeros
bad = [r for r in iivcm if not str(r["adaptabrasil_iivcm"]).replace(",", ".").replace(" ", "").replace('"', "")[:4].replace(".", "").isdigit()]
print("linhas com valor nao numerico:", len(bad))
# faixa de valores
vals = [float(r["adaptabrasil_iivcm"].replace(",", ".")) for r in iivcm if r["adaptabrasil_iivcm"].strip()]
print(f"faixa IIVCM: min={min(vals):.1f} max={max(vals):.1f} media={sum(vals)/len(vals):.1f}")
# medias por UF dos prioritarios
for uf in UFS:
    v = [float(r["adaptabrasil_iivcm"].replace(",", ".")) for r in prio if r["uf"].strip().upper() == uf]
    print(f"  {uf}: n={len(v)} media={sum(v)/len(v):.2f}" if v else f"  {uf}: sem dados")

print()
print("=" * 70)
print("AUDITORIA 5: WORKBOOK GERADO")
print("=" * 70)
from openpyxl import load_workbook
wb = load_workbook(os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo1_Amazonia2050.xlsx"))
print("abas:", wb.sheetnames)
ws = wb["Matriz_UF"]
print("\nMatriz_UF (indicador | AC AP AM MA MT PA RO RR TO | status):")
for row in ws.iter_rows(min_row=2, values_only=True):
    print(" ", row[0], "|", row[1][:45], "|", row[4:13], "|", row[13])
ws = wb["Catalogo_Indicadores"]
print("\nCatalogo:", ws.max_row - 1, "indicadores;", "status:", Counter(r[8] for r in ws.iter_rows(min_row=2, values_only=True)))
