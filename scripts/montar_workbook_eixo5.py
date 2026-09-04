#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Monta 'Indicadores_Resultado_Eixo5_Amazonia2050.xlsx' — Eixo 5: Governança
e parcerias (Estratégia Amazônia 2050), 9 estados da Amazônia Legal.

Entradas:
- dados/eixo5/capag_uf_ano.csv  (scripts/eixo5_capag.py  <- CKAN Tesouro Transparente)
- dados/eixo5/pd_uf_ano.csv     (scripts/eixo5_pd.py     <- MCTI + IBGE/SIDRA)
"""
import csv, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.abspath(__file__))
PASTA = os.path.dirname(BASE)
DADOS = os.path.join(PASTA, "dados")
UFS = ["AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO"]
NOMES_UF = {"AC": "Acre", "AP": "Amapá", "AM": "Amazonas", "MA": "Maranhão",
            "MT": "Mato Grosso", "PA": "Pará", "RO": "Rondônia", "RR": "Roraima", "TO": "Tocantins"}

# ---------- CAPAG por UF/ano (STN) ----------
capag = {}  # uf -> ano -> {'capag': str, notas letras e valores}
with open(os.path.join(DADOS, "eixo5", "capag_uf_ano.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        capag.setdefault(r["uf"], {})[int(r["ano"])] = {
            "capag": r["capag"],
            "ind1": r["indicador1"], "ind2": r["indicador2"], "ind3": r["indicador3"],
            "n1": r["nota1"], "n2": r["nota2"], "n3": r["nota3"],
        }
ANOS_CAPAG = sorted({a for d in capag.values() for a in d})

def letra_capag(v):
    """Letra base da classificação (A+ → A; C* → C; 'Suspensa' → None)."""
    v = (v or "").strip().upper()
    return v[0] if v and v[0] in "ABC" else None

def conta_capag(ano, estrito=False):
    """Nº de estados da AL com classificação A ou B no ano."""
    n = 0
    for uf in UFS:
        v = capag.get(uf, {}).get(ano, {}).get("capag", "")
        if estrito:
            n += 1 if v in ("A", "B") else 0
        else:
            n += 1 if letra_capag(v) in ("A", "B") else 0
    return n

# ---------- P&D por UF/ano (MCTI + SIDRA) ----------
pd_data = {}  # uf -> ano -> dict
with open(os.path.join(DADOS, "eixo5", "pd_uf_ano.csv"), encoding="utf-8") as f:
    for r in csv.DictReader(f):
        ano = int(r["ano"])
        d = {"pd_mi": r["pd_mi"], "pct_rec": r["pct_receita"],
             "pib_mi": r["pib_mi"], "pct_pib": r["pct_pib"]}
        if any(d.values()):
            pd_data.setdefault(r["uf"], {})[ano] = d
ANOS_PD = sorted({a for d in pd_data.values() for a in d})

def fv(s):
    return "" if s in (None, "") else s

# ---------- Catálogo Eixo 5 ----------
CATALOGO = [
    ("I5.1.1", "5.1", "Financiamento climático captado/executado",
     "Mobilizar recursos para financiamento climático e ambiental equivalentes a, no mínimo, 0,1% do PIB da AL",
     "Valor anual e acumulado dos recursos captados e efetivamente executados, por estado, origem e instrumento financeiro",
     "R$", "Estados (programas jurisdicionais de crédito de carbono, financiamento climático e fundos multilaterais)",
     2050, "coleta manual (não há base consolidada pública por estado; documentos estaduais e relatórios de fundos)"),
    ("I5.2.1", "5.2", "Taxa de alavancagem (blended finance)",
     "Estruturar pelo menos 2 mecanismos regionais de financiamento misto (blended finance)",
     "Aumento do volume de capital privado mobilizado para cada R$ de recurso público aportado nesses mecanismos",
     "R$ privado / R$ público", "Estados", 2050,
     "coleta manual (mecanismos ainda por estruturar)"),
    ("I5.2.2", "5.2", "Governança regional / recursos executados pelo CAL",
     "Institucionalizar modelo permanente de governança regional coordenado pelo CAL, com participação dos 9 estados",
     "Nº de mecanismos de governança regional instituídos e em operação; volume de recursos (R$) captados e executados de forma centralizada pelo Consórcio da Amazônia Legal",
     "nº / R$", "CAL (consorcioamazonialegal.gov.br/orcamento-anual)", 2050,
     "pendente via automação (site do CAL em Wix — tabela carrega via JS autenticado; coleta manual junto ao CAL. Ref. da ficha técnica: orçamento executado 2026 = R$ 6.120.000; crescimento médio de R$ 483.566/ano no histórico 2019-2026)"),
    ("I5.3.1", "5.3", "Arranjo jurídico climático estadual",
     "Desenvolver e implementar arranjo jurídico-institucional climático regional com diretrizes comuns entre os 9 estados",
     "Número ou percentual de estados que incorporaram as diretrizes em normas ou instrumentos próprios",
     "nº / %", "Estados", 2050, "coleta manual (diretrizes regionais em desenvolvimento)"),
    ("I5.3.2", "5.3", "Interoperabilidade de dados prioritários",
     "Implantar sistema regional de governança e interoperabilidade de dados prioritários, com adesão dos 9 estados",
     "Percentual das bases de dados prioritárias integradas ou interoperáveis",
     "% / bases", "Estados", 2050, "coleta manual"),
    ("I5.4.1", "5.4", "Dispêndio em P&D (% do PIB estadual)",
     "Ampliar o investimento público e privado em P&D nos estados da AL, alcançando o equivalente a 1% do PIB estadual ao ano",
     "Dispêndio público e privado em P&D como percentual do PIB de cada estado",
     "% do PIB", "Estados e CAL; proxy: MCTI (dispêndio estadual em P&D) + IBGE/SIDRA (PIB por UF)", 2050,
     "parcial (componente público estadual coletada — MCTI t.1.2.2.5/1.2.2.7, 2000-2024, e P&D/PIB via SIDRA t5938; dispêndio empresarial/privado exige PINTEC por UF, disponível só para UFs selecionadas)"),
    ("I5.5.1", "5.5", "CAPAG A ou B (capacidade de pagamento)",
     "Fortalecer a capacidade de gestão pública em todos os estados, com foco em transparência e eficiência",
     "Número e percentual de estados com classificação A ou B na CAPAG",
     "grau A-C", "STN — Tesouro Transparente (Capacidade de Pagamento dos Estados)", 2050,
     "coletado (CKAN do Tesouro Transparente, 2018-2025; 2024 = revisão de 04/2025)"),
    ("I5.5.2", "5.5", "Transparência pública (EBT 360)",
     "Atender 90% dos critérios da principal avaliação nacional de transparência pública vigente",
     "Número e percentual de estados que atingem o patamar mínimo estabelecido (Escala Brasil Transparente 360)",
     "0 a 10", "CGU — Escala Brasil Transparente 360 (Mapa Brasil Transparente)", 2050,
     "pendente (sistema MBT/CGU fora do ar para atualização — retorno previsto nov/2026; portal dadosabertos.cgu.gov.br inacessível no dia da coleta)"),
    ("I5.5.3", "5.5", "Modelos regionais de governança compartilhada",
     "Consolidar modelos regionais de governança compartilhada para temas prioritários da AL, com equipe, orçamento e plano de implementação",
     "Número de modelos construídos e em operação com equipe delegada para a sua implementação",
     "modelos", "Estados e CAL", 2050, "coleta manual"),
]

# ---------- Estilo ----------
HDR_FILL = PatternFill("solid", fgColor="1B4332")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=13, color="1B4332")
NOTE_FONT = Font(italic=True, size=9, color="555555")
BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)
WRAP = Alignment(wrap_text=True, vertical="top")

def estilo_header(ws, ncols, row=1):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
        cell.border = BORDER

def estilo_corpo(ws, ncols, primeira=2):
    for r in range(primeira, ws.max_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).border = BORDER
            ws.cell(row=r, column=c).alignment = WRAP

def largura(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def rodape(ws, texto, negrito=False):
    """Anexa uma linha nova com o texto na coluna A (evita sobrescrever a última linha)."""
    ws.append([texto])
    c = ws.cell(row=ws.max_row, column=1)
    if negrito:
        c.font = Font(bold=True)
    else:
        c.font = NOTE_FONT

wb = Workbook()

# ---------- Sobre ----------
ws = wb.active
ws.title = "Sobre"
ws["A1"] = "ESTRATÉGIA REGIONAL AMAZÔNIA 2050 — Eixo 5: Governança e parcerias"
ws["A1"].font = TITLE_FONT
ws["A3"] = "Indicadores de RESULTADO do Eixo 5, organizados para os 9 estados da Amazônia Legal."
ws["A4"] = "Base: '3. Matriz Metas x Indicadores .xlsx' (aba atualizada, coluna INDICADOR DE RESULTADO) e fichas técnicas."
ws["A6"] = "Data de coleta: 30/08/2026."
ws["A7"] = "Status: coletado = dado oficial obtido; parcial = parte do indicador; pendente = fonte indisponível no dia; coleta manual = levantamento documental/estadual."
ws["A9"] = "FONTES E LIMITAÇÕES"
ws["A9"].font = Font(bold=True)
ws["A10"] = "• CAPAG (I5.5.1): CKAN do Tesouro Transparente (STN) — notas anuais 2018-2025 por UF com os 3 indicadores (endividamento, poupança corrente, liquidez). Para 2024 usou-se a revisão publicada em 04/2025. Valores especiais mantidos como publicados: AP 'C*' (2019) e 'Suspensa' (2020). A graduação A+/B+ foi introduzida pela STN nas avaliações recentes; a contagem 'A ou B' da aba 5.5.1 considera a letra base (A+ ou B+ contam como A ou B)."
ws["A11"] = "• P&D (I5.4.1): MCTI, Indicadores Nacionais de C,T&I — t.1.2.2.5 (dispêndios dos governos estaduais em P&D, R$ milhões correntes, 2000-2024) e t.1.2.2.7 (% P&D/receita total). A série cobre o dispêndio PÚBLICO ESTADUAL (incl. IES estaduais); não inclui dispêndio privado nem federal. P&D % PIB calculado com o PIB por UF do IBGE/SIDRA t5938 (disponível até 2023)."
ws["A12"] = "• EBT 360 (I5.5.2): o sistema Mapa Brasil Transparente (CGU) está temporariamente fora do ar ('previsão de retorno: novembro/2026') e o portal dadosabertos.cgu.gov.br não resolveu no dia da coleta — coletar na próxima edição/avaliação disponível."
ws["A13"] = "• CAL (I5.2.2): a tabela 'Orçamento Anual' do site do CAL (Wix) carrega via _api/cloud-data com autenticação de sessão (WDE0117 sem token) — sem download aberto; referência oficial na ficha técnica citada na aba 5.2.2."
ws["A14"] = "• Indicadores I5.1.1, I5.2.1, I5.3.1, I5.3.2 e I5.5.3: fontes da matriz são os próprios estados/CAL — exigem levantamento documental (normas, diários oficiais, relatórios)."
ws["A16"] = "Scripts: scripts/eixo5_capag.py, scripts/eixo5_pd.py, scripts/montar_workbook_eixo5.py. Dados brutos: dados/eixo5/."
ws["A16"].font = NOTE_FONT
largura(ws, [135])

# ---------- Catálogo ----------
ws = wb.create_sheet("Catalogo_Indicadores")
headers = ["Código", "Linha de Ação", "Indicador (nome curto)", "Meta", "Indicador de RESULTADO (texto da matriz)",
           "Unidade", "Fonte", "Prazo", "Status da coleta"]
ws.append(headers)
for row in CATALOGO:
    ws.append(list(row))
estilo_header(ws, len(headers))
estilo_corpo(ws, len(headers))
largura(ws, [9, 9, 30, 42, 55, 13, 34, 8, 48])
ws.freeze_panes = "A2"

# ---------- Matriz_UF ----------
ws = wb.create_sheet("Matriz_UF")
headers = ["Código", "Indicador de RESULTADO (resumo)", "Unidade", "Ano ref.", "AC", "AP", "AM", "MA", "MT", "PA", "RO", "RR", "TO", "AL (agregado)", "Status", "Fonte"]
ws.append(headers)
for cod, la, curto, meta, indicador, unid, fonte, prazo, status in CATALOGO:
    if cod == "I5.5.1":
        vals = [capag.get(uf, {}).get(2025, {}).get("capag", "—") for uf in UFS]
        n_ab = conta_capag(2025)
        agreg, anoref, status_ = f"{n_ab}/9 em A ou B", "2025 (STN)", "coletado (CAPAG 2018-2025)"
    elif cod == "I5.4.1":
        vals = []
        for uf in UFS:
            d = pd_data.get(uf, {}).get(2023, {})
            v = d.get("pct_pib")
            vals.append(round(float(v), 3) if v else "—")
        tot_pd = sum(float(pd_data.get(uf, {}).get(2023, {}).get("pd_mi") or 0) for uf in UFS)
        tot_pib = sum(float(pd_data.get(uf, {}).get(2023, {}).get("pib_mi") or 0) for uf in UFS)
        agreg = f"{tot_pd / tot_pib * 100:.3f}% (P&D público/PIB)"
        anoref, status_ = "2023 (MCTI + SIDRA)", "parcial (público estadual; privado exige PINTEC)"
    else:
        vals = ["—"] * 9
        agreg, anoref, status_ = "—", "—", status
    ws.append([cod, curto, unid, anoref] + vals + [agreg, status_, fonte])
estilo_header(ws, len(headers))
estilo_corpo(ws, len(headers))
largura(ws, [9, 42, 14, 17, 8, 8, 8, 8, 8, 8, 8, 8, 8, 20, 30, 30])
ws.freeze_panes = "E2"

# ---------- 5.5.1 CAPAG ----------
ws = wb.create_sheet("5.5.1_CAPAG_STN")
headers = ["UF", "Estado"] + [str(a) for a in ANOS_CAPAG] + ["Nº indicadores com nota A/B (2025)"]
ws.append(headers)
for uf in UFS:
    row = [uf, NOMES_UF[uf]]
    for a in ANOS_CAPAG:
        row.append(capag.get(uf, {}).get(a, {}).get("capag", "—"))
    d25 = capag.get(uf, {}).get(2025, {})
    n = sum(1 for k in ("n1", "n2", "n3") if letra_capag(d25.get(k)) in ("A", "B"))
    row.append(n)
    ws.append(row)
row = ["AL", "Amazônia Legal (9 UF)"]
for a in ANOS_CAPAG:
    row.append(f"{conta_capag(a)}/9")
ws.append(row)
estilo_header(ws, len(headers))
estilo_corpo(ws, len(headers))
largura(ws, [6, 22] + [9] * len(ANOS_CAPAG) + [26])

rodape(ws, "Contagem anual de estados da AL com CAPAG cuja letra base é A ou B (A+ e B+ contam como A e B):", negrito=True)
ws.append(["Ano"] + [str(a) for a in ANOS_CAPAG])
ws.append(["A ou B (nº UF)"] + [conta_capag(a) for a in ANOS_CAPAG])
ws.append(["% da AL"] + [round(conta_capag(a) / 9 * 100, 1) for a in ANOS_CAPAG])
ws.append(["A/B estritos (nº UF)"] + [conta_capag(a, estrito=True) for a in ANOS_CAPAG])
estilo_corpo(ws, len(ANOS_CAPAG) + 1, primeira=ws.max_row - 3)
rodape(ws, "Fonte: STN — Tesouro Transparente, 'Capacidade de Pagamento dos Estados e do Distrito Federal' (CKAN; capag-estados). CAPAG = combinação das notas de endividamento, poupança corrente e liquidez. Escala: A+ > A > B+ > B > C (A+/B+ aplicados a partir da avaliação 2024). 2024 = revisão de 04/2025. AP: 'C*' em 2019 (nota com ressalva) e 'Suspensa' em 2020, conforme publicado.")

# ---------- 5.4.1 P&D ----------
ws = wb.create_sheet("5.4.1_PnD_MCTI_IBGE")
headers = ["UF", "Estado"] + [str(a) for a in ANOS_PD]
ws.append(headers)
for uf in UFS:
    row = [uf, NOMES_UF[uf]]
    for a in ANOS_PD:
        v = pd_data.get(uf, {}).get(a, {}).get("pd_mi")
        row.append(fv(v))
    ws.append(row)
row = ["AL", "Amazônia Legal (9 UF)"]
for a in ANOS_PD:
    tot = sum(float(pd_data.get(uf, {}).get(a, {}).get("pd_mi") or 0) for uf in UFS)
    row.append(round(tot, 2))
ws.append(row)
estilo_header(ws, len(headers))
estilo_corpo(ws, len(headers))
largura(ws, [6, 22] + [10] * len(ANOS_PD))
rodape(ws, "Dispêndios dos governos estaduais em P&D, em R$ milhões correntes (MCTI, t.1.2.2.5, 2000-2024; cédula vazia = não publicado/nulo na fonte; a linha AL soma apenas os valores publicados).")

# % P&D/PIB (anos com PIB publicado)
anoss_pib = sorted({a for uf in UFS for a, d in pd_data.get(uf, {}).items() if d.get("pib_mi")})
rodape(ws, "P&D estadual como % do PIB da UF (P&D ÷ PIB corrente, SIDRA t5938 — IBGE):", negrito=True)
ws.append(["UF"] + [str(a) for a in anoss_pib])
for uf in UFS:
    row = [uf]
    for a in anoss_pib:
        v = pd_data.get(uf, {}).get(a, {}).get("pct_pib")
        row.append(fv(v))
    ws.append(row)
row = ["AL"]
for a in anoss_pib:
    pd_tot = sum(float(pd_data.get(uf, {}).get(a, {}).get("pd_mi") or 0) for uf in UFS)
    pib_tot = sum(float(pd_data.get(uf, {}).get(a, {}).get("pib_mi") or 0) for uf in UFS)
    row.append(round(pd_tot / pib_tot * 100, 4) if pib_tot else "—")
ws.append(row)
estilo_corpo(ws, len(anoss_pib) + 1, primeira=ws.max_row - len(UFS) - 1)
rodape(ws, "Fonte: MCTI — Indicadores Nacionais de C,T&I (t.1.2.2.5 e t.1.2.2.7, dados dos Balanços Gerais dos Estados coletados junto à COREM/STN e secretarias estaduais de C&T); PIB por UF: IBGE/SIDRA t5938 (PIB a preços correntes, mil R$ → R$ mi). Meta da matriz: 1% do PIB (público + privado) — a série aqui é apenas o componente público estadual.")

# ---------- 5.2.2 CAL (referência) ----------
ws = wb.create_sheet("5.2.2_CAL_orcamento")
ws.append(["Referência — Recursos captados/executados centralizadamente pelo Consórcio da Amazônia Legal"])
ws["A1"].font = TITLE_FONT
ws["A3"] = "Tentativa de coleta automatizada (30/08/2026):"
ws["A4"] = "• Site: https://www.consorcioamazonialegal.gov.br/orcamento-anual (página 'Transparência > Orçamento Anual', com filtros Ano/Situação/Instrumento e exportação CSV/JSON/TXT/PDF na interface)."
ws["A5"] = "• A página é Wix (CMS); a tabela carrega via _api/cloud-data com token de sessão — chamada sem token retorna 'WDE0117: MetaSite not found'; sem JSON estático embutido no HTML e sem endpoint dinâmico público localizado (dynamic-pages-router retornou 404)."
ws["A6"] = "• Conclusão: coleta manual junto ao CAL (o próprio consórcio publica os valores) ou Extração fornecida pela Secretaria-Executiva."
ws["A8"] = "Referência oficial da ficha técnica (Metodologia de Projeção do CAL):"
ws["A9"] = "• Orçamento base 2026 (valor executado): R$ 6.120.000,00"
ws["A10"] = "• Taxa média de crescimento anual do histórico 2019-2026: R$ 483.566,00/ano"
ws["A11"] = "• Equação da ficha: Orçamento Projetado = 6.120.000 + 483.566 × (Ano Alvo − 2026)"
ws["A13"] = "Projeção da ficha para marcos da Estratégia:"
ws["A14"] = "2030: R$ " + f"{6120000 + 483566 * 4:,.0f}".replace(",", ".")
ws["A15"] = "2035: R$ " + f"{6120000 + 483566 * 9:,.0f}".replace(",", ".")
ws["A16"] = "2050: R$ " + f"{6120000 + 483566 * 24:,.0f}".replace(",", ".")
ws["A18"] = "Uso exclusivo como referência documentada na ficha técnica do CAL; substituir pela série executada oficial quando obtida."
ws["A18"].font = NOTE_FONT
largura(ws, [130])

caminho = os.path.join(PASTA, "entregaveis", "Indicadores_Resultado_Eixo5_Amazonia2050.xlsx")
wb.save(caminho)
print("OK ->", os.path.relpath(caminho, PASTA))
